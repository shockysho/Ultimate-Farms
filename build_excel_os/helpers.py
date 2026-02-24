"""
helpers.py — Shared utility functions for building the Excel OS workbook.
Table creation, data validation, conditional formatting, protection, named ranges.
"""

from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Protection, numbers
from openpyxl.utils import get_column_letter, quote_sheetname

from . import config as C


def create_excel_table(ws, table_name, headers, data_rows, start_row=1, start_col=1,
                       style_name="TableStyleMedium2", calculated_columns=None,
                       col_widths=None, number_formats=None):
    """
    Create an Excel Table on the given worksheet with data, formatting, and optional calculated columns.

    Args:
        ws: openpyxl Worksheet
        table_name: str — unique table name (no spaces)
        headers: list of str — column headers
        data_rows: list of lists — cell values (None for formula cells)
        start_row: int — first row (1-based)
        start_col: int — first column (1-based)
        style_name: str — Excel table style
        calculated_columns: dict {col_offset: formula_string} — 0-based column index within table
            The formula uses structured refs like "[@Column]-[@Other]" (no leading =)
        col_widths: list of int/float — column widths (same order as headers)
        number_formats: dict {col_offset: format_string} — 0-based

    Returns:
        (Table, end_row) tuple
    """
    num_cols = len(headers)
    num_rows = len(data_rows)

    # Write headers
    for ci, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + ci, value=header)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
        cell.alignment = C.HEADER_ALIGN

    # Write data rows
    for ri, row_data in enumerate(data_rows):
        for ci, value in enumerate(row_data):
            cell = ws.cell(row=start_row + 1 + ri, column=start_col + ci)
            if value is not None:
                cell.value = value

    # Write calculated column formulas into cells
    if calculated_columns:
        for col_offset, formula in calculated_columns.items():
            col_idx = start_col + col_offset
            for ri in range(num_rows):
                cell = ws.cell(row=start_row + 1 + ri, column=col_idx)
                cell.value = f"={formula}"

    # Apply number formats
    if number_formats:
        for col_offset, fmt in number_formats.items():
            col_idx = start_col + col_offset
            for ri in range(num_rows):
                ws.cell(row=start_row + 1 + ri, column=col_idx).number_format = fmt

    # Calculate table range
    end_row = start_row + num_rows
    if num_rows == 0:
        end_row = start_row + 1  # Tables need at least one data row placeholder
        for ci in range(num_cols):
            ws.cell(row=start_row + 1, column=start_col + ci, value="")

    end_col = start_col + num_cols - 1
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

    # Create Table object
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style

    # Manually build TableColumn list (avoids _initialise_columns issues)
    tab.tableColumns = [TableColumn(id=i + 1, name=h) for i, h in enumerate(headers)]

    ws.add_table(tab)

    # Set column widths
    if col_widths:
        for ci, w in enumerate(col_widths):
            if w:
                ws.column_dimensions[get_column_letter(start_col + ci)].width = w

    return tab, end_row


def add_dropdown_validation(ws, cell_range, source_sheet_title, source_range,
                            error_message="Invalid entry. Select from the list."):
    """
    Add a dropdown (list) data validation referencing a range on another sheet.
    """
    formula = f"={quote_sheetname(source_sheet_title)}!{source_range}"
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,  # OOXML: False = show the dropdown arrow
        showErrorMessage=True,
        errorTitle="Invalid Entry",
        error=error_message,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)
    return dv


def add_inline_dropdown(ws, cell_range, values_list, error_message="Select from the list."):
    """
    Add a dropdown using an inline comma-separated list of values.
    Use for short, static lists (< 255 chars total).
    """
    formula = '"' + ','.join(str(v) for v in values_list) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Entry",
        error=error_message,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)
    return dv


def add_integer_validation(ws, cell_range, min_val=0, max_val=999999):
    """Add whole-number validation."""
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=str(min_val),
        formula2=str(max_val),
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Number",
        error=f"Enter a whole number between {min_val} and {max_val}.",
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)
    return dv


def add_date_validation(ws, cell_range):
    """Add date validation (no future dates)."""
    dv = DataValidation(
        type="date",
        operator="lessThanOrEqual",
        formula1="TODAY()",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Date",
        error="Date cannot be in the future.",
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)
    return dv


def create_named_range(wb, name, sheet_title, cell_range):
    """
    Create a workbook-scoped named range.
    cell_range should be absolute, e.g. "$A$2:$A$50"
    """
    ref = f"{quote_sheetname(sheet_title)}!{cell_range}"
    defn = DefinedName(name, attr_text=ref)
    wb.defined_names[name] = defn
    return defn


def add_traffic_light_cf(ws, cell_range, green_op, green_val, yellow_op=None, yellow_val=None,
                         red_op=None, red_val=None):
    """
    Add Green/Yellow/Red conditional formatting.

    Simple usage (numeric thresholds, higher=better):
        add_traffic_light_cf(ws, "J2:J100",
            green_op="greaterThanOrEqual", green_val="0.88",
            yellow_op="greaterThanOrEqual", yellow_val="0.85",
            red_op="lessThan", red_val="0.85")

    For text-based status (G/Y/R):
        Use add_text_traffic_light_cf instead.
    """
    # Green rule
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator=green_op,
        formula=[str(green_val)],
        fill=C.ALERT_GREEN_FILL,
        font=C.ALERT_GREEN_FONT,
    ))
    # Yellow rule (applied after green, lower priority)
    if yellow_op and yellow_val is not None:
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator=yellow_op,
            formula=[str(yellow_val)],
            fill=C.ALERT_YELLOW_FILL,
            font=C.ALERT_YELLOW_FONT,
        ))
    # Red rule (applied last, lowest priority)
    if red_op and red_val is not None:
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator=red_op,
            formula=[str(red_val)],
            fill=C.ALERT_RED_FILL,
            font=C.ALERT_RED_FONT,
        ))


def add_text_status_cf(ws, cell_range):
    """
    Add conditional formatting for text-based status: Green/Yellow/Red or G/Y/R or OK/WARNING/ALERT.
    """
    for text, fill, font in [
        ("Green", C.ALERT_GREEN_FILL, C.ALERT_GREEN_FONT),
        ("OK", C.ALERT_GREEN_FILL, C.ALERT_GREEN_FONT),
        ("G", C.ALERT_GREEN_FILL, C.ALERT_GREEN_FONT),
        ("Yellow", C.ALERT_YELLOW_FILL, C.ALERT_YELLOW_FONT),
        ("WARNING", C.ALERT_YELLOW_FILL, C.ALERT_YELLOW_FONT),
        ("Y", C.ALERT_YELLOW_FILL, C.ALERT_YELLOW_FONT),
        ("Red", C.ALERT_RED_FILL, C.ALERT_RED_FONT),
        ("ALERT", C.ALERT_RED_FILL, C.ALERT_RED_FONT),
        ("R", C.ALERT_RED_FILL, C.ALERT_RED_FONT),
        ("FLAG", C.ALERT_RED_FILL, C.ALERT_RED_FONT),
        ("BREACH", C.ALERT_RED_FILL, C.ALERT_RED_FONT),
    ]:
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="equal",
            formula=[f'"{text}"'],
            fill=fill,
            font=font,
        ))


def protect_sheet(ws, unlocked_columns=None, start_row=2, end_row=1000,
                  password="UltimateFarms2026"):
    """
    Protect entire sheet, optionally unlocking specific columns for data entry.

    Args:
        ws: worksheet
        unlocked_columns: list of 1-based column indices to unlock, or None for full lock
        start_row: first data row to unlock
        end_row: last data row to unlock
        password: protection password
    """
    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.formatCells = False
    ws.protection.sort = False
    ws.protection.autoFilter = False

    if unlocked_columns:
        for col_idx in unlocked_columns:
            col_letter = get_column_letter(col_idx)
            for row in range(start_row, end_row + 1):
                ws[f"{col_letter}{row}"].protection = Protection(locked=False)


def write_section_header(ws, row, col, text, merge_end_col=None):
    """Write a formatted section header."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = C.SECTION_HEADER_FONT
    cell.fill = C.SECTION_HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if merge_end_col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=merge_end_col,
        )


def apply_header_formatting(ws, start_row, start_col, num_cols):
    """Apply header formatting to an existing row."""
    for ci in range(num_cols):
        cell = ws.cell(row=start_row, column=start_col + ci)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
        cell.alignment = C.HEADER_ALIGN


def set_col_widths(ws, widths, start_col=1):
    """Set column widths from a list."""
    for ci, w in enumerate(widths):
        if w:
            ws.column_dimensions[get_column_letter(start_col + ci)].width = w


def freeze_panes(ws, cell="A2"):
    """Freeze panes at the specified cell (default: freeze header row)."""
    ws.freeze_panes = cell
