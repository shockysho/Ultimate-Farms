"""
tab_daily_entry.py -- Tab 2: Daily Entry (single-page data entry form).

A streamlined front-end for the recordkeeper to enter daily production data
for all 10 cages.  Auto-calculates totals, lay %, and yesterday comparison.
Generates transfer rows matching tblDailyCageLog schema for copy-paste into
Daily Log.

This sheet was originally hand-built in Excel by the farm owner; this module
codifies it into the Python builder so it regenerates on every build.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from . import config as C
from .config import T


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ENTRY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
AUTO_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # light green
REF_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")    # light blue
INACTIVE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
OK_FONT = Font(color="006100")
WARN_FONT = Font(color="9C0006")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

NUM_CAGES = len(C.HOUSING_DATA)  # 10


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------

def build_tab_daily_entry(wb):
    """Build Tab 2: Daily Entry -- single-page production entry form."""
    ws = wb.create_sheet(title=C.TAB_NAMES[2])

    ms = f"'{C.MASTER_SHEET}'"  # quoted sheet name for formulas
    prod = T("daily_cage_log")

    # ── Row 1: Title ──────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    ws["A1"].value = "DAILY PRODUCTION ENTRY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    # ── Row 3: Date ───────────────────────────────────────────────────────
    ws["A3"].value = "Date:"
    ws["A3"].font = Font(bold=True)
    ws["B3"].value = "=TODAY()"
    ws["B3"].number_format = C.FMT_DATE

    # ── Row 5-6: Summary banner ───────────────────────────────────────────
    ws.merge_cells("A5:B5")
    ws["A5"].value = "TODAY'S SUMMARY"
    ws["A5"].font = SECTION_FONT

    sum_headers = ["Total Eggs", "Total Crates", "Total Deaths", "Farm Lay %",
                   "Active Sides", "", "", "vs Yesterday", "Eggs +/-", "Deaths +/-"]
    for ci, h in enumerate(sum_headers):
        cell = ws.cell(row=5, column=3 + ci, value=h)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center")

    first_data = 10
    last_data = first_data + NUM_CAGES - 1

    # C6: Total eggs (array formula via SUMPRODUCT to avoid CSE issues)
    ws["C6"].value = f'=SUMPRODUCT((H{first_data}:H{last_data}<>"")*(H{first_data}:H{last_data}<>"--")*(ISNUMBER(H{first_data}:H{last_data}))*H{first_data}:H{last_data})'
    ws["C6"].number_format = C.FMT_INTEGER

    # D6: Total crates
    ws["D6"].value = "=INT(C6/30)"
    ws["D6"].number_format = C.FMT_INTEGER

    # E6: Total deaths
    ws["E6"].value = f"=SUM(J{first_data}:J{last_data})"
    ws["E6"].number_format = C.FMT_INTEGER

    # F6: Farm lay %
    ws["F6"].value = (
        f'=IFERROR(SUMPRODUCT((E{first_data}:E{last_data}="ACTIVE")'
        f'*H{first_data}:H{last_data})'
        f'/SUMPRODUCT((E{first_data}:E{last_data}="ACTIVE")'
        f'*D{first_data}:D{last_data})/30,0)'
    )
    ws["F6"].number_format = C.FMT_PERCENT

    # G6: Active sides
    ws["G6"].value = f'=COUNTIF(E{first_data}:E{last_data},"ACTIVE")'

    # J6: vs Yesterday eggs
    ws["J6"].value = (
        f'=IFERROR(IF(SUM(M{first_data}:M{last_data})=0,"--",'
        f'IF(C6>SUM(M{first_data}:M{last_data}),'
        f'"+ "&TEXT(C6-SUM(M{first_data}:M{last_data}),"#,##0"),'
        f'"- "&TEXT(SUM(M{first_data}:M{last_data})-C6,"#,##0"))),"--")'
    )

    # K6: vs Yesterday deaths
    ws["K6"].value = (
        f'=IFERROR(IF(SUM(J{first_data}:J{last_data})>0,'
        f'IF(SUM(J{first_data}:J{last_data})>3,'
        f'"!! "&SUM(J{first_data}:J{last_data})&" deaths",'
        f'SUM(J{first_data}:J{last_data})&" deaths"),"None"),"--")'
    )

    # ── Row 8-9: Production entry headers ─────────────────────────────────
    ws.merge_cells(f"A8:E8")
    ws["A8"].value = "PRODUCTION ENTRY"
    ws["A8"].font = SECTION_FONT

    ws.merge_cells(f"F8:G8")
    ws["F8"].value = ">> ENTER DATA HERE <<"
    ws["F8"].font = Font(bold=True, color="9C0006")
    ws["F8"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"H8:I8")
    ws["H8"].value = ">> AUTO-CALC <<"
    ws["H8"].font = Font(bold=True, color="006100")
    ws["H8"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"J8:L8")
    ws["J8"].value = ">> ENTER <<"
    ws["J8"].font = Font(bold=True, color="9C0006")
    ws["J8"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"M8:N8")
    ws["M8"].value = "YESTERDAY REF"
    ws["M8"].font = Font(bold=True, color="4472C4")
    ws["M8"].alignment = Alignment(horizontal="center")

    entry_headers = ["Cage", "Side", "Cohort", "Birds", "Status",
                     "Crates", "Singles", "Total Eggs", "Lay %",
                     "Deaths", "Death Cause", "Notes", "Prev Eggs", "% Chg"]
    for ci, h in enumerate(entry_headers):
        cell = ws.cell(row=9, column=ci + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Housing data rows (A70:J79 on Master Data)
    housing_start_row = 70  # tblHousing first data row on Master Data

    for i in range(NUM_CAGES):
        r = first_data + i
        mr = housing_start_row + i  # Master Data row

        # A: Cage ID from Master Data
        ws.cell(row=r, column=1, value=f"={ms}!A{mr}")
        # B: Side from Master Data
        ws.cell(row=r, column=2, value=f"={ms}!E{mr}")
        # C: Cohort ID from Master Data
        ws.cell(row=r, column=3, value=f"={ms}!G{mr}")
        # D: Bird Count from Master Data
        ws.cell(row=r, column=4, value=f"={ms}!F{mr}")
        # E: Status (ACTIVE/INACTIVE)
        ws.cell(row=r, column=5, value=f'=IF({ms}!H{mr}="Yes","ACTIVE","INACTIVE")')

        # F: Crates (USER ENTRY -- starts blank)
        cell_f = ws.cell(row=r, column=6)
        cell_f.fill = ENTRY_FILL
        cell_f.protection = Protection(locked=False)

        # G: Singles (USER ENTRY -- default 0)
        cell_g = ws.cell(row=r, column=7, value=0)
        cell_g.fill = ENTRY_FILL
        cell_g.protection = Protection(locked=False)

        # H: Total Eggs (auto-calc)
        ws.cell(row=r, column=8,
                value=f'=IF(E{r}="INACTIVE","--",IF(F{r}="","",F{r}*30+G{r}))')
        ws.cell(row=r, column=8).fill = AUTO_FILL
        ws.cell(row=r, column=8).number_format = C.FMT_INTEGER

        # I: Lay % per cage (auto-calc)
        ws.cell(row=r, column=9,
                value=f'=IF(OR(E{r}="INACTIVE",D{r}=0,H{r}="--",H{r}=""),"--",H{r}/D{r})')
        ws.cell(row=r, column=9).fill = AUTO_FILL
        ws.cell(row=r, column=9).number_format = C.FMT_PERCENT

        # J: Deaths (USER ENTRY)
        cell_j = ws.cell(row=r, column=10)
        cell_j.fill = ENTRY_FILL
        cell_j.protection = Protection(locked=False)

        # K: Death Cause (USER ENTRY -- dropdown)
        cell_k = ws.cell(row=r, column=11)
        cell_k.fill = ENTRY_FILL
        cell_k.protection = Protection(locked=False)

        # L: Notes (USER ENTRY)
        cell_l = ws.cell(row=r, column=12)
        cell_l.fill = ENTRY_FILL
        cell_l.protection = Protection(locked=False)

        # M: Previous day's eggs (SUMIFS from Daily Log)
        ws.cell(row=r, column=13,
                value=(
                    f'=IFERROR(SUMIFS({prod}[Total Eggs],'
                    f'{prod}[Date],TODAY()-1,'
                    f'{prod}[Cage ID],A{r}),0)'
                ))
        ws.cell(row=r, column=13).fill = REF_FILL
        ws.cell(row=r, column=13).number_format = C.FMT_INTEGER

        # N: % change vs yesterday
        ws.cell(row=r, column=14,
                value=f'=IF(OR(M{r}=0,H{r}="--",F{r}=""),"--",(H{r}-M{r})/M{r})')
        ws.cell(row=r, column=14).fill = REF_FILL
        ws.cell(row=r, column=14).number_format = C.FMT_PERCENT

    # ── Data validation on entry columns ──────────────────────────────────
    # Crates: whole number >= 0
    dv_crates = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0",
                               allow_blank=True, showErrorMessage=True,
                               errorTitle="Invalid", error="Enter whole crates >= 0")
    dv_crates.sqref = f"F{first_data}:F{last_data}"
    ws.add_data_validation(dv_crates)

    # Singles: whole number >= 0
    dv_singles = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0",
                                allow_blank=True, showErrorMessage=True,
                                errorTitle="Invalid", error="Enter singles >= 0")
    dv_singles.sqref = f"G{first_data}:G{last_data}"
    ws.add_data_validation(dv_singles)

    # Deaths: whole number >= 0
    dv_deaths = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0",
                               allow_blank=True, showErrorMessage=True,
                               errorTitle="Invalid", error="Enter deaths >= 0")
    dv_deaths.sqref = f"J{first_data}:J{last_data}"
    ws.add_data_validation(dv_deaths)

    # Death Cause dropdown
    causes = ",".join(["Predator", "Disease", "Heat Stress", "Injury",
                       "Egg Bound", "Prolapse", "Cannibalism", "Unknown", "Other"])
    dv_cause = DataValidation(type="list", formula1=f'"{causes}"',
                              allow_blank=True, showDropDown=False)
    dv_cause.sqref = f"K{first_data}:K{last_data}"
    ws.add_data_validation(dv_cause)

    # ── Conditional formatting: grey out INACTIVE rows ────────────────────
    for i in range(NUM_CAGES):
        r = first_data + i
        rng = f"A{r}:N{r}"
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'$E${r}="INACTIVE"'],
            fill=INACTIVE_FILL
        ))
        # Red highlight on deaths > 0
        ws.conditional_formatting.add(f"J{r}", FormulaRule(
            formula=[f'AND(ISNUMBER(J{r}),J{r}>0)'],
            fill=C.ALERT_RED_FILL, font=C.ALERT_RED_FONT
        ))
        # Red highlight if death entered but no cause
        ws.conditional_formatting.add(f"K{r}", FormulaRule(
            formula=[f'AND(ISNUMBER(J{r}),J{r}>0,K{r}="")'],
            fill=C.ALERT_RED_FILL
        ))
        # % change > 10% amber, > 20% red
        ws.conditional_formatting.add(f"N{r}", FormulaRule(
            formula=[f'AND(ISNUMBER(N{r}),ABS(N{r})>0.2)'],
            fill=C.ALERT_RED_FILL, font=C.ALERT_RED_FONT
        ))
        ws.conditional_formatting.add(f"N{r}", FormulaRule(
            formula=[f'AND(ISNUMBER(N{r}),ABS(N{r})>0.1)'],
            fill=C.ALERT_YELLOW_FILL, font=C.ALERT_YELLOW_FONT
        ))

    # ── Row 21-23: Egg Grading section ────────────────────────────────────
    grade_start = last_data + 2
    ws.merge_cells(f"A{grade_start}:N{grade_start}")
    ws[f"A{grade_start}"].value = "EGG GRADING (optional -- leave blank if not grading today)"
    ws[f"A{grade_start}"].font = SECTION_FONT

    gr = grade_start + 1
    grade_headers = ["", "Large", "Medium", "Small/Peewee", "Cracked/Broken",
                     "Grade Total", "Prod Total", "Match?"]
    for ci, h in enumerate(grade_headers):
        cell = ws.cell(row=gr, column=ci + 1, value=h)
        cell.font = Font(bold=True, size=10)

    gr2 = gr + 1
    ws.cell(row=gr2, column=1, value="Whole Farm")
    # B-E: user entry for grading (unlocked)
    for c in range(2, 6):
        cell = ws.cell(row=gr2, column=c)
        cell.fill = ENTRY_FILL
        cell.protection = Protection(locked=False)
    # F: Grade total
    ws.cell(row=gr2, column=6, value=f'=IF(SUM(B{gr2}:E{gr2})=0,"",SUM(B{gr2}:E{gr2}))')
    ws.cell(row=gr2, column=6).fill = AUTO_FILL
    # G: Production total
    ws.cell(row=gr2, column=7, value="=C6")
    ws.cell(row=gr2, column=7).fill = AUTO_FILL
    # H: Match check
    ws.cell(row=gr2, column=8,
            value=f'=IF(F{gr2}="","",IF(F{gr2}=G{gr2},"OK","MISMATCH"))')
    ws.cell(row=gr2, column=8).fill = AUTO_FILL

    ws.conditional_formatting.add(f"H{gr2}", FormulaRule(
        formula=[f'H{gr2}="MISMATCH"'],
        fill=MISMATCH_FILL, font=WARN_FONT
    ))

    # ── Validation checks section ─────────────────────────────────────────
    vc_start = gr2 + 2
    ws.merge_cells(f"A{vc_start}:N{vc_start}")
    ws[f"A{vc_start}"].value = "VALIDATION CHECKS"
    ws[f"A{vc_start}"].font = SECTION_FONT

    checks = [
        (
            "All active cages entered?",
            (
                f'=IF(COUNTIF(E{first_data}:E{last_data},"ACTIVE")=0,"No active cages",'
                f'IF(COUNTIFS(E{first_data}:E{last_data},"ACTIVE",F{first_data}:F{last_data},"<>")'
                f'=COUNTIF(E{first_data}:E{last_data},"ACTIVE"),"OK - All entered","MISSING entries"))'
            ),
        ),
        (
            "Deaths have causes?",
            (
                f'=IF(SUM(J{first_data}:J{last_data})=0,"OK - No deaths",'
                f'IF(SUMPRODUCT((J{first_data}:J{last_data}>0)*(K{first_data}:K{last_data}=""))>0,'
                f'"MISSING - deaths without cause","OK - All causes entered"))'
            ),
        ),
        (
            "Grade totals match?",
            f'=IF(F{gr2}="","Grading skipped",IF(F{gr2}=G{gr2},"OK - Match","MISMATCH - Off by "&ABS(F{gr2}-G{gr2})))',
        ),
        (
            "Any anomalies?",
            (
                f'=IF(SUMPRODUCT((ISNUMBER(N{first_data}:N{last_data}))*(ABS(N{first_data}:N{last_data})>0.2))>0,'
                f'"WARNING - Large swings detected","OK - Normal variation")'
            ),
        ),
    ]

    vr = vc_start + 1
    for label, formula in checks:
        ws.cell(row=vr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=vr, column=2, value=formula)
        # Green/red CF on check results
        ws.conditional_formatting.add(f"B{vr}", FormulaRule(
            formula=[f'LEFT(B{vr},2)="OK"'],
            font=OK_FONT
        ))
        ws.conditional_formatting.add(f"B{vr}", FormulaRule(
            formula=[f'OR(LEFT(B{vr},4)="MISS",LEFT(B{vr},4)="MISM",LEFT(B{vr},4)="WARN")'],
            font=WARN_FONT, fill=C.ALERT_RED_FILL
        ))
        vr += 1

    # ── Transfer rows section ─────────────────────────────────────────────
    tr_start = vr + 1
    ws.merge_cells(f"A{tr_start}:N{tr_start}")
    ws[f"A{tr_start}"].value = (
        "TRANSFER TO DAILY LOG -- Copy rows below into 'Daily Log' sheet (Ctrl+C the range, paste as values)"
    )
    ws[f"A{tr_start}"].font = SECTION_FONT

    # Transfer headers matching tblDailyCageLog columns
    transfer_headers = [
        "Date", "Cage ID", "Side", "Crates", "Singles",
        "Large", "Medium", "Small/Peewee", "Cracked/Broken",
        "Deaths", "Death Cause", "Disease Sub", "Mort Action", "Notes",
        "Total Eggs", "Equiv Crates", "Rem Singles", "Cracked %", "Large %",
    ]
    thr = tr_start + 1
    for ci, h in enumerate(transfer_headers):
        cell = ws.cell(row=thr, column=ci + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # One transfer row per cage
    for i in range(NUM_CAGES):
        tr = thr + 1 + i
        dr = first_data + i  # data entry row

        # Only output if cage is ACTIVE and crates entered
        skip = f'OR($E{dr}="INACTIVE",$F{dr}="")'

        # A: Date
        ws.cell(row=tr, column=1, value=f'=IF({skip},"",$B$3)')
        ws.cell(row=tr, column=1).number_format = C.FMT_DATE
        # B: Cage ID
        ws.cell(row=tr, column=2, value=f'=IF({skip},"",A{dr})')
        # C: Side
        ws.cell(row=tr, column=3, value=f'=IF({skip},"",B{dr})')
        # D: Crates
        ws.cell(row=tr, column=4, value=f'=IF({skip},"",F{dr})')
        # E: Singles
        ws.cell(row=tr, column=5, value=f'=IF({skip},"",G{dr})')

        # F-I: Egg grading distribution (proportional from whole-farm grading)
        for gc, gcol in enumerate(["B", "C", "D", "E"], start=6):
            col_letter = gcol
            ws.cell(row=tr, column=gc,
                    value=(
                        f'=IF(OR({skip},${col_letter}${gr2}=""),"",ROUND(${col_letter}${gr2}'
                        f'*H{dr}/IF(C6=0,1,C6),0))'
                    ))

        # J: Deaths
        ws.cell(row=tr, column=10, value=f'=IF({skip},"",IF(J{dr}="",0,J{dr}))')
        # K: Death Cause
        ws.cell(row=tr, column=11, value=f'=IF({skip},"",K{dr})')
        # L: Disease Sub (blank -- entered later if needed)
        ws.cell(row=tr, column=12, value=f'=IF({skip},"","")')
        # M: Mortality Action (blank)
        ws.cell(row=tr, column=13, value=f'=IF({skip},"","")')
        # N: Notes
        ws.cell(row=tr, column=14, value=f'=IF({skip},"",L{dr})')
        # O: Total Eggs
        ws.cell(row=tr, column=15, value=f'=IF({skip},"",H{dr})')
        # P: Equivalent Crates
        ws.cell(row=tr, column=16, value=f'=IF({skip},"",INT(O{tr}/30))')
        # Q: Remainder Singles
        ws.cell(row=tr, column=17, value=f'=IF({skip},"",MOD(O{tr},30))')
        # R: Cracked %
        ws.cell(row=tr, column=18,
                value=f'=IF(OR({skip},O{tr}=0,$I${gr2}=""),"",I{tr}/O{tr})')
        ws.cell(row=tr, column=18).number_format = C.FMT_PERCENT
        # S: Large %
        ws.cell(row=tr, column=19,
                value=f'=IF(OR({skip},O{tr}=0,$F${gr2}=""),"",F{tr}/O{tr})')
        ws.cell(row=tr, column=19).number_format = C.FMT_PERCENT

    # ── Instructions row ──────────────────────────────────────────────────
    inst_row = thr + 1 + NUM_CAGES + 1
    ws.merge_cells(f"A{inst_row}:N{inst_row}")
    ws[f"A{inst_row}"].value = (
        "INSTRUCTIONS: Review validation checks above. If all pass, "
        "select A{s}:S{e}, Copy, go to Daily Log, paste as values at next empty row.".format(
            s=thr + 1, e=thr + NUM_CAGES)
    )
    ws[f"A{inst_row}"].font = Font(italic=True, color="4472C4")

    # ── Column widths ─────────────────────────────────────────────────────
    widths = [10, 6, 10, 7, 9, 8, 8, 10, 8, 8, 14, 20, 10, 8, 10, 10, 8, 8, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26 else ""].width = w
    from openpyxl.utils import get_column_letter
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet protection (lock everything except entry cells) ─────────────
    ws.protection.sheet = True
    ws.protection.password = "UltimateFarms2026"
    ws.protection.formatCells = False
    ws.protection.sort = False
    ws.protection.autoFilter = False

    # Freeze top rows
    ws.freeze_panes = f"A{first_data}"

    print("  Tab 2 (Daily Entry): Production entry form created")
    return ws
