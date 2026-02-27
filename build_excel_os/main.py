"""
main.py -- Orchestrator for Ultimate Farms Excel OS v3.0 (9-tab Verification Core).

Usage:
    python -m build_excel_os.main
    python -m build_excel_os.main --legacy-file "<path>"
"""

import os
import sys
import zipfile
from pathlib import Path
from openpyxl import Workbook

from . import config as C
from .sample_data import generate_all_sample_data
from .import_legacy_data import import_all_legacy_data
from .tab6_master_data import build_tab6_master_data
from .tab_daily_entry import build_tab_daily_entry
from .tab_inputs import build_inputs
from .tab7_engine import build_tab7_engine
from .tab8_analytics import build_tab8_analytics
from .tab1_dashboard import build_tab1_dashboard


OUTPUT_DIR = Path(__file__).resolve().parent.parent
OUTPUT_FILE = OUTPUT_DIR / "Ultimate_Farms_Excel_OS_v3.0.xlsx"


def apply_tab_colors(wb):
    """Apply tab colors to every worksheet."""
    tab_title_to_number = {v: k for k, v in C.TAB_NAMES.items()}
    for ws in wb.worksheets:
        tab_num = tab_title_to_number.get(ws.title)
        if tab_num and tab_num in C.TAB_COLORS:
            ws.sheet_properties.tabColor = C.TAB_COLORS[tab_num]


def reorder_sheets(wb):
    """Reorder sheets: Dashboard first, then Daily Entry through Analytics."""
    desired_order = [C.TAB_NAMES[i] for i in range(1, 10)]
    name_to_ws = {ws.title: ws for ws in wb.worksheets}
    ordered = []
    for name in desired_order:
        if name in name_to_ws:
            ordered.append(name_to_ws.pop(name))
    for ws in name_to_ws.values():
        ordered.append(ws)
    wb._sheets = ordered


def remove_default_sheet(wb):
    """Remove the default 'Sheet' created by openpyxl."""
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]


def strip_calc_chain(filepath):
    """Remove calcChain.xml from xlsx to avoid formula chain corruption."""
    tmp = str(filepath) + ".tmp"
    with zipfile.ZipFile(str(filepath), "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == "xl/calcChain.xml":
                    continue
                zout.writestr(item, zin.read(item.filename))
    os.replace(tmp, str(filepath))


def build():
    """Main build pipeline."""
    print("=" * 60)
    print("  Ultimate Farms Excel OS v3.0 (Verification Core) -- Build Starting")
    print("=" * 60)

    # Check for --legacy-file flag
    legacy_file = None
    if len(sys.argv) > 2 and sys.argv[1] == "--legacy-file":
        legacy_file = sys.argv[2]
    elif len(sys.argv) > 1 and not sys.argv[1].startswith("-"):
        # Also accept positional argument
        legacy_file = sys.argv[1]

    # 1. Create workbook
    wb = Workbook()
    print("\n[1/8] Workbook created")

    # 2. Generate or import data
    dynamic_customers = []
    dynamic_vendors = []
    if legacy_file:
        print(f"[2/8] Importing legacy data from: {legacy_file}")
        data, dynamic_customers, dynamic_vendors = import_all_legacy_data(legacy_file)
    else:
        print("[2/8] Generating sample data...")
        data = generate_all_sample_data()
    print(f"       {sum(len(v) for v in data.values() if isinstance(v, list))} total data rows generated")

    # 3. Build Tab 7: Master Data (must be first -- dropdowns depend on it)
    print("[3/8] Building Tab 7: Master Data (11 tables)...")
    master_rows = build_tab6_master_data(wb, dynamic_customers=dynamic_customers,
                                          dynamic_vendors=dynamic_vendors)

    # 4. Build Tab 2: Daily Entry (production entry form)
    print("[4/8] Building Tab 2: Daily Entry (entry form)...")
    build_tab_daily_entry(wb)

    # 5. Build Tabs 3-6: Input tabs
    print("[5/8] Building Tabs 3-6: Input Tabs (15 tables)...")
    build_inputs(wb, data, master_rows)

    # 6. Build Tab 8: Engine (5 recon + Ghost Money + 41 fraud flags)
    print("[6/8] Building Tab 8: Engine (8 sections: Target Resolver + 5 Recon + Ghost Money + 41 Fraud Flags)...")
    build_tab7_engine(wb)

    # 7. Build Tab 9: Analytics
    print("[7/8] Building Tab 9: Analytics (7 sections incl. Ghost Money Trends)...")
    build_tab8_analytics(wb)

    # 8. Build Tab 1: Dashboard (Ghost Money + 5 recon engines)
    print("[8/8] Building Tab 1: Dashboard (Ghost Money + 5 Recon Engines)...")
    build_tab1_dashboard(wb)

    # Post-processing
    print("\n--- Post-processing ---")
    print("  Applying tab colors...")
    apply_tab_colors(wb)

    print("  Reordering sheets (Dashboard > Daily Entry > Daily Log > ... > Analytics)...")
    reorder_sheets(wb)

    remove_default_sheet(wb)

    print(f"\n  Saving to: {OUTPUT_FILE}")
    wb.save(str(OUTPUT_FILE))

    print("  Stripping calcChain.xml...")
    strip_calc_chain(OUTPUT_FILE)

    # Summary
    print("\n" + "=" * 60)
    print(f"  BUILD COMPLETE -- {len(wb.sheetnames)} tabs generated")
    print(f"  File: {OUTPUT_FILE}")
    print(f"  Size: {os.path.getsize(OUTPUT_FILE) / 1024:.0f} KB")
    print("=" * 60)

    print("\nTab listing:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i:2d}. {name}")

    return wb


if __name__ == "__main__":
    build()
