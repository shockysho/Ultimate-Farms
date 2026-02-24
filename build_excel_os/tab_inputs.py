"""
tab_inputs.py -- Build Tabs 2-5: Input sheets (multi-table per sheet).
Each function places multiple Excel Tables on a single worksheet,
stacked vertically with section headers and 2-row gaps.

All dropdown validations reference Tab 6 (Master Data) using master_rows.
"""

from openpyxl.utils import get_column_letter

from . import config as C
from .config import T
from .helpers import (
    create_excel_table, add_dropdown_validation, add_inline_dropdown,
    add_integer_validation, add_date_validation,
    protect_sheet, freeze_panes, set_col_widths, write_section_header,
    add_traffic_light_cf, add_text_status_cf,
)

MS = C.MASTER_SHEET  # "Master Data"


def _dd_range(master_rows, table_key, col_letter):
    """Build absolute range string for dropdown validation from master_rows.
    E.g. _dd_range(master_rows, "staff", "A") -> "$A$102:$A$113"
    """
    tbl_name = T(table_key)
    hdr, end = master_rows[tbl_name]
    return f"${col_letter}${hdr + 1}:${col_letter}${end}"


# ============================================================================
# Tab 2: Daily Log
# ============================================================================

def build_tab2_daily_log(wb, data, master_rows):
    """Tab 2: Daily Log -- 5 tables on one sheet.

    Section A: tblDailyCageLog (merged production+mortality)
    Section B: tblEnvironmental
    Section C: tblWaterConsumption
    Section D: tblBiosecurity
    Section E: tblLabor
    """
    ws = wb.create_sheet(title=C.TAB_NAMES[3])
    row = 1

    # ── Section A: Daily Cage Log ──────────────────────────────────────
    write_section_header(ws, row, 1, "DAILY CAGE LOG (Production + Mortality)", merge_end_col=20)
    row += 1

    headers_a = [
        "Date", "Cage ID", "Side", "Crates Collected", "Singles Collected",
        "Grade: Large", "Grade: Medium", "Grade: Small/Peewee", "Grade: Cracked/Broken",
        "Deaths", "Death Cause", "Disease Sub-category", "Mortality Action", "Notes",
        # Auto-calcs
        "Total Eggs", "Equivalent Crates", "Remainder Singles",
        "Cracked %", "Large %", "Grade Validation",
    ]

    rows_a = []
    for r in data["daily_cage_log"]:
        rows_a.append([
            r[0], r[1], r[2], r[3], r[4],  # Date, Cage, Side, Crates, Singles
            r[5], r[6], r[7], r[8],          # Large, Medium, Small, Cracked
            r[9], r[10], r[11], r[12], r[13],  # Deaths, Cause, Sub, Action, Notes
            None, None, None, None, None, None,  # Formula columns
        ])

    calc_a = {
        14: "([@Crates Collected]*30)+[@Singles Collected]",
        15: "INT([@Total Eggs]/30)",
        16: "MOD([@Total Eggs],30)",
        17: 'IF([@Total Eggs]=0,0,[@Grade: Cracked/Broken]/[@Total Eggs])',
        18: 'IF([@Total Eggs]=0,0,[@Grade: Large]/[@Total Eggs])',
        19: 'IF([@Grade: Large]+[@Grade: Medium]+[@Grade: Small/Peewee]+[@Grade: Cracked/Broken]=[@Total Eggs],"OK","MISMATCH")',
    }

    tab_a, end_a = create_excel_table(
        ws, T("daily_cage_log"), headers_a, rows_a, start_row=row,
        col_widths=[12, 10, 6, 14, 14, 12, 12, 16, 18, 8, 14, 16, 16, 20, 10, 14, 14, 10, 10, 14],
        calculated_columns=calc_a,
        number_formats={17: C.FMT_PERCENT, 18: C.FMT_PERCENT},
    )

    # Dropdowns for cage log
    dr = end_a
    add_dropdown_validation(ws, f"B{row+1}:B{dr}", MS, _dd_range(master_rows, "housing", "A"))
    add_inline_dropdown(ws, f"C{row+1}:C{dr}", C.DROPDOWN_LISTS["sides"])
    add_integer_validation(ws, f"D{row+1}:D{dr}", 0, 999)
    add_integer_validation(ws, f"E{row+1}:E{dr}", 0, 29)
    add_inline_dropdown(ws, f"K{row+1}:K{dr}", C.DROPDOWN_LISTS["mortality_cause"])
    add_inline_dropdown(ws, f"L{row+1}:L{dr}", C.DROPDOWN_LISTS["disease_subcategory"])
    add_inline_dropdown(ws, f"M{row+1}:M{dr}", C.DROPDOWN_LISTS["mortality_action"])

    add_traffic_light_cf(ws, f"R{row+1}:R{dr}",
                         green_op="lessThanOrEqual", green_val="0.04",
                         yellow_op="lessThanOrEqual", yellow_val="0.06",
                         red_op="greaterThan", red_val="0.06")
    add_text_status_cf(ws, f"T{row+1}:T{dr}")

    unlock_a = list(range(1, 15))  # Columns A-N (input), O-T locked (formulas)
    protect_sheet(ws, unlocked_columns=unlock_a, start_row=row+1, end_row=max(dr, 500))

    row = end_a + 3

    # ── Section B: Environmental ──────────────────────────────────────
    write_section_header(ws, row, 1, "ENVIRONMENTAL LOG", merge_end_col=14)
    row += 1

    headers_b = [
        "Date", "House/Zone", "Time of Reading", "Temperature (C)",
        "Hours Above 31C", "Cooling Status", "Feed Refusal (AM)",
        "Feed Refusal (PM)", "Trough Locations Checked", "Water Status", "Notes",
        "Heat Stress Score", "Feed Refusal Flag", "Max Temp Today",
    ]

    rows_b = []
    for r in data["environmental"]:
        rows_b.append(list(r) + [None, None, None])

    calc_b = {
        11: '[@Hours Above 31C]',
        12: 'IF(OR([@Feed Refusal (AM)]="Full",[@Feed Refusal (PM)]="Full"),"YES","NO")',
        13: '[@Temperature (C)]',
    }

    tab_b, end_b = create_excel_table(
        ws, T("environmental"), headers_b, rows_b, start_row=row,
        col_widths=[12, 14, 12, 14, 14, 14, 14, 14, 25, 12, 20, 14, 14, 12],
        calculated_columns=calc_b,
        number_formats={3: C.FMT_DECIMAL_1, 4: C.FMT_DECIMAL_1},
    )

    add_inline_dropdown(ws, f"F{row+1}:F{end_b}", C.DROPDOWN_LISTS["cooling_status"])
    add_inline_dropdown(ws, f"G{row+1}:G{end_b}", C.DROPDOWN_LISTS["feed_refusal"])
    add_inline_dropdown(ws, f"H{row+1}:H{end_b}", C.DROPDOWN_LISTS["feed_refusal"])
    add_inline_dropdown(ws, f"J{row+1}:J{end_b}", C.DROPDOWN_LISTS["water_status"])

    row = end_b + 3

    # ── Section C: Water Consumption ──────────────────────────────────
    write_section_header(ws, row, 1, "WATER CONSUMPTION LOG", merge_end_col=10)
    row += 1

    headers_c = [
        "Date", "House/Zone", "Meter Reading AM (L)", "Meter Reading PM (L)",
        "Estimation Method", "Notes",
        "Daily Consumption (L)", "Liters per 1000 Birds", "7-Day Average", "Variance from Avg",
    ]

    rows_c = []
    for r in data["water_consumption"]:
        rows_c.append(list(r) + [None, None, None, None])

    housing_tbl = T("housing")
    water_tbl = T("water_consumption")
    calc_c = {
        6: '[@Meter Reading PM (L)]-[@Meter Reading AM (L)]',
        7: f'IF(SUMIFS({housing_tbl}[Bird Count],{housing_tbl}[House],[@House/Zone])=0,0,[@Daily Consumption (L)]/(SUMIFS({housing_tbl}[Bird Count],{housing_tbl}[House],[@House/Zone])/1000))',
        8: f'IFERROR(AVERAGEIFS({water_tbl}[Daily Consumption (L)],{water_tbl}[Date],">="&([@Date]-6),{water_tbl}[Date],"<="&[@Date],{water_tbl}[House/Zone],[@House/Zone]),0)',
        9: 'IF([@7-Day Average]=0,0,([@Daily Consumption (L)]-[@7-Day Average])/[@7-Day Average])',
    }

    tab_c, end_c = create_excel_table(
        ws, T("water_consumption"), headers_c, rows_c, start_row=row,
        col_widths=[12, 14, 18, 18, 16, 20, 18, 18, 14, 14],
        calculated_columns=calc_c,
        number_formats={2: C.FMT_INTEGER, 3: C.FMT_INTEGER, 6: C.FMT_INTEGER,
                        7: C.FMT_DECIMAL_1, 8: C.FMT_DECIMAL_1, 9: C.FMT_PERCENT},
    )

    add_inline_dropdown(ws, f"E{row+1}:E{end_c}", ["Flow meter", "Polytank level", "Manual count"])

    row = end_c + 3

    # ── Section D: Biosecurity ────────────────────────────────────────
    write_section_header(ws, row, 1, "DAILY FORTRESS CHECKLIST (BIOSECURITY)", merge_end_col=14)
    row += 1

    headers_d = [
        "Date", "Footbath OK", "Disinfectant Fresh", "Restricted Entry OK",
        "Houses Locked", "Rodent Control Checked", "Dead Bird Disposal OK",
        "PPE Compliance", "Time Separation OK", "Right-side Door OK",
        "Supervisor Sign-off", "Notes",
        "Compliance Score %", "Critical Breach",
    ]

    rows_d = []
    for r in data["biosecurity"]:
        rows_d.append(list(r) + [None, None])

    calc_d = {
        12: '(COUNTIF(INDIRECT("B"&ROW()&":J"&ROW()),"Yes"))/9',
        13: 'IF(OR([@Footbath OK]="No",[@Restricted Entry OK]="No",[@Dead Bird Disposal OK]="No"),"BREACH","OK")',
    }

    tab_d, end_d = create_excel_table(
        ws, T("biosecurity"), headers_d, rows_d, start_row=row,
        col_widths=[12, 12, 14, 16, 12, 18, 18, 14, 16, 16, 16, 20, 14, 12],
        calculated_columns=calc_d,
        number_formats={12: C.FMT_PERCENT},
    )

    for col_l in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        add_inline_dropdown(ws, f"{col_l}{row+1}:{col_l}{end_d}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"K{row+1}:K{end_d}", MS, _dd_range(master_rows, "staff", "A"))

    add_text_status_cf(ws, f"N{row+1}:N{end_d}")
    add_traffic_light_cf(ws, f"M{row+1}:M{end_d}",
                         green_op="greaterThanOrEqual", green_val="0.9",
                         yellow_op="greaterThanOrEqual", yellow_val="0.8",
                         red_op="lessThan", red_val="0.8")

    row = end_d + 3

    # ── Section E: Labor ──────────────────────────────────────────────
    write_section_header(ws, row, 1, "LABOR / COMPLIANCE LOG", merge_end_col=15)
    row += 1

    headers_e = [
        "Date", "Staff ID", "Team", "Shift", "Present",
        "Arrival Time", "Tasks Assigned", "Tasks Completed",
        "Verified By", "Violations", "Strike Issued",
        "Strike Reason", "Supervisor Notes",
        "Completion Rate %", "Attendance (30d)",
    ]

    rows_e = []
    for r in data["labor"]:
        rows_e.append(list(r) + [None, None])

    labor_tbl = T("labor")
    calc_e = {
        13: 'IF([@Tasks Assigned]=0,0,[@Tasks Completed]/[@Tasks Assigned])',
        14: (
            f'IF(COUNTIFS({labor_tbl}[Staff ID],[@Staff ID],{labor_tbl}[Date],">="&([@Date]-29),{labor_tbl}[Date],"<="&[@Date])=0,0,'
            f'COUNTIFS({labor_tbl}[Staff ID],[@Staff ID],{labor_tbl}[Date],">="&([@Date]-29),{labor_tbl}[Date],"<="&[@Date],{labor_tbl}[Present],"Yes")'
            f'/COUNTIFS({labor_tbl}[Staff ID],[@Staff ID],{labor_tbl}[Date],">="&([@Date]-29),{labor_tbl}[Date],"<="&[@Date]))'
        ),
    }

    tab_e, end_e = create_excel_table(
        ws, T("labor"), headers_e, rows_e, start_row=row,
        col_widths=[12, 10, 10, 12, 8, 12, 14, 14, 12, 16, 12, 20, 20, 14, 14],
        calculated_columns=calc_e,
        number_formats={13: C.FMT_PERCENT, 14: C.FMT_PERCENT},
    )

    add_dropdown_validation(ws, f"B{row+1}:B{end_e}", MS, _dd_range(master_rows, "staff", "A"))
    add_inline_dropdown(ws, f"C{row+1}:C{end_e}", C.DROPDOWN_LISTS["team"])
    add_inline_dropdown(ws, f"D{row+1}:D{end_e}", C.DROPDOWN_LISTS["shift"])
    add_inline_dropdown(ws, f"E{row+1}:E{end_e}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"I{row+1}:I{end_e}", MS, _dd_range(master_rows, "staff", "A"))
    add_inline_dropdown(ws, f"J{row+1}:J{end_e}", C.DROPDOWN_LISTS["violations"])
    add_inline_dropdown(ws, f"K{row+1}:K{end_e}", C.DROPDOWN_LISTS["strike_level"])

    add_traffic_light_cf(ws, f"N{row+1}:N{end_e}",
                         green_op="greaterThanOrEqual", green_val="0.9",
                         yellow_op="greaterThanOrEqual", yellow_val="0.8",
                         red_op="lessThan", red_val="0.8")

    freeze_panes(ws, "A2")
    print("  Tab 3 (Daily Log): 5 tables on single sheet")
    return ws


# ============================================================================
# Tab 3: Feed & Inventory
# ============================================================================

def build_tab3_feed_inventory(wb, data, master_rows):
    """Tab 3: Feed & Inventory -- 4 tables on one sheet.

    Section A: tblFeedConsumption
    Section B: tblIngredientMovement
    Section C: tblFeedMixBatchHeader
    Section D: tblInventoryCount
    """
    ws = wb.create_sheet(title=C.TAB_NAMES[4])
    row = 1

    # ── Section A: Feed Consumption ───────────────────────────────────
    write_section_header(ws, row, 1, "DAILY FEED CONSUMPTION LOG", merge_end_col=13)
    row += 1

    headers_a = [
        "Date", "Feeding Round", "Cohort ID", "Cage ID (optional)",
        "Feed Type", "Qty Issued (kg)", "Qty Returned (kg)",
        "Issued By", "Verified By", "Notes",
        "Net Consumed (kg)", "Intake/Bird (g)", "Daily FCR",
    ]

    rows_a = []
    for r in data["feed_consumption"]:
        rows_a.append(list(r) + [None, None, None])

    flock_tbl = T("flock")
    prod_tbl = T("daily_cage_log")
    calc_a = {
        10: '[@Qty Issued (kg)]-[@Qty Returned (kg)]',
        11: f'IF(SUMIFS({flock_tbl}[Current Bird Count],{flock_tbl}[Cohort ID],[@Cohort ID])=0,0,([@Net Consumed (kg)]*1000)/SUMIFS({flock_tbl}[Current Bird Count],{flock_tbl}[Cohort ID],[@Cohort ID]))',
        12: f'IF(SUMIFS({prod_tbl}[Total Eggs],{prod_tbl}[Date],[@Date])=0,0,[@Net Consumed (kg)]/(SUMIFS({prod_tbl}[Total Eggs],{prod_tbl}[Date],[@Date])/12))',
    }

    tab_a, end_a = create_excel_table(
        ws, T("feed_consumption"), headers_a, rows_a, start_row=row,
        col_widths=[12, 12, 12, 16, 14, 14, 14, 12, 12, 20, 14, 14, 10],
        calculated_columns=calc_a,
        number_formats={5: C.FMT_DECIMAL_1, 6: C.FMT_DECIMAL_1, 10: C.FMT_DECIMAL_1,
                        11: C.FMT_DECIMAL_1, 12: C.FMT_DECIMAL_2},
    )

    add_inline_dropdown(ws, f"B{row+1}:B{end_a}", C.DROPDOWN_LISTS["feeding_round"])
    add_dropdown_validation(ws, f"C{row+1}:C{end_a}", MS, _dd_range(master_rows, "flock", "A"))
    add_dropdown_validation(ws, f"E{row+1}:E{end_a}", MS, _dd_range(master_rows, "item_master", "A"))
    add_dropdown_validation(ws, f"H{row+1}:H{end_a}", MS, _dd_range(master_rows, "staff", "A"))
    add_dropdown_validation(ws, f"I{row+1}:I{end_a}", MS, _dd_range(master_rows, "staff", "A"))

    row = end_a + 3

    # ── Section B: Ingredient Movement ────────────────────────────────
    write_section_header(ws, row, 1, "INGREDIENT INVENTORY MOVEMENT", merge_end_col=15)
    row += 1

    headers_b = [
        "Date", "Ingredient", "Movement Type", "Quantity (kg)", "Bag Count",
        "Supplier", "Batch/Lot #", "Price/kg (if Received)", "Receipt/Invoice Ref",
        "Weight Verified", "Issued By", "Verified By", "Notes",
        "Line Value (GHS)", "Running Stock",
    ]

    rows_b = []
    for r in data["ingredient_movement"]:
        rows_b.append(list(r) + [None, None])

    ing_tbl = T("ingredient_movement")
    calc_b = {
        13: 'IF([@Movement Type]="Received",[@Quantity (kg)]*[@Price/kg (if Received)],0)',
        14: (
            f'SUMIFS({ing_tbl}[Quantity (kg)],{ing_tbl}[Ingredient],[@Ingredient],{ing_tbl}[Movement Type],"Received")'
            f'-SUMIFS({ing_tbl}[Quantity (kg)],{ing_tbl}[Ingredient],[@Ingredient],{ing_tbl}[Movement Type],"Issued-to-Mixing")'
            f'-SUMIFS({ing_tbl}[Quantity (kg)],{ing_tbl}[Ingredient],[@Ingredient],{ing_tbl}[Movement Type],"Wasted")'
        ),
    }

    tab_b, end_b = create_excel_table(
        ws, T("ingredient_movement"), headers_b, rows_b, start_row=row,
        col_widths=[12, 12, 16, 12, 10, 10, 14, 16, 16, 12, 10, 10, 20, 14, 14],
        calculated_columns=calc_b,
        number_formats={3: C.FMT_DECIMAL_1, 7: C.FMT_DECIMAL_2, 13: C.FMT_CURRENCY, 14: C.FMT_DECIMAL_1},
    )

    add_dropdown_validation(ws, f"B{row+1}:B{end_b}", MS, _dd_range(master_rows, "item_master", "A"))
    add_inline_dropdown(ws, f"C{row+1}:C{end_b}", C.DROPDOWN_LISTS["movement_type"])
    add_dropdown_validation(ws, f"F{row+1}:F{end_b}", MS, _dd_range(master_rows, "vendor", "A"))
    add_inline_dropdown(ws, f"J{row+1}:J{end_b}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"K{row+1}:K{end_b}", MS, _dd_range(master_rows, "staff", "A"))
    add_dropdown_validation(ws, f"L{row+1}:L{end_b}", MS, _dd_range(master_rows, "staff", "A"))

    row = end_b + 3

    # ── Section C: Feed Mix Batch Header ──────────────────────────────
    write_section_header(ws, row, 1, "FEED MIXING / FORMULATION BATCH LOG", merge_end_col=13)
    row += 1

    headers_c = [
        "Batch ID", "Date Mixed", "Feed Type Produced", "Formula Reference",
        "Output Qty (kg)", "Output Qty (bags)", "Mixed By",
        "QC: Moisture", "QC: Particle Size", "QC: Visual", "Notes",
        "Cost/kg", "Yield Efficiency",
    ]

    rows_c = []
    for r in data["feed_mixing"]:
        rows_c.append(list(r) + [None, None])

    calc_c = {
        11: 'IF([@Output Qty (kg)]=0,0,0)',
        12: 'IF([@Output Qty (kg)]=0,0,0.97)',
    }

    tab_c, end_c = create_excel_table(
        ws, T("feed_mix_header"), headers_c, rows_c, start_row=row,
        col_widths=[10, 12, 24, 16, 14, 14, 12, 12, 14, 12, 20, 10, 14],
        calculated_columns=calc_c,
        number_formats={4: C.FMT_DECIMAL_1, 5: C.FMT_DECIMAL_1, 11: C.FMT_CURRENCY, 12: C.FMT_PERCENT},
    )

    add_dropdown_validation(ws, f"D{row+1}:D{end_c}", MS, _dd_range(master_rows, "feed_formulation", "A"))
    add_dropdown_validation(ws, f"G{row+1}:G{end_c}", MS, _dd_range(master_rows, "staff", "A"))
    add_inline_dropdown(ws, f"H{row+1}:H{end_c}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"I{row+1}:I{end_c}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"J{row+1}:J{end_c}", C.DROPDOWN_LISTS["qc_visual"])

    row = end_c + 3

    # ── Section D: Inventory Count ────────────────────────────────────
    write_section_header(ws, row, 1, "INVENTORY COUNT LOG", merge_end_col=13)
    row += 1

    headers_d = [
        "Date", "Trigger", "Item", "Storage Location",
        "Expected Stock", "Physical Count",
        "Variance", "Variance %",
        "Explanation", "Counted By", "Verified By", "Approved By", "Notes",
    ]

    rows_d = [list(r) for r in data["inventory_count"]]

    calc_d = {
        6: '[@Physical Count]-[@Expected Stock]',
        7: 'IF([@Expected Stock]=0,0,[@Variance]/[@Expected Stock])',
    }

    tab_d, end_d = create_excel_table(
        ws, T("inventory_count"), headers_d, rows_d, start_row=row,
        col_widths=[12, 14, 12, 14, 14, 14, 10, 10, 25, 12, 12, 12, 20],
        calculated_columns=calc_d,
        number_formats={4: C.FMT_DECIMAL_1, 5: C.FMT_DECIMAL_1, 6: C.FMT_DECIMAL_1, 7: C.FMT_PERCENT},
    )

    add_inline_dropdown(ws, f"B{row+1}:B{end_d}", C.DROPDOWN_LISTS["count_trigger"])
    add_dropdown_validation(ws, f"C{row+1}:C{end_d}", MS, _dd_range(master_rows, "item_master", "A"))
    add_dropdown_validation(ws, f"J{row+1}:J{end_d}", MS, _dd_range(master_rows, "staff", "A"))
    add_dropdown_validation(ws, f"K{row+1}:K{end_d}", MS, _dd_range(master_rows, "staff", "A"))

    add_traffic_light_cf(ws, f"H{row+1}:H{end_d}",
                         green_op="between", green_val="-0.02",
                         yellow_op="between", yellow_val="-0.05",
                         red_op="lessThan", red_val="-0.05")

    freeze_panes(ws, "A2")
    protect_sheet(ws)
    print("  Tab 4 (Feed & Inventory): 4 tables on single sheet")
    return ws


# ============================================================================
# Tab 4: Sales & Procurement
# ============================================================================

def build_tab4_sales_procurement(wb, data, master_rows):
    """Tab 4: Sales & Procurement -- 2 tables on one sheet.

    Section A: tblSales
    Section B: tblProcurement
    """
    ws = wb.create_sheet(title=C.TAB_NAMES[5])
    row = 1

    # ── Section A: Sales ──────────────────────────────────────────────
    write_section_header(ws, row, 1, "SALES TRANSACTION LOG", merge_end_col=23)
    row += 1

    headers_a = [
        "Date/Time", "Invoice/Receipt ID", "Customer ID", "Product", "Egg Grade Sold",
        "Qty (crates)", "Qty (singles)", "Unit Price", "Payment Method", "Payment Status",
        "Evidence Ref", "MoMo Verified", "Delivery Status", "Dispatch Auth By",
        "Sold By", "Proof Link", "Notes",
        "Total Eggs", "Line Total (GHS)", "Revenue Today",
        "Fraud: Dup Invoice", "Fraud: Paid No Evidence", "Fraud: Price Dev",
    ]

    rows_a = []
    for r in data["sales"]:
        rows_a.append(list(r) + [None, None, None, None, None, None])

    sales_tbl = T("sales")
    calc_a = {
        17: '([@Qty (crates)]*30)+[@Qty (singles)]',
        18: '[@Qty (crates)]*[@Unit Price]',
        19: f'SUMIFS({sales_tbl}[Line Total (GHS)],{sales_tbl}[Date/Time],">="&INT([@Date/Time]),{sales_tbl}[Date/Time],"<"&(INT([@Date/Time])+1))',
        20: f'IF(COUNTIFS({sales_tbl}[Invoice/Receipt ID],[@Invoice/Receipt ID])>1,"RED FLAG","OK")',
        21: 'IF(AND([@Payment Status]="Paid",[@Evidence Ref]=""),"RED FLAG","OK")',
        22: f'IF(ABS([@Unit Price]-{C.PRICES["egg_crate_default"]})/{C.PRICES["egg_crate_default"]}>{C.THRESHOLDS["price_deviation_tolerance_pct"]},"YELLOW FLAG","OK")',
    }

    tab_a, end_a = create_excel_table(
        ws, T("sales"), headers_a, rows_a, start_row=row,
        col_widths=[16, 16, 10, 12, 14, 10, 10, 10, 14, 12, 16, 12, 14, 12, 10, 14, 16,
                    10, 14, 14, 14, 18, 14],
        calculated_columns=calc_a,
        number_formats={7: C.FMT_CURRENCY, 18: C.FMT_CURRENCY, 19: C.FMT_CURRENCY},
    )

    add_dropdown_validation(ws, f"C{row+1}:C{end_a}", MS, _dd_range(master_rows, "customer_master", "A"))
    add_inline_dropdown(ws, f"D{row+1}:D{end_a}", C.DROPDOWN_LISTS["egg_product"])
    add_inline_dropdown(ws, f"E{row+1}:E{end_a}", C.DROPDOWN_LISTS["egg_grade_sold"])
    add_inline_dropdown(ws, f"I{row+1}:I{end_a}", C.DROPDOWN_LISTS["payment_method"])
    add_inline_dropdown(ws, f"J{row+1}:J{end_a}", C.DROPDOWN_LISTS["payment_status"])
    add_inline_dropdown(ws, f"L{row+1}:L{end_a}", ["Yes", "No", "N/A"])
    add_inline_dropdown(ws, f"M{row+1}:M{end_a}", C.DROPDOWN_LISTS["delivery_status"])
    add_dropdown_validation(ws, f"N{row+1}:N{end_a}", MS, _dd_range(master_rows, "staff", "A"))
    add_dropdown_validation(ws, f"O{row+1}:O{end_a}", MS, _dd_range(master_rows, "staff", "A"))

    add_text_status_cf(ws, f"U{row+1}:U{end_a}")
    add_text_status_cf(ws, f"V{row+1}:V{end_a}")
    add_text_status_cf(ws, f"W{row+1}:W{end_a}")

    row = end_a + 3

    # ── Section B: Procurement ────────────────────────────────────────
    write_section_header(ws, row, 1, "PROCUREMENT / EXPENSE LOG", merge_end_col=22)
    row += 1

    headers_b = [
        "Date/Time", "Vendor", "Category", "Item/Service", "Quantity", "Unit",
        "Unit Cost", "Total Cost", "Payment Method", "Payment Ref",
        "Receipt/Invoice ID", "Requested By", "Approval Status", "Approved By",
        "Approval Date", "Received Verified", "Verified By", "Price Benchmark", "Notes",
        "Fraud: Above Threshold", "Fraud: Price Above Benchmark", "Fraud: Dup Receipt",
    ]

    rows_b = []
    for r in data["procurement"]:
        rows_b.append(list(r) + [None, None, None])

    proc_tbl = T("procurement")
    item_tbl = T("item_master")
    calc_b = {
        7: '[@Quantity]*[@Unit Cost]',
        17: f'IFERROR(INDEX({item_tbl}[Reorder Threshold (units)],MATCH([@Item/Service],{item_tbl}[Item Name],0)),"")',
        19: f'IF(AND([@Total Cost]>{C.THRESHOLDS["procurement_approval_threshold"]},[@Approval Status]<>"Approved"),"RED FLAG","OK")',
        20: 'IF(AND(ISNUMBER([@Price Benchmark]),[@Unit Cost]>[@Price Benchmark]*1.15),"YELLOW FLAG","OK")',
        21: f'IF(COUNTIFS({proc_tbl}[Receipt/Invoice ID],[@Receipt/Invoice ID])>1,"RED FLAG","OK")',
    }

    tab_b, end_b = create_excel_table(
        ws, T("procurement"), headers_b, rows_b, start_row=row,
        col_widths=[16, 10, 16, 20, 10, 8, 12, 12, 14, 14, 16, 12, 14, 12, 14, 14, 12, 14, 20,
                    16, 20, 14],
        calculated_columns=calc_b,
        number_formats={6: C.FMT_CURRENCY, 7: C.FMT_CURRENCY, 17: C.FMT_CURRENCY},
    )

    add_dropdown_validation(ws, f"B{row+1}:B{end_b}", MS, _dd_range(master_rows, "vendor", "A"))
    add_inline_dropdown(ws, f"C{row+1}:C{end_b}", C.DROPDOWN_LISTS["procurement_category"])
    add_inline_dropdown(ws, f"I{row+1}:I{end_b}", C.DROPDOWN_LISTS["payment_method"])
    add_dropdown_validation(ws, f"L{row+1}:L{end_b}", MS, _dd_range(master_rows, "staff", "A"))
    add_inline_dropdown(ws, f"M{row+1}:M{end_b}", C.DROPDOWN_LISTS["approval_status"])
    add_dropdown_validation(ws, f"N{row+1}:N{end_b}", MS, _dd_range(master_rows, "staff", "A"))
    add_inline_dropdown(ws, f"P{row+1}:P{end_b}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"Q{row+1}:Q{end_b}", MS, _dd_range(master_rows, "staff", "A"))

    add_text_status_cf(ws, f"T{row+1}:T{end_b}")
    add_text_status_cf(ws, f"U{row+1}:U{end_b}")
    add_text_status_cf(ws, f"V{row+1}:V{end_b}")

    freeze_panes(ws, "A2")
    protect_sheet(ws)
    print("  Tab 5 (Sales & Procurement): 2 tables on single sheet")
    return ws


# ============================================================================
# Tab 5: Events
# ============================================================================

def build_tab5_events(wb, data, master_rows):
    """Tab 5: Events -- 4 tables on one sheet.

    Section A: tblEquipment
    Section B: tblVisitorLog
    Section C: tblHealthIncident
    Section D: tblMedication
    """
    ws = wb.create_sheet(title=C.TAB_NAMES[6])
    row = 1

    # ── Section A: Equipment ──────────────────────────────────────────
    write_section_header(ws, row, 1, "EQUIPMENT & ENVIRONMENTAL STATUS LOG", merge_end_col=17)
    row += 1

    headers_a = [
        "Date/Time", "Equipment", "Location", "Status", "Downtime Hours",
        "Downtime Start", "Downtime End", "Cause Category", "Impact on Birds",
        "Resolution Action", "Resolution Status", "Cost of Repair",
        "Technician", "SC Notified", "Owner Notified", "Notes",
        "Uptime %",
    ]

    rows_a = []
    for r in data["equipment"]:
        rows_a.append(list(r) + [None])

    calc_a = {16: 'IF([@Downtime Hours]=0,1,(24-[@Downtime Hours])/24)'}

    tab_a, end_a = create_excel_table(
        ws, T("equipment"), headers_a, rows_a, start_row=row,
        col_widths=[16, 22, 12, 8, 12, 14, 14, 18, 14, 20, 14, 12, 16, 10, 12, 20, 10],
        calculated_columns=calc_a,
        number_formats={4: C.FMT_DECIMAL_1, 11: C.FMT_CURRENCY, 16: C.FMT_PERCENT},
    )

    add_inline_dropdown(ws, f"B{row+1}:B{end_a}", C.DROPDOWN_LISTS["equipment_list"])
    add_inline_dropdown(ws, f"D{row+1}:D{end_a}", C.DROPDOWN_LISTS["equipment_status"])
    add_inline_dropdown(ws, f"H{row+1}:H{end_a}", C.DROPDOWN_LISTS["cause_category_equip"])
    add_inline_dropdown(ws, f"I{row+1}:I{end_a}", C.DROPDOWN_LISTS["impact_on_birds"])
    add_inline_dropdown(ws, f"K{row+1}:K{end_a}", C.DROPDOWN_LISTS["resolution_status"])
    add_inline_dropdown(ws, f"M{row+1}:M{end_a}", C.DROPDOWN_LISTS["technician"])
    add_inline_dropdown(ws, f"N{row+1}:N{end_a}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"O{row+1}:O{end_a}", C.DROPDOWN_LISTS["yes_no"])

    add_text_status_cf(ws, f"D{row+1}:D{end_a}")

    row = end_a + 3

    # ── Section B: Visitor Log ────────────────────────────────────────
    write_section_header(ws, row, 1, "VISITOR & VEHICLE LOG", merge_end_col=11)
    row += 1

    headers_b = [
        "Date/Time In", "Date/Time Out", "Visitor/Driver Name", "Vehicle (License)",
        "Company", "Reason for Visit", "Zone Accessed",
        "Disinfection Done", "Approved By", "Camera Ref", "Notes",
    ]

    rows_b = data["visitor"]

    tab_b, end_b = create_excel_table(
        ws, T("visitor_log"), headers_b, rows_b, start_row=row,
        col_widths=[16, 16, 20, 16, 16, 14, 20, 14, 12, 12, 20],
    )

    add_inline_dropdown(ws, f"F{row+1}:F{end_b}", C.DROPDOWN_LISTS["visitor_reason"])
    add_inline_dropdown(ws, f"G{row+1}:G{end_b}", C.DROPDOWN_LISTS["zone_access"])
    add_inline_dropdown(ws, f"H{row+1}:H{end_b}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"I{row+1}:I{end_b}", MS, _dd_range(master_rows, "staff", "A"))

    row = end_b + 3

    # ── Section C: Health Incident ────────────────────────────────────
    write_section_header(ws, row, 1, "HEALTH INCIDENT LOG", merge_end_col=14)
    row += 1

    headers_c = [
        "Date/Time", "Cage/Cohort", "Symptom Category", "Specific Symptoms",
        "Severity (1-5)", "Birds Affected", "Photo Evidence",
        "FHIS Consultation", "FHIS Recommendation", "Action Taken",
        "Vet Called", "Vet Response", "Resolution Status", "Notes",
    ]

    rows_c = data["health_incident"]

    tab_c, end_c = create_excel_table(
        ws, T("health_incident"), headers_c, rows_c, start_row=row,
        col_widths=[16, 12, 16, 30, 12, 12, 12, 14, 20, 14, 10, 20, 14, 20],
    )

    add_dropdown_validation(ws, f"B{row+1}:B{end_c}", MS, _dd_range(master_rows, "housing", "A"))
    add_inline_dropdown(ws, f"C{row+1}:C{end_c}", C.DROPDOWN_LISTS["symptom_category"])
    add_inline_dropdown(ws, f"E{row+1}:E{end_c}", C.DROPDOWN_LISTS["severity_1_5"])
    add_inline_dropdown(ws, f"G{row+1}:G{end_c}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"H{row+1}:H{end_c}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"J{row+1}:J{end_c}", C.DROPDOWN_LISTS["health_action"])
    add_inline_dropdown(ws, f"K{row+1}:K{end_c}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"M{row+1}:M{end_c}", C.DROPDOWN_LISTS["health_resolution"])

    row = end_c + 3

    # ── Section D: Medication ─────────────────────────────────────────
    write_section_header(ws, row, 1, "MEDICATION & VACCINATION LOG", merge_end_col=14)
    row += 1

    headers_d = [
        "Date", "Type", "Product Name", "Dose/Concentration",
        "Route", "Cohort/Cage", "Birds Treated",
        "Administered By", "Vet Authorized", "Batch/Lot #",
        "Withdrawal Period (days)", "Next Dose Due", "Notes",
        "Withdrawal End Date",
    ]

    rows_d = []
    for r in data["medication"]:
        rows_d.append(list(r) + [None])

    calc_d = {
        13: 'IF([@Withdrawal Period (days)]>0,[@Date]+[@Withdrawal Period (days)],"")',
    }

    tab_d, end_d = create_excel_table(
        ws, T("medication"), headers_d, rows_d, start_row=row,
        col_widths=[12, 14, 14, 18, 14, 12, 12, 14, 12, 14, 18, 14, 20, 16],
        calculated_columns=calc_d,
    )

    add_inline_dropdown(ws, f"B{row+1}:B{end_d}", C.DROPDOWN_LISTS["med_type"])
    add_dropdown_validation(ws, f"C{row+1}:C{end_d}", MS, _dd_range(master_rows, "item_master", "A"))
    add_inline_dropdown(ws, f"E{row+1}:E{end_d}", C.DROPDOWN_LISTS["med_route"])
    add_dropdown_validation(ws, f"F{row+1}:F{end_d}", MS, _dd_range(master_rows, "flock", "A"))
    add_dropdown_validation(ws, f"H{row+1}:H{end_d}", MS, _dd_range(master_rows, "staff", "A"))
    add_inline_dropdown(ws, f"I{row+1}:I{end_d}", C.DROPDOWN_LISTS["yes_no"])

    freeze_panes(ws, "A2")
    protect_sheet(ws)
    print("  Tab 6 (Events): 4 tables on single sheet")
    return ws


# ============================================================================
# Public entry point -- called by main.py
# ============================================================================

def build_inputs(wb, data, master_rows):
    """Build all four input tabs (Tabs 2-5)."""
    build_tab2_daily_log(wb, data, master_rows)
    build_tab3_feed_inventory(wb, data, master_rows)
    build_tab4_sales_procurement(wb, data, master_rows)
    build_tab5_events(wb, data, master_rows)
