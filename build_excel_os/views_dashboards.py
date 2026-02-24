"""
views_dashboards.py — Tabs 39-40: Owner Dashboard + Manager Dashboard.
Read-only views pulling from all layers. Exception-driven: Green=ignore, Yellow=investigate, Red=act.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import config as C
from .helpers import (
    protect_sheet, freeze_panes, write_section_header,
    add_text_status_cf, set_col_widths,
)

TN = C.TABLE_NAMES
TAN = C.TAB_NAMES

DASH_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
DASH_SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="4472C4")
KPI_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
KPI_VALUE_FONT = Font(name="Calibri", size=14, bold=True)


def _write_kpi_row(ws, row, col, label, formula, target_text, status_formula, fmt=None):
    """Write a single KPI row: Label | Value | Target | Status."""
    ws.cell(row=row, column=col, value=label).font = KPI_LABEL_FONT
    cell_val = ws.cell(row=row, column=col + 1)
    cell_val.value = formula
    cell_val.font = KPI_VALUE_FONT
    if fmt:
        cell_val.number_format = fmt
    ws.cell(row=row, column=col + 2, value=target_text)
    ws.cell(row=row, column=col + 3, value=status_formula)


def build_tab39_owner_dashboard(wb):
    """Tab 39: Owner Dashboard — exception-driven, <1 hour/week review."""
    ws = wb.create_sheet(title=TAN[39])

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"].value = "ULTIMATE FARMS — OWNER DASHBOARD"
    ws["A1"].font = DASH_TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"].value = 'Exception-driven: Green = ignore | Yellow = investigate | Red = act now'
    ws["A2"].font = DASH_SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    prod = TN[1]
    feed = TN[4]
    mort = TN[3]
    sales = TN[8]
    equip = TN[10]
    bio = TN[11]
    labor = TN[15]
    flock = TN[20]
    target = TN[28]
    recon_eggs = TN[29]
    recon_cash = TN[30]
    fraud = TN[32]

    r = 4
    # ─── SECTION A: BIOLOGICAL (Age-Adjusted) ───
    write_section_header(ws, r, 1, "A — BIOLOGICAL PERFORMANCE (Age-Adjusted)", merge_end_col=6)
    r += 1
    headers = ["Metric", "Current", "Target", "Phase", "7d Trend", "Status"]
    for ci, h in enumerate(headers):
        cell = ws.cell(row=r, column=ci + 1, value=h)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
        cell.alignment = C.HEADER_ALIGN
    r += 1

    bio_kpis = [
        ("Lay % (FL2024A)",
         f'=IFERROR(AVERAGEIFS({prod}[Large %],{prod}[Date],">="&(TODAY()-6)),0)',
         f'=IFERROR(INDEX({target}[Effective Target: Lay %],MATCH("FL2024A",{target}[Cohort ID],0)),"N/A")',
         f'=IFERROR(INDEX({target}[Production Phase],MATCH("FL2024A",{target}[Cohort ID],0)),"N/A")',
         "→", None),
        ("Lay % (FL2024B)",
         f'=IFERROR(AVERAGEIFS({prod}[Large %],{prod}[Date],">="&(TODAY()-6)),0)',
         f'=IFERROR(INDEX({target}[Effective Target: Lay %],MATCH("FL2024B",{target}[Cohort ID],0)),"N/A")',
         f'=IFERROR(INDEX({target}[Production Phase],MATCH("FL2024B",{target}[Cohort ID],0)),"N/A")',
         "→", None),
        ("FCR (Overall)",
         f'=IFERROR(SUM({feed}[Net Consumed (kg)])/(SUM({prod}[Total Eggs])/12),0)',
         f'=IFERROR(INDEX({target}[Effective Target: FCR],1),"2.0")',
         "", "→", None),
        ("Total Mortality (today)",
         f'=IFERROR(SUMIFS({mort}[Death Count],{mort}[Date],TODAY()),0)',
         f'="<{C.THRESHOLDS["mortality_daily_yellow"]}"',
         "", "", None),
        ("Disease Mortality (today)",
         f'=IFERROR(SUMIFS({mort}[Death Count],{mort}[Date],TODAY(),{mort}[Cause Category],"Disease"),0)',
         f'="<{C.THRESHOLDS["disease_mortality_immediate_red"]}"',
         "", "", None),
        ("Cracked %",
         f'=IFERROR(SUMIFS({prod}[Grade: Cracked/Broken],{prod}[Date],TODAY())/SUMIFS({prod}[Total Eggs],{prod}[Date],TODAY()),0)',
         f'="<{C.THRESHOLDS["cracked_pct_yellow"]*100}%"',
         "", "→", None),
        ("Large %",
         f'=IFERROR(SUMIFS({prod}[Grade: Large],{prod}[Date],TODAY())/SUMIFS({prod}[Total Eggs],{prod}[Date],TODAY()),0)',
         f'=">60%"',
         "", "→", None),
    ]

    for kpi in bio_kpis:
        for ci, val in enumerate(kpi):
            cell = ws.cell(row=r, column=ci + 1, value=val)
            if ci == 1:
                cell.font = KPI_VALUE_FONT
        # Status formula
        ws.cell(row=r, column=6, value='=IF(ISNUMBER(B{r}),IF(B{r}>=C{r},"Green",IF(B{r}>=C{r}*0.95,"Yellow","Red")),"N/A")'.format(r=r))
        r += 1

    add_text_status_cf(ws, f"F5:F{r-1}")

    # ─── SECTION B: OPERATIONAL ───
    r += 1
    write_section_header(ws, r, 1, "B — OPERATIONAL STATUS", merge_end_col=6)
    r += 1
    op_headers = ["Metric", "Current", "Target", "Status"]
    for ci, h in enumerate(op_headers):
        ws.cell(row=r, column=ci + 1, value=h).font = C.HEADER_FONT
        ws.cell(row=r, column=ci + 1).fill = C.HEADER_FILL
    r += 1

    op_kpis = [
        ("Feed Days on Hand", '=7', ">=7 days", '=IF(B{r}>=7,"Green",IF(B{r}>=3,"Yellow","Red"))'),
        ("Egg Stock (crates)",
         f'=IFERROR(SUMIFS({prod}[Total Eggs],{prod}[Date],TODAY())/30,0)',
         "<=500", '=IF(B{r}<=500,"Green",IF(B{r}<=1000,"Yellow","Red"))'),
        ("Equipment Status",
         f'=IFERROR(COUNTIFS({equip}[Status],"Red"),"0")',
         "0 Red items", '=IF(B{r}=0,"Green",IF(B{r}<=1,"Yellow","Red"))'),
        ("Biosecurity Compliance",
         f'=IFERROR(AVERAGEIFS({bio}[Compliance Score %],{bio}[Date],">="&(TODAY()-6)),1)',
         "100%", '=IF(B{r}>=0.95,"Green",IF(B{r}>=0.8,"Yellow","Red"))'),
        ("Team Attendance",
         f'=IFERROR(COUNTIFS({labor}[Date],TODAY(),{labor}[Present],"Yes"),0)',
         f'>="{C.THRESHOLDS["attendance_min_crew"]}"',
         '=IF(B{r}>={thresh},"Green","Red")'.format(r='{r}', thresh=C.THRESHOLDS["attendance_min_crew"])),
    ]

    for label, formula, target, status in op_kpis:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=formula).font = KPI_VALUE_FONT
        ws.cell(row=r, column=3, value=target)
        ws.cell(row=r, column=4, value=status.format(r=r))
        r += 1

    add_text_status_cf(ws, f"D{r-5}:D{r-1}")

    # ─── SECTION C: FINANCIAL ───
    r += 1
    write_section_header(ws, r, 1, "C — FINANCIAL (Month-to-Date)", merge_end_col=6)
    r += 1
    for ci, h in enumerate(["Metric", "Current", "Target", "Status"]):
        ws.cell(row=r, column=ci + 1, value=h).font = C.HEADER_FONT
        ws.cell(row=r, column=ci + 1).fill = C.HEADER_FILL
    r += 1

    fin_kpis = [
        ("Revenue MTD",
         f'=SUMIFS({sales}[Line Total (GHS)],{sales}[Date/Time],">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),{sales}[Date/Time],"<="&TODAY())',
         "Prior month pace", ""),
        ("Cash Recon Variance (cumul.)",
         f'=IFERROR(SUM({recon_cash}[Daily Revenue Variance]),0)',
         "Zero", '=IF(ABS(B{r})<100,"Green",IF(ABS(B{r})<500,"Yellow","Red"))'),
        ("Outstanding Receivables",
         f'=IFERROR(SUM({sales}[Line Total (GHS)])-SUM({recon_cash}[Total Payments]),0)',
         "<5000 GHS", '=IF(B{r}<5000,"Green",IF(B{r}<10000,"Yellow","Red"))'),
    ]

    for label, formula, target, status in fin_kpis:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        cell = ws.cell(row=r, column=2, value=formula)
        cell.font = KPI_VALUE_FONT
        cell.number_format = C.FMT_CURRENCY
        ws.cell(row=r, column=3, value=target)
        if status:
            ws.cell(row=r, column=4, value=status.format(r=r))
        r += 1

    add_text_status_cf(ws, f"D{r-3}:D{r-1}")

    # ─── SECTION D: FRAUD & APPROVALS ───
    r += 1
    write_section_header(ws, r, 1, "D — FRAUD FLAGS & PENDING ACTIONS", merge_end_col=6)
    r += 1
    for ci, h in enumerate(["Item", "Count", "Action"]):
        ws.cell(row=r, column=ci + 1, value=h).font = C.HEADER_FONT
        ws.cell(row=r, column=ci + 1).fill = C.HEADER_FILL
    r += 1

    fraud_items = [
        ("Open Red Fraud Flags",
         f'=COUNTIFS({fraud}[Severity],"Red",{fraud}[Status],"Open")',
         "Investigate immediately"),
        ("Open Yellow Fraud Flags",
         f'=COUNTIFS({fraud}[Severity],"Yellow",{fraud}[Status],"Open")',
         "Review this week"),
        ("Pending Procurement Approvals",
         f'=COUNTIFS({TN[9]}[Approval Status],"Pending")',
         "Approve or reject"),
        ("Unresolved Health Incidents",
         f'=COUNTIFS({TN[13]}[Resolution Status],"Open")',
         "Follow up with vet/SC"),
    ]

    for label, formula, action in fraud_items:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=formula).font = KPI_VALUE_FONT
        ws.cell(row=r, column=3, value=action)
        r += 1

    # ─── SECTION E: NEXT ACTIONS ───
    r += 1
    write_section_header(ws, r, 1, "E — NEXT ACTIONS (Priority-Ordered)", merge_end_col=6)
    r += 1
    actions = [
        "1. Review cull recommendations (Prediction Engine → Flock Registry)",
        "2. Check feed/ingredient reorder dates (Prediction Engine)",
        "3. Upcoming vaccinations due (Medication Log → next dose)",
        "4. Vet-call trigger forecast (Prediction Engine → mortality slope)",
        "5. Equipment maintenance due (Equipment Log → patterns)",
        "6. Churn-risk customers to contact (CRM → Red churn risk)",
        "7. Cycle counts scheduled for today (Scheduler)",
    ]
    for action in actions:
        ws.cell(row=r, column=1, value=action)
        r += 1

    # Formatting
    set_col_widths(ws, [30, 16, 18, 16, 10, 10])
    freeze_panes(ws, "A4")
    protect_sheet(ws)

    print("  Tab 39: Owner Dashboard created")
    return ws


def build_tab40_manager_dashboard(wb):
    """Tab 40: Manager Dashboard — daily ops for Site Coordinator."""
    ws = wb.create_sheet(title=TAN[40])

    ws.merge_cells("A1:F1")
    ws["A1"].value = "ULTIMATE FARMS — MANAGER DASHBOARD (Daily Ops)"
    ws["A1"].font = DASH_TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"].value = "Site Coordinator daily workflow — check completeness, review exceptions, plan actions"
    ws["A2"].font = DASH_SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    prod = TN[1]
    env = TN[2]
    mort = TN[3]
    feed = TN[4]
    water = TN[5]
    bio = TN[11]
    labor = TN[15]
    recon_eggs = TN[29]
    recon_cash = TN[30]
    recon_feed = TN[31]
    equip = TN[10]
    health = TN[13]

    r = 4
    # ─── SECTION A: DATA COMPLETENESS ───
    write_section_header(ws, r, 1, "A — DATA COMPLETENESS (Did everyone log everything today?)", merge_end_col=4)
    r += 1
    for ci, h in enumerate(["Check", "Status", "Missing"]):
        ws.cell(row=r, column=ci + 1, value=h).font = C.HEADER_FONT
        ws.cell(row=r, column=ci + 1).fill = C.HEADER_FILL
    r += 1

    completeness = [
        ("Production Log today",
         f'=IF(COUNTIFS({prod}[Date],TODAY())>={len(C.HOUSING_DATA)},"Complete","MISSING")',
         f'=IF(COUNTIFS({prod}[Date],TODAY())>={len(C.HOUSING_DATA)},"",{len(C.HOUSING_DATA)}-COUNTIFS({prod}[Date],TODAY())&" cage entries missing")'),
        ("Environmental Log today",
         f'=IF(COUNTIFS({env}[Date],TODAY())>=5,"Complete","MISSING")',
         f'=IF(COUNTIFS({env}[Date],TODAY())>=5,"","Need readings for all houses")'),
        ("Mortality Log today",
         f'=IF(COUNTIFS({mort}[Date],TODAY())>=0,"Logged","Check")',
         '=""'),
        ("Feed Consumption today",
         f'=IF(COUNTIFS({feed}[Date],TODAY())>=6,"Complete","MISSING")',
         f'=IF(COUNTIFS({feed}[Date],TODAY())>=6,"","Missing feeding rounds")'),
        ("Water Consumption today",
         f'=IF(COUNTIFS({water}[Date],TODAY())>=5,"Complete","MISSING")',
         '=""'),
        ("Biosecurity Checklist today",
         f'=IF(COUNTIFS({bio}[Date],TODAY())>=1,"Done","NOT DONE")',
         '=""'),
        ("Labor Log today",
         f'=IF(COUNTIFS({labor}[Date],TODAY())>=6,"Complete","MISSING")',
         f'=IF(COUNTIFS({labor}[Date],TODAY())>=6,"","Missing staff entries")'),
    ]

    for label, status, missing in completeness:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=status)
        ws.cell(row=r, column=3, value=missing)
        r += 1

    add_text_status_cf(ws, f"B{r-7}:B{r-1}")

    # ─── SECTION B: TODAY'S CYCLE COUNTS ───
    r += 1
    write_section_header(ws, r, 1, "B — TODAY'S CYCLE COUNTS (from Scheduler)", merge_end_col=4)
    r += 1
    ws.cell(row=r, column=1, value="→ Check Tab 38 (Cycle Count Scheduler) for today's count list")
    r += 1

    # ─── SECTION C: OPEN ITEMS ───
    r += 1
    write_section_header(ws, r, 1, "C — OPEN ITEMS REQUIRING ATTENTION", merge_end_col=4)
    r += 1
    for ci, h in enumerate(["Item", "Count", "Action"]):
        ws.cell(row=r, column=ci + 1, value=h).font = C.HEADER_FONT
        ws.cell(row=r, column=ci + 1).fill = C.HEADER_FILL
    r += 1

    open_items = [
        ("Open Health Incidents",
         f'=COUNTIFS({health}[Resolution Status],"Open")',
         "Follow up / escalate"),
        ("Equipment at Yellow/Red",
         f'=COUNTIFS({equip}[Status],"Yellow")+COUNTIFS({equip}[Status],"Red")',
         "Schedule repair / notify owner"),
        ("Pending Procurement",
         f'=COUNTIFS({TN[9]}[Approval Status],"Pending")',
         "Submit to owner for approval"),
    ]

    for label, formula, action in open_items:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=formula).font = KPI_VALUE_FONT
        ws.cell(row=r, column=3, value=action)
        r += 1

    # ─── SECTION D: YESTERDAY'S RECONCILIATION ───
    r += 1
    write_section_header(ws, r, 1, "D — YESTERDAY'S RECONCILIATION STATUS", merge_end_col=4)
    r += 1
    for ci, h in enumerate(["Reconciliation", "Variance", "Status"]):
        ws.cell(row=r, column=ci + 1, value=h).font = C.HEADER_FONT
        ws.cell(row=r, column=ci + 1).fill = C.HEADER_FILL
    r += 1

    recons = [
        ("Egg Reconciliation",
         f'=IFERROR(SUMIFS({recon_eggs}[Variance (crates)],{recon_eggs}[Date],TODAY()-1),"N/A")',
         '=IF(ISNUMBER(B{r}),IF(ABS(B{r})<=0.5,"Green",IF(ABS(B{r})<=2,"Yellow","Red")),"N/A")'),
        ("Cash Reconciliation",
         f'=IFERROR(SUMIFS({recon_cash}[Daily Revenue Variance],{recon_cash}[Date],TODAY()-1),"N/A")',
         '=IF(ISNUMBER(B{r}),IF(ABS(B{r})<100,"Green",IF(ABS(B{r})<500,"Yellow","Red")),"N/A")'),
        ("Feed Reconciliation",
         f'=IFERROR(SUMIFS({recon_feed}[Discrepancy (kg)],{recon_feed}[Date],TODAY()-1),"N/A")',
         '=IF(ISNUMBER(B{r}),IF(ABS(B{r})<25,"Green",IF(ABS(B{r})<50,"Yellow","Red")),"N/A")'),
    ]

    for label, formula, status in recons:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=formula).font = KPI_VALUE_FONT
        ws.cell(row=r, column=3, value=status.format(r=r))
        r += 1

    add_text_status_cf(ws, f"C{r-3}:C{r-1}")

    # ─── SECTION E: HANDOVER PREP ───
    r += 1
    write_section_header(ws, r, 1, "E — WEEKLY HANDOVER PREP (Shows on handover day)", merge_end_col=4)
    r += 1
    ws.cell(row=r, column=1, value="• Verify all inventory totals before handover")
    ws.cell(row=r + 1, column=1, value="• Document equipment status for incoming team")
    ws.cell(row=r + 2, column=1, value="• Brief on outstanding issues and pending actions")
    ws.cell(row=r + 3, column=1, value="• Ensure all logs signed off for the rotation period")

    set_col_widths(ws, [30, 16, 30, 16])
    freeze_panes(ws, "A4")
    protect_sheet(ws)

    print("  Tab 40: Manager Dashboard created")
    return ws


def build_dashboards(wb):
    build_tab39_owner_dashboard(wb)
    build_tab40_manager_dashboard(wb)
