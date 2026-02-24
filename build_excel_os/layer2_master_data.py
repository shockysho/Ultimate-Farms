"""
layer2_master_data.py — Build 11 master data tabs (Layer 2).
These are locked reference tables that populate dropdowns and lookups across the workbook.
Built FIRST because Layer 1 inputs reference them.
"""

from datetime import date
from openpyxl.utils import get_column_letter

from . import config as C
from .helpers import (
    create_excel_table, add_inline_dropdown, protect_sheet,
    freeze_panes, set_col_widths, write_section_header, create_named_range,
    add_text_status_cf,
)


def build_tab17_config(wb):
    """Tab 17: Config & Thresholds — single source of truth for all system parameters."""
    ws = wb.create_sheet(title=C.TAB_NAMES[17])

    headers = ["Section", "Parameter", "Value", "Unit", "Description"]
    rows = [
        # Global parameters
        ("Global", "Farm Name", C.FARM_NAME, "", ""),
        ("Global", "Farm Location", C.FARM_LOCATION, "", ""),
        ("Global", "Currency", C.CURRENCY, "", ""),
        ("Global", "Crate Size", C.CRATE_SIZE, "eggs", "Standard crate = 30 eggs"),
        ("Global", "Feeding Schedule AM", 0.40, "%", "40% morning feed"),
        ("Global", "Feeding Schedule PM", 0.30, "%", "30% afternoon feed"),
        ("Global", "Feeding Schedule EVE", 0.30, "%", "30% evening feed"),
        ("Global", "Target Intake Per Bird", 118, "g/day", "Default 118g (range 115-120g)"),
        ("Global", "Working Days Per Week", 6, "days", ""),
        # Alert thresholds
        ("Alert", "Mortality Daily Yellow", C.THRESHOLDS["mortality_daily_yellow"], "birds", "4-6 birds"),
        ("Alert", "Mortality Daily Red", C.THRESHOLDS["mortality_daily_red"], "birds", "> 6 birds OR cluster"),
        ("Alert", "Lay Rate Yellow", C.THRESHOLDS["lay_rate_yellow"], "decimal", "< 88% for 2 days"),
        ("Alert", "Lay Rate Red", C.THRESHOLDS["lay_rate_red"], "decimal", "< 85% for 2 consecutive days"),
        ("Alert", "Cash Mismatch Yellow (crates)", C.THRESHOLDS["cash_mismatch_yellow_crates"], "crates", "1-2 crates discrepancy"),
        ("Alert", "Cash Mismatch Red (crates)", C.THRESHOLDS["cash_mismatch_red_crates"], "crates", "> 2 crates OR any cash gap"),
        ("Alert", "Feed Stock Yellow (days)", C.THRESHOLDS["feed_stock_yellow_days"], "days", "3-7 days on hand"),
        ("Alert", "Feed Stock Red (days)", C.THRESHOLDS["feed_stock_red_days"], "days", "< 3 days on hand"),
        ("Alert", "Egg Stock Yellow (crates)", C.THRESHOLDS["egg_stock_yellow_crates"], "crates", "500-1000 crates"),
        ("Alert", "Egg Stock Red (crates)", C.THRESHOLDS["egg_stock_red_crates"], "crates", "> 1000 crates"),
        ("Alert", "Equipment Uptime Yellow", C.THRESHOLDS["equipment_uptime_yellow"], "decimal", "80-90%"),
        ("Alert", "Equipment Uptime Red", C.THRESHOLDS["equipment_uptime_red"], "decimal", "< 80%"),
        ("Alert", "Cracked % Yellow", C.THRESHOLDS["cracked_pct_yellow"], "decimal", "> 6% for 2 consecutive days"),
        ("Alert", "Large % Yellow (Peak)", C.THRESHOLDS["large_pct_yellow_peak"], "decimal", "< 60% for 3 days"),
        ("Alert", "Large % Yellow (Post-peak)", C.THRESHOLDS["large_pct_yellow_postpeak"], "decimal", "< 55% for 3 days"),
        ("Alert", "FCR Yellow (Peak)", C.THRESHOLDS["fcr_peak_yellow"], "kg/dz", "> 2.2 for 3 days"),
        ("Alert", "FCR Red (Peak)", C.THRESHOLDS["fcr_peak_red"], "kg/dz", "> 2.5 for 3 days"),
        ("Alert", "Disease Mortality Immediate Red", C.THRESHOLDS["disease_mortality_immediate_red"], "birds/day", "Absolute trigger"),
        ("Alert", "Heat Stress Hours Red", C.THRESHOLDS["heat_stress_hours_red"], "hours", "> 2 hrs in any zone"),
        ("Alert", "Heat Stress Day Yellow", C.THRESHOLDS["heat_stress_hours_day_yellow"], "hours/day", "> 6 cumulative hours"),
        ("Alert", "Water Drop % Yellow", C.THRESHOLDS["water_drop_pct_yellow"], "decimal", "> 20% drop vs 7-day avg"),
        ("Alert", "Water Spike % Yellow", C.THRESHOLDS["water_spike_pct_yellow"], "decimal", "> 30% spike vs 7-day avg"),
        ("Alert", "Biosecurity Score Yellow", C.THRESHOLDS["biosecurity_score_yellow"], "decimal", "< 80% compliance"),
        ("Alert", "Min Crew (Attendance Red)", C.THRESHOLDS["attendance_min_crew"], "staff", "< 6 present"),
        # Fraud detection
        ("Fraud", "Price Deviation Tolerance", C.THRESHOLDS["price_deviation_tolerance_pct"], "decimal", "15% tolerance"),
        ("Fraud", "Procurement Approval Threshold", C.THRESHOLDS["procurement_approval_threshold"], "GHS", "Above this = requires approval"),
        ("Fraud", "Cash Deposit Window", C.THRESHOLDS["cash_deposit_window_hours"], "hours", "Cash must be deposited within this"),
        ("Fraud", "Feed Discrepancy Yellow (bags)", C.THRESHOLDS["feed_discrepancy_bags_yellow"], "bags", "> 0.5 bags/day"),
        ("Fraud", "Ingredient Shrinkage Yellow", C.THRESHOLDS["ingredient_shrinkage_yellow"], "decimal", "> 2% over 7 days"),
        ("Fraud", "Ingredient Shrinkage Red", C.THRESHOLDS["ingredient_shrinkage_red"], "decimal", "> 5%"),
        # Cycle count
        ("CycleCount", "Class A Frequency (days)", 7, "days", "High-value items"),
        ("CycleCount", "Class B Frequency (days)", 14, "days", "Moderate-value items"),
        ("CycleCount", "Class C Frequency (days)", 30, "days", "Low-value items"),
        ("CycleCount", "Variance Threshold %", 0.05, "decimal", "Flag if > 5%"),
        # Cull logic
        ("Cull", "Late-lay Start Week", 61, "weeks", ""),
        ("Cull", "Consecutive Below-target Days", 14, "days", "Trigger for prepare status"),
        ("Cull", "Margin Floor Weeks", 4, "weeks", "Below floor for N weeks"),
        ("Cull", "Mortality Slope Threshold", 0.002, "rate", "Sustained slope trigger"),
    ]

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[17], headers, rows,
                                       col_widths=[12, 35, 12, 10, 45])
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab18_staff(wb):
    """Tab 18: Staff Registry."""
    ws = wb.create_sheet(title=C.TAB_NAMES[18])

    headers = ["Staff ID", "Name", "Role", "Team", "Hire Date", "Phone", "Status"]
    rows = [[s[0], s[1], s[2], s[3], s[4], s[5], s[6]] for s in C.STAFF_DATA]

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[18], headers, rows,
                                       col_widths=[10, 20, 18, 14, 12, 16, 10])
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab19_housing(wb):
    """Tab 19: Housing Map."""
    ws = wb.create_sheet(title=C.TAB_NAMES[19])

    headers = ["Cage ID", "House", "Zone", "Row", "Side", "Bird Count",
               "Cohort ID", "Active", "Total Birds", "Equipment Status"]
    rows = []
    for h in C.HOUSING_DATA:
        rows.append([h[0], h[1], h[2], h[3], h[4], h[5], h[6], h[7], None, "Green"])

    # Total Birds is a formula summing bird counts per house
    calculated = {8: 'SUMIFS([Bird Count],[House],[@ House],[Active],"Yes")'}

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[19], headers, rows,
                                       col_widths=[10, 12, 10, 6, 6, 12, 12, 8, 12, 14],
                                       calculated_columns=calculated)
    add_text_status_cf(ws, f"J2:J{end_row}")
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab20_flock(wb):
    """Tab 20: Flock Registry."""
    ws = wb.create_sheet(title=C.TAB_NAMES[20])

    headers = ["Cohort ID", "Breed", "Arrival Date", "Initial Count", "Status",
               "Current Bird Count", "Current Age (weeks)", "Production Phase",
               "Projected Cull Start", "Projected Cull End", "Cull Status", "Cull Triggers"]
    rows = []
    for f in C.FLOCK_DATA:
        rows.append([
            f[0], f[1], f[2], f[3], f[4],
            None, None, None,  # Formulas
            None, None, "Monitor", ""
        ])

    calculated = {
        5: f'[@Initial Count]-SUMIFS({C.TABLE_NAMES[3]}[Death Count],{C.TABLE_NAMES[3]}[Cage ID],"*")',
        6: 'INT((TODAY()-[@Arrival Date])/7)',
        7: 'IF([@Current Age (weeks)]<25,"Ramp-up",IF([@Current Age (weeks)]<=45,"Peak",IF([@Current Age (weeks)]<=60,"Post-peak","Late-lay")))',
    }

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[20], headers, rows,
                                       col_widths=[12, 22, 12, 12, 10, 16, 16, 14, 14, 14, 12, 20],
                                       calculated_columns=calculated)
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab21_breeder_curves(wb):
    """Tab 21: Breeder Curve Tables — Lohmann Brown Classic performance by age."""
    ws = wb.create_sheet(title=C.TAB_NAMES[21])

    headers = ["Week", "Expected Lay %", "Expected Egg Wt (g)", "Expected Large %",
               "Expected Feed Intake (g/bird/day)", "Expected FCR (kg/dozen)",
               "Mortality Band (% monthly)", "Phase"]
    rows = []
    for entry in C.BREEDER_CURVE:
        rows.append(list(entry))

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[21], headers, rows,
                                       col_widths=[8, 14, 18, 14, 28, 22, 22, 12],
                                       number_formats={
                                           1: C.FMT_PERCENT,
                                           3: C.FMT_PERCENT,
                                           5: C.FMT_DECIMAL_2,
                                           6: C.FMT_PERCENT,
                                       })
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab22_owner_overrides(wb):
    """Tab 22: Owner Target Overrides."""
    ws = wb.create_sheet(title=C.TAB_NAMES[22])

    headers = ["Override ID", "Scope", "Cohort ID", "Breed", "Metric",
               "Start Week", "End Week", "Override Target", "Min Band", "Max Band",
               "Reason", "Set By", "Set Date"]
    # Sample: owner adjusts lay target for older flock
    rows = [
        [1, "Cohort-specific", "FL2024A", "Lohmann Brown Classic", "Lay %",
         45, 60, 0.88, 0.85, 0.92,
         "Adjusted for Ghana heat conditions", "Owner", "2025-12-01"],
        [2, "Breed-wide", "", "Lohmann Brown Classic", "Feed intake",
         25, 80, 120, 115, 125,
         "Tropical climate adjustment", "Owner", "2025-12-01"],
    ]

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[22], headers, rows,
                                       col_widths=[10, 14, 12, 22, 12, 10, 10, 14, 10, 10, 30, 10, 12])
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab23_customer_crm(wb):
    """Tab 23: Customer CRM."""
    ws = wb.create_sheet(title=C.TAB_NAMES[23])

    headers = ["Customer ID", "Name", "Type", "Contact Person", "Phone",
               "Location", "Credit Allowed", "Credit Limit (GHS)",
               "Typical Order (crates/wk)", "Reliability (1-5)",
               "Price Sensitivity (1-5)", "Preferred Payment",
               "Churn Risk", "Customer Since",
               "Last Order Date", "Lifetime Crates", "Notes"]
    rows = []
    for c in C.CUSTOMER_DATA:
        rows.append([
            c[0], c[1], c[2], c[3], c[4], c[5],
            c[6], c[7], c[8], c[9], c[10], c[11],
            None, "2024-01-15",  # Churn risk = formula, Customer Since
            None, None, ""  # Last order, lifetime = formulas
        ])

    calculated = {
        12: 'IF([@Last Order Date]="","N/A",IF(TODAY()-[@Last Order Date]>14,"Red",IF(TODAY()-[@Last Order Date]>7,"Yellow","Green")))',
    }

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[23], headers, rows,
                                       col_widths=[10, 22, 14, 16, 16, 18, 12, 14, 18, 12, 14, 16, 10, 14, 14, 14, 20],
                                       calculated_columns=calculated,
                                       number_formats={7: C.FMT_CURRENCY})

    add_text_status_cf(ws, f"M2:M{end_row}")
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab24_customer_profile(wb):
    """Tab 24: Customer Profile (single-view lookup)."""
    ws = wb.create_sheet(title=C.TAB_NAMES[24])

    headers = ["Customer ID", "Name", "Type", "Location",
               "Avg Order (4wk)", "Order Frequency (/month)",
               "Volume Trend", "Outstanding Balance (GHS)",
               "Payment Reliability %", "Last Contact Notes"]
    # Populated with formulas referencing CRM and Sales
    rows = []
    for c in C.CUSTOMER_DATA:
        rows.append([c[0], c[1], c[2], c[5],
                      None, None, None, None, None, ""])

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[24], headers, rows,
                                       col_widths=[10, 22, 14, 18, 14, 18, 14, 18, 18, 25])
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab25_vendor(wb):
    """Tab 25: Vendor / Supplier Master."""
    ws = wb.create_sheet(title=C.TAB_NAMES[25])

    headers = ["Vendor ID", "Name", "Category", "Contact", "Phone",
               "Location", "Payment Terms", "Reliability (1-5)",
               "Price Score (1-5)", "Approved", "Notes"]
    rows = []
    for v in C.VENDOR_DATA:
        rows.append([v[0], v[1], v[2], v[3], v[4], v[5], v[6], v[7], v[8], v[9], ""])

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[25], headers, rows,
                                       col_widths=[10, 22, 16, 14, 16, 14, 12, 14, 14, 10, 20])
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab26_item_master(wb):
    """Tab 26: Inventory Item Master."""
    ws = wb.create_sheet(title=C.TAB_NAMES[26])

    headers = ["Item ID", "Item Name", "Category", "Unit", "Bag Weight (kg)",
               "Reorder Threshold (units)", "Reorder Threshold (days)",
               "Risk Class", "Preferred Supplier", "Storage Location", "Notes"]
    rows = []
    for item in C.ITEM_DATA:
        rows.append([
            item[0], item[1], item[2], item[3], item[4],
            item[5], item[6], item[7], item[8] or "",
            "Main Store", ""
        ])

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[26], headers, rows,
                                       col_widths=[10, 25, 14, 8, 14, 20, 20, 10, 16, 14, 20])
    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_tab27_feed_formulation(wb):
    """Tab 27: Feed Formulation Reference — target formulas and ingredient proportions."""
    ws = wb.create_sheet(title=C.TAB_NAMES[27])

    # Formula summary table
    headers = ["Formula ID", "Name", "Age Range", "Target ME (kcal/kg)",
               "Target CP (%)", "Target Calcium (%)", "Target Avail P (%)"]
    rows = []
    for fid, fdata in C.FEED_FORMULAS.items():
        rows.append([fid, fdata["name"], fdata["age_range"],
                      fdata["target_ME_kcal"], fdata["target_CP_pct"],
                      fdata["target_calcium_pct"], fdata["target_phos_pct"]])

    tab, end_row = create_excel_table(ws, C.TABLE_NAMES[27], headers, rows,
                                       col_widths=[12, 28, 14, 18, 12, 16, 16])

    # Ingredient proportions table (starts below)
    start_row = end_row + 3
    write_section_header(ws, start_row - 1, 1, "Ingredient Proportions by Formula", merge_end_col=7)

    ing_headers = ["Formula ID", "Ingredient Code", "Ingredient Name",
                   "Target %", "Min %", "Max %"]
    ing_rows = []
    for fid, fdata in C.FEED_FORMULAS.items():
        for ing in fdata["ingredients"]:
            ing_rows.append([fid, ing[0], ing[1], ing[2], ing[3], ing[4]])

    create_excel_table(ws, C.TABLE_NAMES["27b"], ing_headers, ing_rows,
                       start_row=start_row, col_widths=[12, 14, 22, 10, 8, 8],
                       number_formats={3: C.FMT_DECIMAL_1, 4: C.FMT_DECIMAL_1, 5: C.FMT_DECIMAL_1})

    freeze_panes(ws)
    protect_sheet(ws)
    return ws


def build_layer2(wb):
    """Build all 11 Layer 2 (Master Data) tabs and create named ranges."""
    build_tab17_config(wb)
    build_tab18_staff(wb)
    build_tab19_housing(wb)
    build_tab20_flock(wb)
    build_tab21_breeder_curves(wb)
    build_tab22_owner_overrides(wb)
    build_tab23_customer_crm(wb)
    build_tab24_customer_profile(wb)
    build_tab25_vendor(wb)
    build_tab26_item_master(wb)
    build_tab27_feed_formulation(wb)

    # Create named ranges for dropdown references
    num_staff = len(C.STAFF_DATA) + 1
    num_housing = len(C.HOUSING_DATA) + 1
    num_flock = len(C.FLOCK_DATA) + 1
    num_customer = len(C.CUSTOMER_DATA) + 1
    num_vendor = len(C.VENDOR_DATA) + 1
    num_item = len(C.ITEM_DATA) + 1

    create_named_range(wb, "StaffIDs", C.TAB_NAMES[18], f"$A$2:$A${num_staff}")
    create_named_range(wb, "StaffNames", C.TAB_NAMES[18], f"$B$2:$B${num_staff}")
    create_named_range(wb, "CageIDs", C.TAB_NAMES[19], f"$A$2:$A${num_housing}")
    create_named_range(wb, "HouseNames", C.TAB_NAMES[19], f"$B$2:$B${num_housing}")
    create_named_range(wb, "FlockIDs", C.TAB_NAMES[20], f"$A$2:$A${num_flock}")
    create_named_range(wb, "CustomerIDs", C.TAB_NAMES[23], f"$A$2:$A${num_customer}")
    create_named_range(wb, "CustomerNames", C.TAB_NAMES[23], f"$B$2:$B${num_customer}")
    create_named_range(wb, "VendorIDs", C.TAB_NAMES[25], f"$A$2:$A${num_vendor}")
    create_named_range(wb, "VendorNames", C.TAB_NAMES[25], f"$B$2:$B${num_vendor}")
    create_named_range(wb, "ItemIDs", C.TAB_NAMES[26], f"$A$2:$A${num_item}")
    create_named_range(wb, "ItemNames", C.TAB_NAMES[26], f"$B$2:$B${num_item}")
    create_named_range(wb, "BreederWeeks", C.TAB_NAMES[21], f"$A$2:$A${len(C.BREEDER_CURVE)+1}")
    create_named_range(wb, "FormulaIDs", C.TAB_NAMES[27], f"$A$2:$A${len(C.FEED_FORMULAS)+1}")

    print("  Layer 2: 11 master data tabs + 13 named ranges created")
