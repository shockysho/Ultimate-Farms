"""
tab1_dashboard.py -- Tab 1: Dashboard (Owner + Manager views on one sheet).

v3.0 Verification Core: Ghost Money section added as Section A (Owner),
5 reconciliation engines in Manager view.

All table references use T("key") from config.
"""

from openpyxl.styles import Font, Alignment

from . import config as C
from .config import T
from .helpers import (
    protect_sheet,
    freeze_panes,
    write_section_header,
    add_text_status_cf,
    set_col_widths,
)

# ---------------------------------------------------------------------------
# Fonts
# ---------------------------------------------------------------------------
DASH_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
DASH_SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="4472C4")
KPI_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
KPI_VALUE_FONT = Font(name="Calibri", size=14, bold=True)
GHOST_FONT = Font(name="Calibri", size=11, bold=True, color="9C0006")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _write_kpi_row(ws, row, col, label, formula, target_text, status_formula, fmt=None):
    """Write a single KPI row: Label | Value | Target | Status."""
    ws.cell(row=row, column=col, value=label).font = KPI_LABEL_FONT
    cell_val = ws.cell(row=row, column=col + 1)
    cell_val.value = formula
    cell_val.font = KPI_VALUE_FONT
    if fmt:
        cell_val.number_format = fmt
    ws.cell(row=row, column=col + 2, value=target_text)
    ws.cell(row=row, column=col + 3, value=status_formula)


# ---------------------------------------------------------------------------
# Owner Dashboard (top half)
# ---------------------------------------------------------------------------

def _build_owner_section(ws, start_row):
    """Build the Owner Dashboard section starting at *start_row*.

    Returns the next free row after the section.
    """
    # Table-name shortcuts via T()
    prod = T("daily_cage_log")
    feed = T("feed_consumption")
    sales = T("sales")
    equip = T("equipment")
    bio = T("biosecurity")
    labor = T("labor")
    flock = T("flock")
    target = T("target_resolver")
    recon_eggs = T("recon_eggs")
    recon_cash = T("recon_cash")
    fraud = T("fraud_flags")
    procurement = T("procurement")
    health = T("health_incident")
    ghost = T("ghost_money")

    # -- Title ---------------------------------------------------------------
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=6)
    title_cell = ws.cell(row=start_row, column=1,
                         value="ULTIMATE FARMS -- OWNER DASHBOARD")
    title_cell.font = DASH_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=start_row + 1, start_column=1,
                   end_row=start_row + 1, end_column=6)
    sub_cell = ws.cell(
        row=start_row + 1, column=1,
        value="Exception-driven: Green = ignore | Yellow = investigate | Red = act now",
    )
    sub_cell.font = DASH_SUBTITLE_FONT
    sub_cell.alignment = Alignment(horizontal="center")

    r = start_row + 3  # leave one blank row after subtitle

    # ── SECTION A: GHOST MONEY TRACKER ────────────────────────────────────
    write_section_header(ws, r, 1,
                         "A -- GHOST MONEY TRACKER (Where Money Disappears)",
                         merge_end_col=6)
    r += 1

    gm_headers = ["Component", "Today (GHS)", "7-Day Avg (GHS)", "MTD (GHS)", "Cumulative (GHS)", "Status"]
    for ci, h in enumerate(gm_headers):
        cell = ws.cell(row=r, column=ci + 1, value=h)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
        cell.alignment = C.HEADER_ALIGN
    r += 1

    # Ghost Money component column names in tblGhostMoney
    gm_cols = [
        ("Feed Shrinkage", "Feed Shrinkage (GHS)"),
        ("Mortality Over-Target", "Mortality Over-Target (GHS)"),
        ("Egg Variance Loss", "Egg Variance Loss (GHS)"),
        ("Cracked/Damaged", "Cracked/Damaged (GHS)"),
        ("Price Arbitrage Missed", "Price Arbitrage Missed (GHS)"),
        ("Cash Discrepancy", "Cash Discrepancy (GHS)"),
        ("Inventory Carrying", "Inventory Carrying (GHS)"),
    ]

    gm_yel = C.THRESHOLDS["ghost_money_daily_yellow"]
    gm_red = C.THRESHOLDS["ghost_money_daily_red"]

    gm_start = r
    for label, col_name in gm_cols:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        # Today
        cell_today = ws.cell(
            row=r, column=2,
            value=f'=IFERROR(SUMIFS({ghost}[{col_name}],{ghost}[Date],TODAY()),0)',
        )
        cell_today.font = KPI_VALUE_FONT
        cell_today.number_format = C.FMT_CURRENCY
        # 7-Day Avg
        cell_avg = ws.cell(
            row=r, column=3,
            value=f'=IFERROR(AVERAGEIFS({ghost}[{col_name}],{ghost}[Date],">="&(TODAY()-6)),0)',
        )
        cell_avg.number_format = C.FMT_CURRENCY
        # MTD
        cell_mtd = ws.cell(
            row=r, column=4,
            value=f'=IFERROR(SUMIFS({ghost}[{col_name}],{ghost}[Date],">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1)),0)',
        )
        cell_mtd.number_format = C.FMT_CURRENCY
        # Cumulative (last row value)
        cell_cum = ws.cell(
            row=r, column=5,
            value=f'=IFERROR(SUMIFS({ghost}[{col_name}],{ghost}[Date],">="&MIN({ghost}[Date])),0)',
        )
        cell_cum.number_format = C.FMT_CURRENCY
        # Status
        ws.cell(
            row=r, column=6,
            value=f'=IF(B{r}>{gm_red / 7},"Red",IF(B{r}>{gm_yel / 7},"Yellow","Green"))',
        )
        r += 1

    # TOTAL row
    ws.cell(row=r, column=1, value="TOTAL GHOST MONEY").font = GHOST_FONT
    for col_idx in range(2, 6):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(
            row=r, column=col_idx,
            value=f'=SUM({col_letter}{gm_start}:{col_letter}{r - 1})',
        )
        cell.font = GHOST_FONT
        cell.number_format = C.FMT_CURRENCY
    ws.cell(
        row=r, column=6,
        value=f'=IF(B{r}>{gm_red},"Red",IF(B{r}>{gm_yel},"Yellow","Green"))',
    )
    r += 1

    # Ghost Money definition note
    r += 1
    note_cell = ws.cell(
        row=r, column=1,
        value='Ghost Money = value that SHOULD exist but doesn\'t. SUM(Expected - Actual) across all verification loops.',
    )
    note_cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    add_text_status_cf(ws, f"F{gm_start}:F{r - 2}")

    # ── SECTION B: BIOLOGICAL (Age-Adjusted) ────────────────────────────────
    r += 1
    write_section_header(ws, r, 1,
                         "B -- BIOLOGICAL PERFORMANCE (Age-Adjusted)",
                         merge_end_col=6)
    r += 1

    headers_a = ["Metric", "Current", "Target", "Phase", "7d Trend", "Status"]
    for ci, h in enumerate(headers_a):
        cell = ws.cell(row=r, column=ci + 1, value=h)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
        cell.alignment = C.HEADER_ALIGN
    r += 1

    # Lay % = Total Eggs today / bird count, averaged over 7 days
    housing = T("housing")
    lay_formula_a = (
        f'=IFERROR(SUMPRODUCT(({prod}[Total Eggs])*({prod}[Date]>=TODAY()-6)'
        f'*(IFERROR(INDEX({housing}[Cohort ID],MATCH({prod}[Cage ID],{housing}[Cage ID],0)),"")="FL2024A"))'
        f'/(SUMPRODUCT(({prod}[Date]>=TODAY()-6)'
        f'*(IFERROR(INDEX({housing}[Cohort ID],MATCH({prod}[Cage ID],{housing}[Cage ID],0)),"")="FL2024A")'
        f'*(IFERROR(INDEX({flock}[Current Bird Count],MATCH("FL2024A",{flock}[Cohort ID],0)),1)))'
        f'/30),0)'
    )
    lay_formula_b = (
        f'=IFERROR(SUMPRODUCT(({prod}[Total Eggs])*({prod}[Date]>=TODAY()-6)'
        f'*(IFERROR(INDEX({housing}[Cohort ID],MATCH({prod}[Cage ID],{housing}[Cage ID],0)),"")="FL2024B"))'
        f'/(SUMPRODUCT(({prod}[Date]>=TODAY()-6)'
        f'*(IFERROR(INDEX({housing}[Cohort ID],MATCH({prod}[Cage ID],{housing}[Cage ID],0)),"")="FL2024B")'
        f'*(IFERROR(INDEX({flock}[Current Bird Count],MATCH("FL2024B",{flock}[Cohort ID],0)),1)))'
        f'/30),0)'
    )

    # Thresholds
    mort_y = C.THRESHOLDS["mortality_daily_yellow"]
    mort_r = C.THRESHOLDS["mortality_daily_red"]
    disease_r = C.THRESHOLDS["disease_mortality_immediate_red"]
    cracked_y = C.THRESHOLDS["cracked_pct_yellow"]
    large_y = C.THRESHOLDS["large_pct_yellow_peak"]
    fcr_y = C.THRESHOLDS["fcr_peak_yellow"]
    fcr_r = C.THRESHOLDS["fcr_peak_red"]

    bio_kpis = [
        (
            "Lay % (FL2024A)",
            lay_formula_a,
            f'=IFERROR(INDEX({target}[Effective Target: Lay %],MATCH("FL2024A",{target}[Cohort ID],0)),"N/A")',
            f'=IFERROR(INDEX({target}[Production Phase],MATCH("FL2024A",{target}[Cohort ID],0)),"N/A")',
            "->",
            '=IF(ISNUMBER(B{r}),IF(NOT(ISNUMBER(C{r})),"N/A",IF(B{r}>=C{r},"Green",IF(B{r}>=C{r}*0.95,"Yellow","Red"))),"N/A")',
        ),
        (
            "Lay % (FL2024B)",
            lay_formula_b,
            f'=IFERROR(INDEX({target}[Effective Target: Lay %],MATCH("FL2024B",{target}[Cohort ID],0)),"N/A")',
            f'=IFERROR(INDEX({target}[Production Phase],MATCH("FL2024B",{target}[Cohort ID],0)),"N/A")',
            "->",
            '=IF(ISNUMBER(B{r}),IF(NOT(ISNUMBER(C{r})),"N/A",IF(B{r}>=C{r},"Green",IF(B{r}>=C{r}*0.95,"Yellow","Red"))),"N/A")',
        ),
        (
            "FCR (Overall)",
            f'=IFERROR(SUM({feed}[Net Consumed (kg)])/(SUM({prod}[Total Eggs])/12),0)',
            f'=IFERROR(INDEX({target}[Effective Target: FCR],1),"2.0")',
            "",
            "->",
            f'=IF(ISNUMBER(B{{r}}),IF(B{{r}}<={fcr_y},"Green",IF(B{{r}}<={fcr_r},"Yellow","Red")),"N/A")',
        ),
        (
            "Total Mortality (today)",
            f'=IFERROR(SUMIFS({prod}[Deaths],{prod}[Date],TODAY()),0)',
            f"<{mort_y}",
            "",
            "",
            f'=IF(ISNUMBER(B{{r}}),IF(B{{r}}<{mort_y},"Green",IF(B{{r}}<{mort_r},"Yellow","Red")),"N/A")',
        ),
        (
            "Disease Mortality (today)",
            f'=IFERROR(SUMIFS({prod}[Deaths],{prod}[Date],TODAY(),{prod}[Death Cause],"Disease"),0)',
            f"<{disease_r}",
            "",
            "",
            f'=IF(ISNUMBER(B{{r}}),IF(B{{r}}<{disease_r},"Green","Red"),"N/A")',
        ),
        (
            "Cracked %",
            (
                f'=IFERROR(SUMIFS({prod}[Grade: Cracked/Broken],{prod}[Date],TODAY())'
                f'/SUMIFS({prod}[Total Eggs],{prod}[Date],TODAY()),0)'
            ),
            f"<{cracked_y*100:.0f}%",
            "",
            "->",
            f'=IF(ISNUMBER(B{{r}}),IF(B{{r}}<{cracked_y},"Green",IF(B{{r}}<{cracked_y*1.5},"Yellow","Red")),"N/A")',
        ),
        (
            "Large %",
            (
                f'=IFERROR(SUMIFS({prod}[Grade: Large],{prod}[Date],TODAY())'
                f'/SUMIFS({prod}[Total Eggs],{prod}[Date],TODAY()),0)'
            ),
            f">{large_y*100:.0f}%",
            "",
            "->",
            f'=IF(ISNUMBER(B{{r}}),IF(B{{r}}>={large_y},"Green",IF(B{{r}}>={large_y*0.9},"Yellow","Red")),"N/A")',
        ),
    ]

    status_start = r
    for kpi in bio_kpis:
        label, value_formula, target_text, phase, trend, status_tmpl = kpi
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        cell_val = ws.cell(row=r, column=2, value=value_formula)
        cell_val.font = KPI_VALUE_FONT
        ws.cell(row=r, column=3, value=target_text)
        ws.cell(row=r, column=4, value=phase)
        ws.cell(row=r, column=5, value=trend)
        ws.cell(row=r, column=6, value=status_tmpl.format(r=r))
        r += 1

    add_text_status_cf(ws, f"F{status_start}:F{r - 1}")

    # ── SECTION C: OPERATIONAL ──────────────────────────────────────────────
    r += 1
    write_section_header(ws, r, 1, "C -- OPERATIONAL STATUS", merge_end_col=6)
    r += 1

    op_headers = ["Metric", "Current", "Target", "Status"]
    for ci, h in enumerate(op_headers):
        cell = ws.cell(row=r, column=ci + 1, value=h)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
    r += 1

    op_kpis = [
        (
            "Feed Days on Hand",
            "=7",
            ">=7 days",
            '=IF(B{r}>=7,"Green",IF(B{r}>=3,"Yellow","Red"))',
        ),
        (
            "Egg Stock (crates)",
            f'=IFERROR(SUMIFS({prod}[Total Eggs],{prod}[Date],TODAY())/30,0)',
            "<=500",
            '=IF(B{r}<=500,"Green",IF(B{r}<=1000,"Yellow","Red"))',
        ),
        (
            "Equipment Status",
            f'=IFERROR(COUNTIFS({equip}[Status],"Red"),"0")',
            "0 Red items",
            '=IF(B{r}=0,"Green",IF(B{r}<=1,"Yellow","Red"))',
        ),
        (
            "Biosecurity Compliance",
            f'=IFERROR(AVERAGEIFS({bio}[Compliance Score %],{bio}[Date],">="&(TODAY()-6)),1)',
            "100%",
            '=IF(B{r}>=0.95,"Green",IF(B{r}>=0.8,"Yellow","Red"))',
        ),
        (
            "Team Attendance",
            f'=IFERROR(COUNTIFS({labor}[Date],TODAY(),{labor}[Present],"Yes"),0)',
            f'>={C.THRESHOLDS["attendance_min_crew"]}',
            f'=IF(B{{r}}>={C.THRESHOLDS["attendance_min_crew"]},"Green","Red")',
        ),
    ]

    op_start = r
    for label, formula, target_text, status in op_kpis:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=formula).font = KPI_VALUE_FONT
        ws.cell(row=r, column=3, value=target_text)
        ws.cell(row=r, column=4, value=status.format(r=r))
        r += 1

    add_text_status_cf(ws, f"D{op_start}:D{r - 1}")

    # ── SECTION D: FINANCIAL (Month-to-Date) ────────────────────────────────
    r += 1
    write_section_header(ws, r, 1, "D -- FINANCIAL (Month-to-Date)", merge_end_col=6)
    r += 1

    for ci, h in enumerate(["Metric", "Current", "Target", "Status"]):
        cell = ws.cell(row=r, column=ci + 1, value=h)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
    r += 1

    fin_kpis = [
        (
            "Revenue MTD",
            (
                f'=SUMIFS({sales}[Line Total (GHS)],'
                f'{sales}[Date/Time],">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'
                f'{sales}[Date/Time],"<="&TODAY())'
            ),
            "Prior month pace",
            "",
        ),
        (
            "Cash Recon Variance (cumul.)",
            f'=IFERROR(SUM({recon_cash}[Daily Revenue Variance]),0)',
            "Zero",
            '=IF(ABS(B{r})<100,"Green",IF(ABS(B{r})<500,"Yellow","Red"))',
        ),
        (
            "Outstanding Receivables",
            f'=IFERROR(SUM({sales}[Line Total (GHS)])-SUM({recon_cash}[Total Payments]),0)',
            "<5000 GHS",
            '=IF(B{r}<5000,"Green",IF(B{r}<10000,"Yellow","Red"))',
        ),
        (
            "Ghost Money MTD",
            (
                f'=IFERROR(SUMIFS({ghost}[Daily Ghost Money (GHS)],'
                f'{ghost}[Date],">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1)),0)'
            ),
            "<15,000 GHS",
            '=IF(B{r}<15000,"Green",IF(B{r}<30000,"Yellow","Red"))',
        ),
    ]

    fin_start = r
    for label, formula, target_text, status in fin_kpis:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        cell = ws.cell(row=r, column=2, value=formula)
        cell.font = KPI_VALUE_FONT
        cell.number_format = C.FMT_CURRENCY
        ws.cell(row=r, column=3, value=target_text)
        if status:
            ws.cell(row=r, column=4, value=status.format(r=r))
        r += 1

    add_text_status_cf(ws, f"D{fin_start}:D{r - 1}")

    # ── SECTION E: FRAUD FLAGS & PENDING ACTIONS ────────────────────────────
    r += 1
    write_section_header(ws, r, 1,
                         "E -- FRAUD FLAGS & PENDING ACTIONS", merge_end_col=6)
    r += 1

    for ci, h in enumerate(["Item", "Count", "Action"]):
        cell = ws.cell(row=r, column=ci + 1, value=h)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
    r += 1

    fraud_items = [
        (
            "Open Red Fraud Flags",
            f'=COUNTIFS({fraud}[Severity],"Red",{fraud}[Status],"Open")',
            "Investigate immediately",
        ),
        (
            "Open Yellow Fraud Flags",
            f'=COUNTIFS({fraud}[Severity],"Yellow",{fraud}[Status],"Open")',
            "Review this week",
        ),
        (
            "Type E (Data Massage) Flags",
            f'=COUNTIFS({fraud}[Detection Type],"Type E",{fraud}[Status],"Open")',
            "AI-detected -- verify manually",
        ),
        (
            "Pending Procurement Approvals",
            f'=COUNTIFS({procurement}[Approval Status],"Pending")',
            "Approve or reject",
        ),
        (
            "Unresolved Health Incidents",
            f'=COUNTIFS({health}[Resolution Status],"Open")',
            "Follow up with vet/SC",
        ),
    ]

    for label, formula, action in fraud_items:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=formula).font = KPI_VALUE_FONT
        ws.cell(row=r, column=3, value=action)
        r += 1

    # ── SECTION F: NEXT ACTIONS (Priority-Ordered) ──────────────────────────
    r += 1
    write_section_header(ws, r, 1,
                         "F -- NEXT ACTIONS (Priority-Ordered)", merge_end_col=6)
    r += 1

    actions = [
        "1. Review Ghost Money trends -- which component is growing? (Analytics tab)",
        "2. Review cull recommendations (Prediction Engine -> Flock Registry)",
        "3. Check feed/ingredient reorder dates (Prediction Engine)",
        "4. Upcoming vaccinations due (Medication Log -> next dose)",
        "5. Vet-call trigger forecast (Prediction Engine -> mortality slope)",
        "6. Equipment maintenance due (Equipment Log -> patterns)",
        "7. Churn-risk customers to contact (CRM -> Red churn risk)",
        "8. Cycle counts scheduled for today (Scheduler)",
    ]
    for action in actions:
        ws.cell(row=r, column=1, value=action)
        r += 1

    return r  # next free row


# ---------------------------------------------------------------------------
# Manager Dashboard (bottom half)
# ---------------------------------------------------------------------------

def _build_manager_section(ws, start_row):
    """Build the Manager Dashboard section starting at *start_row*.

    Returns the next free row after the section.
    """
    # Table-name shortcuts via T()
    prod = T("daily_cage_log")
    env = T("environmental")
    feed = T("feed_consumption")
    water = T("water_consumption")
    bio = T("biosecurity")
    labor = T("labor")
    recon_eggs = T("recon_eggs")
    recon_cash = T("recon_cash")
    recon_feed = T("recon_feed")
    recon_mort = T("recon_mortality")
    recon_inv = T("recon_inventory")
    equip = T("equipment")
    health = T("health_incident")
    procurement = T("procurement")

    # -- Title ---------------------------------------------------------------
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=6)
    title_cell = ws.cell(
        row=start_row, column=1,
        value="ULTIMATE FARMS -- MANAGER DASHBOARD (Daily Ops)",
    )
    title_cell.font = DASH_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=start_row + 1, start_column=1,
                   end_row=start_row + 1, end_column=6)
    sub_cell = ws.cell(
        row=start_row + 1, column=1,
        value="Site Coordinator daily workflow -- check completeness, review exceptions, plan actions",
    )
    sub_cell.font = DASH_SUBTITLE_FONT
    sub_cell.alignment = Alignment(horizontal="center")

    r = start_row + 3

    num_cages = len(C.HOUSING_DATA)

    # ── SECTION A: DATA COMPLETENESS ────────────────────────────────────────
    write_section_header(
        ws, r, 1,
        "A -- DATA COMPLETENESS (Did everyone log everything today?)",
        merge_end_col=4,
    )
    r += 1

    for ci, h in enumerate(["Check", "Status", "Missing"]):
        cell = ws.cell(row=r, column=ci + 1, value=h)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
    r += 1

    completeness = [
        (
            "Daily Cage Log today",
            f'=IF(COUNTIFS({prod}[Date],TODAY())>={num_cages},"Complete","MISSING")',
            (
                f'=IF(COUNTIFS({prod}[Date],TODAY())>={num_cages},"",'
                f'{num_cages}-COUNTIFS({prod}[Date],TODAY())&" cage entries missing")'
            ),
        ),
        (
            "Environmental Log today",
            f'=IF(COUNTIFS({env}[Date],TODAY())>=5,"Complete","MISSING")',
            f'=IF(COUNTIFS({env}[Date],TODAY())>=5,"","Need readings for all houses")',
        ),
        (
            "Feed Consumption today",
            f'=IF(COUNTIFS({feed}[Date],TODAY())>=6,"Complete","MISSING")',
            f'=IF(COUNTIFS({feed}[Date],TODAY())>=6,"","Missing feeding rounds")',
        ),
        (
            "Water Consumption today",
            f'=IF(COUNTIFS({water}[Date],TODAY())>=5,"Complete","MISSING")',
            '=""',
        ),
        (
            "Biosecurity Checklist today",
            f'=IF(COUNTIFS({bio}[Date],TODAY())>=1,"Done","NOT DONE")',
            '=""',
        ),
        (
            "Labor Log today",
            f'=IF(COUNTIFS({labor}[Date],TODAY())>=6,"Complete","MISSING")',
            f'=IF(COUNTIFS({labor}[Date],TODAY())>=6,"","Missing staff entries")',
        ),
    ]

    comp_start = r
    for label, status, missing in completeness:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=status)
        ws.cell(row=r, column=3, value=missing)
        r += 1

    add_text_status_cf(ws, f"B{comp_start}:B{r - 1}")

    # ── SECTION B: TODAY'S CYCLE COUNTS ─────────────────────────────────────
    r += 1
    write_section_header(ws, r, 1,
                         "B -- TODAY'S CYCLE COUNTS (from Scheduler)",
                         merge_end_col=4)
    r += 1
    ws.cell(row=r, column=1,
            value="-> Check Tab 8 (Analytics) for today's count list")
    r += 1

    # ── SECTION C: OPEN ITEMS ───────────────────────────────────────────────
    r += 1
    write_section_header(ws, r, 1,
                         "C -- OPEN ITEMS REQUIRING ATTENTION",
                         merge_end_col=4)
    r += 1

    for ci, h in enumerate(["Item", "Count", "Action"]):
        cell = ws.cell(row=r, column=ci + 1, value=h)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
    r += 1

    open_items = [
        (
            "Open Health Incidents",
            f'=COUNTIFS({health}[Resolution Status],"Open")',
            "Follow up / escalate",
        ),
        (
            "Equipment at Yellow/Red",
            f'=COUNTIFS({equip}[Status],"Yellow")+COUNTIFS({equip}[Status],"Red")',
            "Schedule repair / notify owner",
        ),
        (
            "Pending Procurement",
            f'=COUNTIFS({procurement}[Approval Status],"Pending")',
            "Submit to owner for approval",
        ),
    ]

    for label, formula, action in open_items:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=formula).font = KPI_VALUE_FONT
        ws.cell(row=r, column=3, value=action)
        r += 1

    # ── SECTION D: YESTERDAY'S RECONCILIATION (5 engines) ─────────────────
    r += 1
    write_section_header(ws, r, 1,
                         "D -- YESTERDAY'S RECONCILIATION STATUS (5 Engines)",
                         merge_end_col=4)
    r += 1

    for ci, h in enumerate(["Reconciliation", "Variance", "Status"]):
        cell = ws.cell(row=r, column=ci + 1, value=h)
        cell.font = C.HEADER_FONT
        cell.fill = C.HEADER_FILL
    r += 1

    recons = [
        (
            "Egg Reconciliation",
            (
                f'=IFERROR(SUMIFS({recon_eggs}[Variance (crates)],'
                f'{recon_eggs}[Date],TODAY()-1),"N/A")'
            ),
            (
                '=IF(ISNUMBER(B{r}),IF(ABS(B{r})<=0.5,"Green",'
                'IF(ABS(B{r})<=2,"Yellow","Red")),"N/A")'
            ),
        ),
        (
            "Cash Reconciliation",
            (
                f'=IFERROR(SUMIFS({recon_cash}[Daily Revenue Variance],'
                f'{recon_cash}[Date],TODAY()-1),"N/A")'
            ),
            (
                '=IF(ISNUMBER(B{r}),IF(ABS(B{r})<100,"Green",'
                'IF(ABS(B{r})<500,"Yellow","Red")),"N/A")'
            ),
        ),
        (
            "Feed Reconciliation",
            (
                f'=IFERROR(SUMIFS({recon_feed}[Discrepancy (kg)],'
                f'{recon_feed}[Date],TODAY()-1),"N/A")'
            ),
            (
                '=IF(ISNUMBER(B{r}),IF(ABS(B{r})<25,"Green",'
                'IF(ABS(B{r})<50,"Yellow","Red")),"N/A")'
            ),
        ),
        (
            "Mortality Reconciliation",
            (
                f'=IFERROR(SUMIFS({recon_mort}[Deaths Today],'
                f'{recon_mort}[Date],TODAY()-1),"N/A")'
            ),
            (
                f'=IF(ISNUMBER(B{{r}}),IF(B{{r}}<{C.THRESHOLDS["mortality_daily_yellow"]},"Green",'
                f'IF(B{{r}}<{C.THRESHOLDS["mortality_daily_red"]},"Yellow","Red")),"N/A")'
            ),
        ),
        (
            "Inventory Reconciliation",
            (
                f'=IFERROR(COUNTIFS({recon_inv}[Status],"Red"),"N/A")'
            ),
            (
                '=IF(ISNUMBER(B{r}),IF(B{r}=0,"Green",'
                'IF(B{r}<=2,"Yellow","Red")),"N/A")'
            ),
        ),
    ]

    recon_start = r
    for label, formula, status in recons:
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=formula).font = KPI_VALUE_FONT
        ws.cell(row=r, column=3, value=status.format(r=r))
        r += 1

    add_text_status_cf(ws, f"C{recon_start}:C{r - 1}")

    # ── SECTION E: WEEKLY HANDOVER PREP ─────────────────────────────────────
    r += 1
    write_section_header(
        ws, r, 1,
        "E -- WEEKLY HANDOVER PREP (Shows on handover day)",
        merge_end_col=4,
    )
    r += 1

    handover_items = [
        "* Verify all inventory totals before handover",
        "* Document equipment status for incoming team",
        "* Brief on outstanding issues and pending actions",
        "* Ensure all logs signed off for the rotation period",
    ]
    for item in handover_items:
        ws.cell(row=r, column=1, value=item)
        r += 1

    return r  # next free row


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_tab1_dashboard(wb):
    """Build Tab 1: Dashboard -- Owner + Manager views on one sheet."""
    ws = wb.create_sheet(title=C.TAB_NAMES[1])

    # ── Owner Dashboard (top half) ──────────────────────────────────────────
    next_row = _build_owner_section(ws, start_row=1)

    # Leave a visible gap (3 blank rows) between the two dashboards
    gap = 3
    manager_start = next_row + gap

    # ── Manager Dashboard (bottom half) ─────────────────────────────────────
    _build_manager_section(ws, start_row=manager_start)

    # ── Final formatting ────────────────────────────────────────────────────
    set_col_widths(ws, [30, 16, 30, 16, 10, 10])
    freeze_panes(ws, "A4")
    protect_sheet(ws)

    print("  Tab 1: Dashboard (Owner + Manager) -- Ghost Money + 5 Recon Engines created")
    return ws
