"""
import_legacy_data.py -- Read ULTIMATE FARMS DAILY RECORDS.xlsx and transform
into the OS v2.0 schema format compatible with generate_all_sample_data() output.
"""

from datetime import date, datetime, time, timedelta
from pathlib import Path
from openpyxl import load_workbook

from . import config as C


# ---------------------------------------------------------------------------
# Date parsing helper
# ---------------------------------------------------------------------------

def _sanitize_year(d):
    """Fix obvious year typos in legacy data (e.g. 2016 instead of 2026)."""
    if d is None:
        return None
    # Farm started operations around 2024-2025; any year before 2024 is a typo
    if d.year < 2024:
        corrected_year = d.year + 10  # e.g. 2016 -> 2026, 2015 -> 2025
        try:
            d = d.replace(year=corrected_year)
        except ValueError:
            pass  # e.g. Feb 29 in non-leap year
    return d


def _parse_date(val):
    """Parse a value that might be a date string, datetime, or date object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return _sanitize_year(val.date())
    if isinstance(val, date):
        return _sanitize_year(val)
    s = str(val).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y ", "%Y-%m-%d %H:%M:%S"):
        try:
            return _sanitize_year(datetime.strptime(s.strip(), fmt).date())
        except ValueError:
            continue
    # Handle date ranges like '9-14/2/2026' -- take the first date
    if "-" in s and "/" in s:
        try:
            first_part = s.split("-")[0]
            rest = s.split("/")[1:]  # month/year
            return _sanitize_year(datetime.strptime(f"{first_part}/{'/'.join(rest)}", "%d/%m/%Y").date())
        except (ValueError, IndexError):
            pass
    return None


def _safe_float(val, default=0.0):
    """Parse a numeric value, returning default if not parseable."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return default
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return default


def _safe_int(val, default=0):
    return int(_safe_float(val, default))


def _safe_str(val):
    """Get a clean string from a cell value."""
    if val is None:
        return ""
    return str(val).strip()


# ---------------------------------------------------------------------------
# Customer / Vendor name resolution
# ---------------------------------------------------------------------------

class NameResolver:
    """Fuzzy name-to-ID resolver with dynamic ID generation."""

    def __init__(self, known_data, id_prefix, id_start, name_index=1):
        self.lookup = {}
        self.known_data = known_data
        for entry in known_data:
            self.lookup[entry[name_index].lower().strip()] = entry[0]
        self.id_prefix = id_prefix
        self.next_id = id_start
        self.dynamic_entries = []
        self._seen_names = {}

    def resolve(self, name):
        if not name or not name.strip():
            return None
        name_lower = name.lower().strip()

        # Check cache of previously resolved dynamic names
        if name_lower in self._seen_names:
            return self._seen_names[name_lower]

        # Exact match
        if name_lower in self.lookup:
            return self.lookup[name_lower]

        # Partial match (either direction)
        for known_name, cid in self.lookup.items():
            if known_name in name_lower or name_lower in known_name:
                self._seen_names[name_lower] = cid
                return cid

        # No match -- create dynamic entry
        new_id = f"{self.id_prefix}{self.next_id:03d}"
        self.next_id += 1
        self._seen_names[name_lower] = new_id
        self.lookup[name_lower] = new_id
        self.dynamic_entries.append((new_id, name.strip()))
        return new_id


# ---------------------------------------------------------------------------
# Cage sheet parsing and transformation
# ---------------------------------------------------------------------------

def _get_cage_bird_weights(cage_ids):
    """Get bird count weights for a list of cage IDs from HOUSING_DATA."""
    weights = {}
    total = 0
    for cage_id, _, _, _, side, count, flock_id, active in C.HOUSING_DATA:
        if cage_id in cage_ids:
            weights[cage_id] = count
            total += count
    return weights, total


def _parse_cage_sheet(ws, cage_name):
    """Parse a cage sheet into raw data rows."""
    rows = []
    # Determine header row and column layout based on sheet name
    header_row = 9
    data_start = 10 if cage_name != "Cage 4" else 11

    # Read header to detect column positions
    headers = {}
    for c in range(1, 30):
        v = ws.cell(row=header_row, column=c).value
        if v:
            headers[str(v).strip().lower()] = c

    for r in range(data_start, 1200):
        date_val = ws.cell(row=r, column=1).value
        d = _parse_date(date_val)
        if d is None:
            # Check if there's any data in the row at all
            has_data = False
            for c in range(2, 6):
                if ws.cell(row=r, column=c).value is not None:
                    has_data = True
                    break
            if not has_data:
                continue
            # Row with data but no date -- might be a continuation, skip
            continue

        row_data = {"date": d}

        # Column mapping varies by cage sheet
        if cage_name == "Cage 1,2&3":
            # A:Date B:Opening Stock C:Age D:Mortality E:Isolation F:Culled G:Recovery
            # H:Feed/bags I:Total Eggs J:Total Eggs/crts K:Avg Egg Weight L:% Production
            # M:Medication N:Closing Stock O:Observations
            row_data["mortality"] = _safe_int(ws.cell(row=r, column=4).value)
            row_data["feed_bags"] = _safe_float(ws.cell(row=r, column=8).value)
            row_data["total_eggs"] = _safe_float(ws.cell(row=r, column=9).value)
            row_data["total_crates"] = _safe_float(ws.cell(row=r, column=10).value)
            row_data["medication"] = _safe_str(ws.cell(row=r, column=13).value)
            row_data["observations"] = _safe_str(ws.cell(row=r, column=15).value)

        elif cage_name == "Cage 4":
            # A:Date B:Opening C:Age D:Mortality E:Sold F:Isolation G:Culled H:Recovery
            # I:Feed/bags J:Total Eggs K:Total Eggs/crts L:Avg Egg Weight M:% Production
            # N:Medication O:Closing Stock P:Observations
            row_data["mortality"] = _safe_int(ws.cell(row=r, column=4).value)
            row_data["sold"] = _safe_int(ws.cell(row=r, column=5).value)
            row_data["feed_bags"] = _safe_float(ws.cell(row=r, column=9).value)
            row_data["total_eggs"] = _safe_float(ws.cell(row=r, column=10).value)
            row_data["total_crates"] = _safe_float(ws.cell(row=r, column=11).value)
            row_data["medication"] = _safe_str(ws.cell(row=r, column=14).value)
            row_data["observations"] = _safe_str(ws.cell(row=r, column=16).value)

        elif cage_name == "Cage 5":
            # A:Date B:Opening C:Age D:Mortality E:Sold F:Isolation G:Isolation centre birds
            # H:Culled I:Recovery J:Feed/bags K:Total Eggs L:Total Eggs/crts
            # M:Avg Egg Weight N:% Production O:Medication P:Closing Stock Q:Observations
            row_data["mortality"] = _safe_int(ws.cell(row=r, column=4).value)
            row_data["sold"] = _safe_int(ws.cell(row=r, column=5).value)
            row_data["isolation"] = _safe_int(ws.cell(row=r, column=6).value)
            row_data["feed_bags"] = _safe_float(ws.cell(row=r, column=10).value)
            row_data["total_eggs"] = _safe_float(ws.cell(row=r, column=11).value)
            row_data["total_crates"] = _safe_float(ws.cell(row=r, column=12).value)
            row_data["medication"] = _safe_str(ws.cell(row=r, column=15).value)
            row_data["observations"] = _safe_str(ws.cell(row=r, column=17).value)

        rows.append(row_data)

    return rows


def _transform_cage_to_daily_log(raw_rows, cage_name):
    """Transform cage rows into tblDailyCageLog format.

    Each legacy row is split proportionally across cage-sides based on bird counts.
    Returns list of rows matching: [date, cage_id, side, crates, singles,
        large, medium, small, cracked, deaths, death_cause, disease_sub, mortality_action, notes]
    """
    cage_ids = C.LEGACY_CAGE_MAP[cage_name]
    weights, total_birds = _get_cage_bird_weights(cage_ids)

    if total_birds == 0:
        return []

    result = []
    for row in raw_rows:
        d = row["date"]
        total_crates = row.get("total_crates", 0) or 0
        total_eggs = row.get("total_eggs", 0) or 0
        deaths_total = row.get("mortality", 0) or 0

        # Build notes
        notes_parts = []
        medication = row.get("medication", "")
        if medication and medication.lower() not in ("plain water", "plain water ", ""):
            notes_parts.append(f"Medication: {medication}")
        obs = row.get("observations", "")
        if obs:
            notes_parts.append(obs)
        sold = row.get("sold", 0)
        if sold:
            notes_parts.append(f"Sold: {sold}")
        notes = "; ".join(notes_parts)

        # Determine mortality action from isolation info (Cage 5 only)
        has_isolation = row.get("isolation", 0) or 0

        # Distribute proportionally across cage-sides
        allocated_crates = 0
        allocated_deaths = 0

        for i, cage_id in enumerate(cage_ids):
            bird_count = weights.get(cage_id, 0)
            fraction = bird_count / total_birds if total_birds > 0 else 0
            is_last = (i == len(cage_ids) - 1)

            # Crates -- proportional, remainder to last cage
            if is_last:
                cage_crates = int(total_crates) - allocated_crates
            else:
                cage_crates = int(round(total_crates * fraction))
                allocated_crates += cage_crates

            # Calculate singles for this cage-side
            cage_eggs = int(round(total_eggs * fraction)) if not is_last else int(total_eggs) - int(round(total_eggs * (1 - fraction))) if len(cage_ids) == 1 else 0
            # Simpler: just compute from cage_crates
            cage_singles = 0  # Legacy doesn't have per-cage singles data

            # Deaths -- proportional, remainder to last cage
            if is_last:
                cage_deaths = deaths_total - allocated_deaths
            else:
                cage_deaths = int(round(deaths_total * fraction))
                allocated_deaths += cage_deaths

            # Get side from HOUSING_DATA
            side = None
            for hid, _, _, _, s, _, _, _ in C.HOUSING_DATA:
                if hid == cage_id:
                    side = s
                    break

            mortality_action = ""
            if has_isolation > 0 and cage_deaths > 0:
                mortality_action = "Isolated section"

            result.append([
                d, cage_id, side,
                max(0, cage_crates), cage_singles,
                0, 0, 0, 0,  # large, medium, small, cracked -- not tracked in legacy
                max(0, cage_deaths),
                "",   # death_cause -- not tracked
                "",   # disease_sub -- not tracked
                mortality_action,
                notes if i == 0 else "",  # only attach notes to first cage-side
            ])

    return result


# ---------------------------------------------------------------------------
# Sales parsing and transformation
# ---------------------------------------------------------------------------

def _parse_sales(ws):
    """Parse SALES sheet, skipping header and total/subtotal rows."""
    rows = []
    last_date = None

    for r in range(5, 1200):  # Data starts row 5
        date_val = ws.cell(row=r, column=1).value
        d = _parse_date(date_val)

        # Detect total/subtotal rows
        raw_date_str = _safe_str(date_val).lower()
        if raw_date_str in ("total", "totals", ""):
            # Check if column F has a value but B (customer name) is empty and it's a sum row
            cust = _safe_str(ws.cell(row=r, column=2).value)
            qty = ws.cell(row=r, column=6).value
            if not cust and raw_date_str == "total":
                continue  # Skip total rows

        if d is not None:
            last_date = d
        elif last_date is None:
            continue  # Skip rows before first date

        # Must have a customer name to be a real data row
        customer_name = _safe_str(ws.cell(row=r, column=2).value)
        if not customer_name:
            continue

        # Skip if this is actually a total row disguised
        if customer_name.lower() in ("total", "totals"):
            continue

        qty_crates = _safe_float(ws.cell(row=r, column=6).value)
        if qty_crates <= 0:
            continue  # Skip rows with no crates

        row_data = {
            "date": last_date,
            "customer_name": customer_name,
            "location": _safe_str(ws.cell(row=r, column=3).value),
            "phone": _safe_str(ws.cell(row=r, column=4).value),
            "invoice": _safe_str(ws.cell(row=r, column=5).value),
            "qty_crates": qty_crates,
            "size": _safe_str(ws.cell(row=r, column=7).value),
            "unit_price": _safe_float(ws.cell(row=r, column=8).value),
            "total": _safe_float(ws.cell(row=r, column=9).value),
            "payment_date": _safe_str(ws.cell(row=r, column=10).value),
            "acc_paid_to": _safe_str(ws.cell(row=r, column=11).value),
            "transaction_id": _safe_str(ws.cell(row=r, column=12).value),
            "sold_by": _safe_str(ws.cell(row=r, column=13).value),
            "remarks": _safe_str(ws.cell(row=r, column=15).value),
        }
        rows.append(row_data)

    return rows


def _transform_sales(raw_rows, customer_resolver):
    """Transform parsed sales rows into tblSales format."""
    result = []
    invoice_counter = 1000

    for row in raw_rows:
        d = row["date"]
        if d is None:
            continue

        # Customer ID resolution
        cust_id = customer_resolver.resolve(row["customer_name"])
        if cust_id is None:
            cust_id = "C014"  # Walk-in fallback

        # Invoice ID
        inv = row["invoice"]
        if not inv or inv.lower() in ("n/r", "random", "none", ""):
            invoice_counter += 1
            inv = f"INV-{invoice_counter}"
        elif not inv.startswith("INV-"):
            inv = f"INV-{inv}"

        # Egg grade from Size
        size = row["size"].lower()
        grade_map = {
            "unsorted": "Mixed", "mixed": "Mixed", "large": "Large",
            "medium": "Medium", "small": "Small", "": "Mixed",
        }
        egg_grade = grade_map.get(size, "Mixed")

        # Payment method inference from "Acc paid to"
        acc = row["acc_paid_to"].lower()
        if "momo" in acc or "mobile" in acc or "mtn" in acc:
            payment_method = "MoMo"
        elif "bank" in acc or "fidelity" in acc or "transfer" in acc:
            payment_method = "Bank Transfer"
        elif acc:
            payment_method = "MoMo"  # Default for non-empty payment accounts
        else:
            payment_method = "Cash"

        # Payment status
        has_evidence = bool(row["transaction_id"])
        has_payment_date = bool(row["payment_date"])
        if has_evidence or has_payment_date:
            payment_status = "Paid"
        else:
            payment_status = "Unpaid"

        # MoMo verified
        evidence_ref = row["transaction_id"]
        momo_verified = "Yes" if payment_method == "MoMo" and evidence_ref else "N/A"

        dt = datetime.combine(d, time(9, 0))

        result.append([
            dt, inv, cust_id, "Egg crates", egg_grade,
            int(row["qty_crates"]), 0,  # qty_crates, qty_singles
            row["unit_price"],
            payment_method, payment_status,
            evidence_ref, momo_verified,
            "Dispatched", "",  # delivery_status, dispatch_auth
            row["sold_by"], "",  # sold_by, proof_link
            row["remarks"],
        ])

    return result


# ---------------------------------------------------------------------------
# Procurement parsing and transformation
# ---------------------------------------------------------------------------

def _parse_procurement(ws):
    """Parse PROCUREMENT sheet."""
    rows = []
    last_date = None

    for r in range(2, 300):  # Headers at row 1, data starts row 2
        date_val = ws.cell(row=r, column=1).value
        d = _parse_date(date_val)

        if d is not None:
            last_date = d
        elif last_date is None:
            continue

        item_desc = _safe_str(ws.cell(row=r, column=3).value)
        if not item_desc:
            # Check if this is a subtotal row (column 5 has a value but no item)
            continue

        qty_raw = _safe_str(ws.cell(row=r, column=2).value)
        # Handle non-numeric quantities like "100pcs", "10L"
        qty = _safe_float(qty_raw)
        if qty == 0 and qty_raw:
            # Try to extract number
            import re
            nums = re.findall(r'[\d.]+', qty_raw)
            qty = float(nums[0]) if nums else 1.0

        row_data = {
            "date": last_date,
            "qty": qty,
            "item_desc": item_desc,
            "unit_price": _safe_float(ws.cell(row=r, column=4).value),
            "total": _safe_float(ws.cell(row=r, column=5).value),
            "bought_by": _safe_str(ws.cell(row=r, column=6).value),
            "shop_name": _safe_str(ws.cell(row=r, column=7).value),
            "location": _safe_str(ws.cell(row=r, column=8).value),
            "receipt": _safe_str(ws.cell(row=r, column=9).value),
            "vendor_phone": _safe_str(ws.cell(row=r, column=10).value),
            "status": _safe_str(ws.cell(row=r, column=11).value),
            "observations": _safe_str(ws.cell(row=r, column=12).value),
        }
        rows.append(row_data)

    return rows


def _infer_category(item_desc):
    """Infer procurement category from item description."""
    desc = item_desc.lower()
    if any(w in desc for w in ("maize", "soya", "bran", "limestone", "premix",
                                "feed", "layer mash", "corn")):
        return "Feed ingredient"
    if any(w in desc for w in ("vaccine", "medication", "antibiotic", "vitamin",
                                "oligo", "calcium", "dewormer", "supplement")):
        return "Medication"
    if any(w in desc for w in ("diesel", "fuel", "petrol", "gas")):
        return "Fuel"
    if any(w in desc for w in ("repair", "belt", "motor", "pipe", "weld",
                                "technician", "workmanship", "battery")):
        return "Repairs"
    if any(w in desc for w in ("tray", "crate", "boot", "glove", "ppe",
                                "disinfectant", "cleantec", "pesticide",
                                "rodent", "bait", "filter")):
        return "Supplies"
    if any(w in desc for w in ("transport", "delivery", "shipping")):
        return "Other"
    return "Other"


def _infer_unit(item_desc):
    """Infer unit from item description."""
    desc = item_desc.lower()
    if any(w in desc for w in ("bag", "bags")):
        return "bags"
    if any(w in desc for w in ("liter", "litre", "fuel", "diesel", "petrol", "oil")):
        return "liters"
    if any(w in desc for w in ("bottle", "bottles")):
        return "bottles"
    if "kg" in desc:
        return "kg"
    return "units"


def _transform_procurement(raw_rows, vendor_resolver):
    """Transform parsed procurement rows into tblProcurement format."""
    result = []
    proc_counter = 500

    for row in raw_rows:
        d = row["date"]
        if d is None:
            continue

        # Vendor resolution
        vendor_name = row["shop_name"] or row["bought_by"] or "Unknown"
        vendor_id = vendor_resolver.resolve(vendor_name)
        if vendor_id is None:
            vendor_id = "V001"

        category = _infer_category(row["item_desc"])
        unit = _infer_unit(row["item_desc"])

        receipt = row["receipt"]
        if receipt and not receipt.startswith("PRC-"):
            receipt = f"PRC-{receipt}"
        elif not receipt:
            proc_counter += 1
            receipt = f"PRC-{proc_counter}"

        approval = "Not Required"
        if row["total"] >= C.THRESHOLDS.get("procurement_approval_threshold", 500):
            approval = "Approved" if row["status"].lower() == "paid" else "Pending"

        dt = datetime.combine(d, time(10, 0))

        result.append([
            dt, vendor_id, category, row["item_desc"],
            row["qty"], unit, row["unit_price"],
            None,  # Total = formula
            "Cash", "",  # payment_method, pay_ref
            receipt,
            row["bought_by"], approval, "",  # requested_by, approval_status, approved_by
            None,  # approval_date
            "Yes", "",  # received_verified, verified_by
            None,  # price_benchmark = formula
            row["observations"],
        ])

    return result


# ---------------------------------------------------------------------------
# Feed parsing and transformation
# ---------------------------------------------------------------------------

def _parse_feed(ws):
    """Parse FEED sheet."""
    rows = []
    for r in range(5, 600):  # Headers at row 4, data starts row 5
        date_val = ws.cell(row=r, column=1).value
        d = _parse_date(date_val)
        if d is None:
            continue

        qty_used = _safe_float(ws.cell(row=r, column=3).value)

        rows.append({
            "date": d,
            "qty_used_bags": qty_used,
            "remarks": _safe_str(ws.cell(row=r, column=6).value),
        })

    return rows


def _transform_feed(raw_rows):
    """Transform feed rows into tblFeedConsumption format.

    Creates 2 rows per day: one per cohort, proportional to bird counts.
    FL2024A = 3210/5000 = 64.2%, FL2024B = 1790/5000 = 35.8%
    """
    fl_a_birds = sum(c[5] for c in C.HOUSING_DATA if c[6] == "FL2024A")
    fl_b_birds = sum(c[5] for c in C.HOUSING_DATA if c[6] == "FL2024B")
    total_birds = fl_a_birds + fl_b_birds
    fl_a_frac = fl_a_birds / total_birds if total_birds > 0 else 0.5

    result = []
    for row in raw_rows:
        d = row["date"]
        qty_bags = row["qty_used_bags"]
        if qty_bags <= 0:
            continue

        total_kg = qty_bags * 50  # 50 kg per bag

        for flock_id, fraction in [("FL2024A", fl_a_frac), ("FL2024B", 1 - fl_a_frac)]:
            cohort_kg = round(total_kg * fraction, 1)
            result.append([
                d, "AM-40%", flock_id, None,  # date, round, cohort_id, cage_id
                "FF001", cohort_kg, 0.0,  # feed_type, qty_kg, qty_wasted
                "", "",  # issued_by, verified_by
                row["remarks"] if flock_id == "FL2024A" else "",
            ])

    return result


# ---------------------------------------------------------------------------
# Crates parsing and transformation
# ---------------------------------------------------------------------------

def _parse_crates(ws):
    """Parse CRATES sheet."""
    rows = []
    for r in range(5, 600):
        date_val = ws.cell(row=r, column=1).value
        d = _parse_date(date_val)
        if d is None:
            continue

        rows.append({
            "date": d,
            "opening": _safe_float(ws.cell(row=r, column=2).value),
            "closing": _safe_float(ws.cell(row=r, column=7).value),
        })

    return rows


def _transform_crates(raw_rows):
    """Transform crates into weekly tblInventoryCount snapshots."""
    result = []
    for i, row in enumerate(raw_rows):
        # Every 7th row (weekly snapshot)
        if i % 7 != 0:
            continue

        d = row["date"]
        expected = row["opening"]
        physical = row["closing"]

        result.append([
            d, "Scheduled", "SUP001", "Main Store",  # date, trigger, item, location
            int(expected), int(physical),
            None, None,  # variance, variance_pct = formulas
            "", "", "",  # explanation, counted_by, verified_by
            "", "",  # approved_by, notes
        ])

    return result


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def import_all_legacy_data(filepath):
    """Read legacy XLSX and return (data_dict, dynamic_customers, dynamic_vendors).

    data_dict has the same keys and row shapes as generate_all_sample_data().
    """
    print(f"  Reading legacy file: {filepath}")
    wb = load_workbook(str(filepath), data_only=True, read_only=True)

    # Initialize resolvers
    customer_resolver = NameResolver(C.CUSTOMER_DATA, "C", 16, name_index=1)
    vendor_resolver = NameResolver(C.VENDOR_DATA, "V", 11, name_index=1)

    # Initialize result dict
    data = {
        "daily_cage_log": [],
        "environmental": [],
        "feed_consumption": [],
        "water_consumption": [],
        "ingredient_movement": [],
        "feed_mixing": [],
        "sales": [],
        "procurement": [],
        "equipment": [],
        "biosecurity": [],
        "visitor": [],
        "health_incident": [],
        "medication": [],
        "labor": [],
        "inventory_count": [],
    }

    # --- Parse cage sheets ---
    for cage_name in ["Cage 1,2&3", "Cage 4", "Cage 5"]:
        if cage_name in wb.sheetnames:
            print(f"    Parsing {cage_name}...")
            raw = _parse_cage_sheet(wb[cage_name], cage_name)
            transformed = _transform_cage_to_daily_log(raw, cage_name)
            data["daily_cage_log"].extend(transformed)
            print(f"      {len(raw)} legacy rows -> {len(transformed)} cage-side rows")

    # Sort daily_cage_log by date then cage_id
    data["daily_cage_log"].sort(key=lambda r: (r[0], r[1]))

    # --- Parse SALES ---
    if "SALES" in wb.sheetnames:
        print("    Parsing SALES...")
        raw_sales = _parse_sales(wb["SALES"])
        data["sales"] = _transform_sales(raw_sales, customer_resolver)
        print(f"      {len(raw_sales)} rows -> {len(data['sales'])} sales records")

    # --- Parse PROCUREMENT ---
    if "PROCUREMENT" in wb.sheetnames:
        print("    Parsing PROCUREMENT...")
        raw_proc = _parse_procurement(wb["PROCUREMENT"])
        data["procurement"] = _transform_procurement(raw_proc, vendor_resolver)
        print(f"      {len(raw_proc)} rows -> {len(data['procurement'])} procurement records")

    # --- Parse FEED ---
    if "FEED" in wb.sheetnames:
        print("    Parsing FEED...")
        raw_feed = _parse_feed(wb["FEED"])
        data["feed_consumption"] = _transform_feed(raw_feed)
        print(f"      {len(raw_feed)} feed days -> {len(data['feed_consumption'])} feed records")

    # --- Parse CRATES ---
    if "CRATES" in wb.sheetnames:
        print("    Parsing CRATES...")
        raw_crates = _parse_crates(wb["CRATES"])
        data["inventory_count"] = _transform_crates(raw_crates)
        print(f"      {len(raw_crates)} crate days -> {len(data['inventory_count'])} inventory snapshots")

    wb.close()

    # --- Compute date range from all data ---
    all_dates = []
    for row in data["daily_cage_log"]:
        if isinstance(row[0], date):
            all_dates.append(row[0])
    for row in data["sales"]:
        if isinstance(row[0], datetime):
            all_dates.append(row[0].date())
    for row in data["feed_consumption"]:
        if isinstance(row[0], date):
            all_dates.append(row[0])

    if all_dates:
        C.DATA_START_DATE = min(all_dates)
        C.DATA_END_DATE = max(all_dates)
        C.NUM_DAYS = (C.DATA_END_DATE - C.DATA_START_DATE).days + 1
        print(f"\n  Date range: {C.DATA_START_DATE} to {C.DATA_END_DATE} ({C.NUM_DAYS} days)")

    # Build dynamic customer/vendor tuples for Master Data
    dynamic_customers = []
    for new_id, name in customer_resolver.dynamic_entries:
        # (ID, Name, Type, Contact, Phone, Location, CreditAllowed, CreditLimit,
        #  TypicalOrder, Reliability, PriceSensitivity, PreferredPayment)
        dynamic_customers.append(
            (new_id, name, "Walk-in", name, "N/A", "Unknown", "No", 0, 1, 3, 3, "Cash")
        )

    dynamic_vendors = []
    for new_id, name in vendor_resolver.dynamic_entries:
        # (ID, Name, Category, Contact, Phone, Location, PaymentTerms,
        #  Reliability, PriceScore, Approved)
        dynamic_vendors.append(
            (new_id, name, "Other", name, "N/A", "Unknown", "COD", 3, 3, "Yes")
        )

    if dynamic_customers:
        print(f"  {len(dynamic_customers)} new customers discovered: {[c[1] for c in dynamic_customers[:5]]}{'...' if len(dynamic_customers) > 5 else ''}")
    if dynamic_vendors:
        print(f"  {len(dynamic_vendors)} new vendors discovered: {[v[1] for v in dynamic_vendors[:5]]}{'...' if len(dynamic_vendors) > 5 else ''}")

    total_rows = sum(len(v) for v in data.values())
    print(f"  Total data rows imported: {total_rows}")

    return data, dynamic_customers, dynamic_vendors
