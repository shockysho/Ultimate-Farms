"""
config.py — Central schema registry for Ultimate Farms Excel OS v2.0
All table names, column schemas, named ranges, formatting constants, and farm parameters.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection, numbers

# ============================================================
# TAB NAMES (worksheet titles, max 31 chars)
# ============================================================
TAB_NAMES = {
    # Layer 1 — Inputs
    1:  "1-Production",
    2:  "2-Environmental",
    3:  "3-Mortality",
    4:  "4-FeedConsumption",
    5:  "5-WaterConsumption",
    6:  "6-IngredientMovement",
    7:  "7-FeedMixing",
    8:  "8-Sales",
    9:  "9-Procurement",
    10: "10-Equipment",
    11: "11-BiosecurityChecklist",
    12: "12-VisitorLog",
    13: "13-HealthIncident",
    14: "14-MedicationVaccine",
    15: "15-LaborCompliance",
    16: "16-InventoryCount",
    # Layer 2 — Master Data
    17: "17-Config",
    18: "18-StaffRegistry",
    19: "19-HousingMap",
    20: "20-FlockRegistry",
    21: "21-BreederCurves",
    22: "22-OwnerOverrides",
    23: "23-CustomerCRM",
    24: "24-CustomerProfile",
    25: "25-VendorMaster",
    26: "26-ItemMaster",
    27: "27-FeedFormulation",
    # Layer 3 — Target Engine
    28: "28-TargetResolver",
    # Layer 4 — Reconciliation
    29: "29-ReconEggs",
    30: "30-ReconCash",
    31: "31-ReconFeed",
    32: "32-FraudFlags",
    # Layer 5 — Financial
    33: "33-ChartOfAccounts",
    34: "34-MonthlyPL",
    35: "35-UnitEconomics",
    # Layer 6 — Analytics
    36: "36-PerformanceAnalytics",
    37: "37-PredictionEngine",
    38: "38-CycleCountScheduler",
    # Views
    39: "39-OwnerDashboard",
    40: "40-ManagerDashboard",
}

# ============================================================
# TABLE NAMES (no spaces, unique across workbook)
# ============================================================
TABLE_NAMES = {
    1:  "tblProduction",
    2:  "tblEnvironmental",
    3:  "tblMortality",
    4:  "tblFeedConsumption",
    5:  "tblWaterConsumption",
    6:  "tblIngredientMovement",
    7:  "tblFeedMixBatchHeader",
    "7b": "tblFeedMixBatchLines",
    8:  "tblSales",
    9:  "tblProcurement",
    10: "tblEquipment",
    11: "tblBiosecurity",
    12: "tblVisitorLog",
    13: "tblHealthIncident",
    14: "tblMedication",
    15: "tblLabor",
    16: "tblInventoryCount",
    17: "tblConfig",
    18: "tblStaff",
    19: "tblHousing",
    20: "tblFlock",
    21: "tblBreederCurves",
    22: "tblOwnerOverrides",
    23: "tblCustomerCRM",
    24: "tblCustomerProfile",
    25: "tblVendor",
    26: "tblItemMaster",
    27: "tblFeedFormulation",
    "27b": "tblFormulaIngredients",
    28: "tblTargetResolver",
    29: "tblReconEggs",
    30: "tblReconCash",
    31: "tblReconFeed",
    "31b": "tblReconFinishedFeed",
    32: "tblFraudFlags",
    33: "tblChartOfAccounts",
    34: "tblMonthlyPL",
    35: "tblUnitEconomics",
    36: "tblPerfAnalytics",
    37: "tblPredictions",
    38: "tblCycleCountSchedule",
    39: "tblOwnerDash",
    40: "tblManagerDash",
}

# ============================================================
# TAB COLORS (hex, no #)
# ============================================================
TAB_COLORS = {
    "L1": "4472C4",    # Blue — Input tabs
    "L2": "70AD47",    # Green — Master Data
    "L3": "ED7D31",    # Orange — Target Engine
    "L4": "FF0000",    # Red — Reconciliation
    "L5": "7030A0",    # Purple — Financial
    "L6": "00B0F0",    # Teal — Analytics
    "View": "FFC000",  # Gold — Dashboards
}

def get_tab_color(tab_num):
    if 1 <= tab_num <= 16:
        return TAB_COLORS["L1"]
    elif 17 <= tab_num <= 27:
        return TAB_COLORS["L2"]
    elif tab_num == 28:
        return TAB_COLORS["L3"]
    elif 29 <= tab_num <= 32:
        return TAB_COLORS["L4"]
    elif 33 <= tab_num <= 35:
        return TAB_COLORS["L5"]
    elif 36 <= tab_num <= 38:
        return TAB_COLORS["L6"]
    else:
        return TAB_COLORS["View"]

# ============================================================
# FORMATTING CONSTANTS
# ============================================================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

ALERT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ALERT_GREEN_FONT = Font(color="006100")
ALERT_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ALERT_YELLOW_FONT = Font(color="9C5700")
ALERT_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ALERT_RED_FONT = Font(color="9C0006")

SECTION_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Number formats
FMT_CURRENCY = '#,##0.00" GHS"'
FMT_PERCENT = '0.0%'
FMT_PERCENT_2 = '0.00%'
FMT_DATE = 'YYYY-MM-DD'
FMT_DATETIME = 'YYYY-MM-DD HH:MM'
FMT_INTEGER = '#,##0'
FMT_DECIMAL_1 = '#,##0.0'
FMT_DECIMAL_2 = '#,##0.00'
FMT_TIME = 'HH:MM'

# ============================================================
# FARM-SPECIFIC CONSTANTS (Ghana, Lohmann Brown)
# ============================================================
FARM_NAME = "Ultimate Farms Ltd"
FARM_LOCATION = "Kumasi, Ashanti Region, Ghana"
CURRENCY = "GHS"
CRATE_SIZE = 30

# Staff
STAFF_DATA = [
    ("STF001", "Kwame Mensah", "Site Coordinator", "Team A", "2023-01-15", "0244-123-4561", "Active"),
    ("STF002", "Ama Darko", "Team Leader", "Team A", "2023-02-01", "0244-123-4562", "Active"),
    ("STF003", "Kofi Asante", "Production Hand", "Team A", "2023-03-01", "0244-123-4563", "Active"),
    ("STF004", "Akua Boateng", "Production Hand", "Team A", "2023-03-01", "0244-123-4564", "Active"),
    ("STF005", "Yaw Owusu", "Feed Handler", "Team A", "2023-04-01", "0244-123-4565", "Active"),
    ("STF006", "Abena Sarpong", "Team Leader", "Team B", "2023-02-01", "0244-123-4566", "Active"),
    ("STF007", "Kwesi Appiah", "Production Hand", "Team B", "2023-03-15", "0244-123-4567", "Active"),
    ("STF008", "Adwoa Frimpong", "Production Hand", "Team B", "2023-03-15", "0244-123-4568", "Active"),
    ("STF009", "Kojo Nyarko", "Feed Handler", "Team B", "2023-04-15", "0244-123-4569", "Active"),
    ("STF010", "Efua Mensah", "Cleaner", "Non-rotating", "2023-05-01", "0244-123-4570", "Active"),
    ("STF011", "Nana Amoako", "Night Watch", "Non-rotating", "2023-01-20", "0244-123-4571", "Active"),
    ("STF012", "Kweku Agyei", "Maintenance Tech", "Non-rotating", "2023-06-01", "0244-123-4572", "Active"),
]

# Housing
HOUSING_DATA = [
    ("H01-A", "House 1", "Row 1", 1, "A", 550, "FL2024A", "Yes"),
    ("H01-B", "House 1", "Row 1", 1, "B", 550, "FL2024A", "Yes"),
    ("H02-A", "House 2", "Row 1", 2, "A", 540, "FL2024A", "Yes"),
    ("H02-B", "House 2", "Row 1", 2, "B", 540, "FL2024A", "Yes"),
    ("H03-A", "House 3", "Row 2", 1, "A", 520, "FL2024A", "Yes"),
    ("H03-B", "House 3", "Row 2", 1, "B", 510, "FL2024A", "Yes"),
    ("H04-A", "House 4", "Row 2", 2, "A", 470, "FL2024B", "Yes"),
    ("H04-B", "House 4", "Row 2", 2, "B", 470, "FL2024B", "Yes"),
    ("H05-A", "House 5", "Row 3", 1, "A", 430, "FL2024B", "Yes"),
    ("H05-B", "House 5", "Row 3", 1, "B", 420, "FL2024B", "Yes"),
]

# Flocks
FLOCK_DATA = [
    ("FL2024A", "Lohmann Brown Classic", "2024-04-15", 3210, "Active"),
    ("FL2024B", "Lohmann Brown Classic", "2024-07-01", 1790, "Active"),
]

# Customers
CUSTOMER_DATA = [
    ("C001", "Kumasi Central Market", "Wholesale", "Nana Agyei", "0244-200-0001", "Kumasi Central", "Yes", 5000, 40, 5, 2, "MoMo"),
    ("C002", "Adum Egg Traders", "Wholesale", "Yaa Asantewaa", "0244-200-0002", "Adum, Kumasi", "Yes", 3000, 30, 4, 3, "MoMo"),
    ("C003", "Kejetia Provisions", "Wholesale", "Kofi Badu", "0244-200-0003", "Kejetia, Kumasi", "Yes", 4000, 25, 4, 2, "MoMo"),
    ("C004", "Bantama Food Court", "Regular Retail", "Ama Serwaa", "0244-200-0004", "Bantama, Kumasi", "No", 0, 8, 5, 3, "MoMo"),
    ("C005", "Oforikrom Bakery", "Regular Retail", "Ernest Osei", "0244-200-0005", "Oforikrom, Kumasi", "No", 0, 10, 4, 2, "MoMo"),
    ("C006", "Suame Chop Bar", "Regular Retail", "Akosua Manu", "0244-200-0006", "Suame, Kumasi", "No", 0, 5, 5, 4, "MoMo"),
    ("C007", "St. Louis SHS Kitchen", "Wholesale", "Sr. Mary", "0244-200-0007", "Kumasi", "Yes", 2000, 20, 5, 1, "Bank Transfer"),
    ("C008", "Prempeh College Kitchen", "Wholesale", "Mr. Adjei", "0244-200-0008", "Kumasi", "Yes", 2000, 15, 4, 1, "Bank Transfer"),
    ("C009", "MaxMart Supermarket", "Wholesale", "Purchasing Dept", "0244-200-0009", "Airport Rd, Kumasi", "Yes", 10000, 50, 3, 4, "Bank Transfer"),
    ("C010", "Ahodwo Residents Coop", "Regular Retail", "Mrs. Mensah", "0244-200-0010", "Ahodwo, Kumasi", "No", 0, 6, 5, 3, "MoMo"),
    ("C011", "Tech Junction Traders", "Agent", "Kwabena Poku", "0244-200-0011", "Tech, Kumasi", "Yes", 2000, 15, 3, 3, "MoMo"),
    ("C012", "Asafo Market Women", "Wholesale", "Madam Grace", "0244-200-0012", "Asafo, Kumasi", "Yes", 3000, 20, 4, 2, "MoMo"),
    ("C013", "Golden Tulip Hotel", "Regular Retail", "Chef Kitchen", "0244-200-0013", "Kumasi", "No", 0, 8, 5, 1, "Bank Transfer"),
    ("C014", "Walk-in Customer", "Walk-in", "Various", "N/A", "Farm Gate", "No", 0, 1, 3, 5, "Cash"),
    ("C015", "Ejisu Egg Depot", "Agent", "Samuel Tetteh", "0244-200-0015", "Ejisu", "Yes", 5000, 35, 4, 3, "MoMo"),
]

# Vendors
VENDOR_DATA = [
    ("V001", "Yedent Agro", "Feed ingredients", "Mr. Yeboah", "0244-300-0001", "Kumasi", "Net 7", 4, 4, "Yes"),
    ("V002", "WIENCO Ghana", "Feed ingredients", "Sales Rep", "0244-300-0002", "Accra", "Net 14", 5, 3, "Yes"),
    ("V003", "Chemico Ltd", "Medications", "Pharmacist", "0244-300-0003", "Kumasi", "COD", 5, 3, "Yes"),
    ("V004", "Afariwaa Farms Supply", "Feed ingredients", "Madam Afari", "0244-300-0004", "Ejura", "COD", 3, 5, "Yes"),
    ("V005", "Ghana Oil Palm Dev Co", "Feed ingredients", "Sales Office", "0244-300-0005", "Kwae", "Net 7", 4, 3, "Yes"),
    ("V006", "Kumasi Fuel Station", "Fuel", "Attendant", "0244-300-0006", "Kumasi", "COD", 5, 4, "Yes"),
    ("V007", "AK Electrical Works", "Repairs", "Mr. Ankrah", "0244-300-0007", "Kumasi", "COD", 4, 3, "Yes"),
    ("V008", "Zoetis Ghana", "Medications", "Vet Sales", "0244-300-0008", "Accra", "Net 14", 5, 2, "Yes"),
    ("V009", "Asante Milling Co", "Feed ingredients", "Mr. Asante", "0244-300-0009", "Kumasi", "Net 7", 3, 5, "Yes"),
    ("V010", "AllFarm Equipment", "Equipment", "Sales Mgr", "0244-300-0010", "Accra", "Net 30", 4, 3, "Yes"),
]

# Inventory items
ITEM_DATA = [
    ("ING001", "Maize (yellow corn)", "Ingredient", "kg", 50, 500, 7, "A", "V001"),
    ("ING002", "Soya bean meal", "Ingredient", "kg", 50, 300, 7, "A", "V002"),
    ("ING003", "Wheat bran", "Ingredient", "kg", 50, 200, 7, "B", "V001"),
    ("ING004", "Limestone (fine)", "Ingredient", "kg", 50, 150, 10, "B", "V004"),
    ("ING005", "Limestone (coarse)", "Ingredient", "kg", 50, 100, 10, "B", "V004"),
    ("ING006", "Vegetable oil (palm)", "Ingredient", "liters", 20, 50, 7, "B", "V005"),
    ("ING007", "Layer premix", "Ingredient", "kg", 25, 75, 14, "A", "V002"),
    ("ING008", "Salt", "Ingredient", "kg", 25, 20, 30, "C", "V001"),
    ("ING009", "DL-Methionine", "Ingredient", "kg", 25, 10, 14, "A", "V002"),
    ("ING010", "Mycotoxin binder", "Ingredient", "kg", 25, 10, 14, "B", "V002"),
    ("FF001", "Layer Mash A (self-milled)", "Finished Feed", "kg", 50, 500, 3, "A", None),
    ("FF002", "Layer Mash B (peak)", "Finished Feed", "kg", 50, 300, 3, "A", None),
    ("MED001", "Newcastle vaccine (LaSota)", "Medication", "bottles", None, 5, 30, "A", "V003"),
    ("MED002", "Infectious Bronchitis vaccine", "Medication", "bottles", None, 3, 30, "A", "V003"),
    ("MED003", "Oxytetracycline powder", "Medication", "bottles", None, 3, 14, "B", "V003"),
    ("MED004", "Vitamin AD3E supplement", "Supplement", "bottles", None, 5, 14, "B", "V003"),
    ("MED005", "Calcium liquid supplement", "Supplement", "liters", None, 10, 14, "B", "V008"),
    ("FUL001", "Diesel", "Fuel", "liters", None, 50, 7, "B", "V006"),
    ("FUL002", "Petrol (generator)", "Fuel", "liters", None, 30, 7, "B", "V006"),
    ("SUP001", "Egg crate trays (paper)", "Packaging", "units", None, 200, 14, "B", None),
    ("SUP002", "Rubber boots (Zone 1)", "PPE", "units", None, 2, 90, "C", None),
    ("SUP003", "Disinfectant (Virkon S)", "Chemicals", "kg", None, 5, 14, "B", "V003"),
    ("SUP004", "Rodent bait stations", "Chemicals", "units", None, 5, 30, "C", None),
    ("SPR001", "Fan belt (egg belt motor)", "Spare parts", "units", None, 2, 30, "B", "V010"),
    ("SPR002", "Water nipple replacements", "Spare parts", "units", None, 10, 30, "C", "V010"),
]

# Feed formulation reference
FEED_FORMULAS = {
    "FRM001": {
        "name": "Layer Mash A (Standard)",
        "age_range": "20-45 weeks",
        "target_ME_kcal": 2750,
        "target_CP_pct": 17.0,
        "target_calcium_pct": 4.0,
        "target_phos_pct": 0.35,
        "ingredients": [
            ("ING001", "Maize", 55.0, 50.0, 60.0),
            ("ING002", "Soya bean meal", 20.0, 18.0, 24.0),
            ("ING004", "Limestone (fine)", 5.5, 5.0, 6.0),
            ("ING005", "Limestone (coarse)", 3.7, 3.0, 4.0),
            ("ING003", "Wheat bran", 8.0, 5.0, 12.0),
            ("ING006", "Vegetable oil", 2.5, 2.0, 3.0),
            ("ING007", "Layer premix", 2.5, 2.5, 2.5),
            ("ING008", "Salt", 0.35, 0.30, 0.40),
            ("ING009", "DL-Methionine", 0.20, 0.15, 0.25),
            ("ING010", "Mycotoxin binder", 0.15, 0.10, 0.20),
        ],
    },
    "FRM002": {
        "name": "Layer Mash B (Peak/High-Ca)",
        "age_range": "45-65 weeks",
        "target_ME_kcal": 2700,
        "target_CP_pct": 16.5,
        "target_calcium_pct": 4.2,
        "target_phos_pct": 0.35,
        "ingredients": [
            ("ING001", "Maize", 53.0, 48.0, 58.0),
            ("ING002", "Soya bean meal", 19.0, 17.0, 22.0),
            ("ING004", "Limestone (fine)", 6.0, 5.5, 6.5),
            ("ING005", "Limestone (coarse)", 4.5, 4.0, 5.0),
            ("ING003", "Wheat bran", 9.5, 7.0, 13.0),
            ("ING006", "Vegetable oil", 2.5, 2.0, 3.0),
            ("ING007", "Layer premix", 2.5, 2.5, 2.5),
            ("ING008", "Salt", 0.35, 0.30, 0.40),
            ("ING009", "DL-Methionine", 0.20, 0.15, 0.25),
            ("ING010", "Mycotoxin binder", 0.15, 0.10, 0.20),
        ],
    },
}

# Breeder curve data (Lohmann Brown Classic, Ghana open-house conditions)
# (week, lay_pct, egg_weight_g, large_pct, feed_intake_g, fcr_kg_per_dozen, mortality_pct_monthly, phase)
BREEDER_CURVE = [
    (18, 0.05, 48, 0.00, 95, None, 0.50, "Ramp-up"),
    (19, 0.20, 50, 0.05, 100, None, 0.45, "Ramp-up"),
    (20, 0.50, 52, 0.10, 105, 4.00, 0.40, "Ramp-up"),
    (21, 0.70, 54, 0.20, 108, 3.20, 0.35, "Ramp-up"),
    (22, 0.82, 55, 0.30, 110, 2.60, 0.35, "Ramp-up"),
    (23, 0.88, 56, 0.40, 112, 2.30, 0.30, "Ramp-up"),
    (24, 0.91, 57, 0.45, 114, 2.10, 0.30, "Ramp-up"),
    (25, 0.92, 58, 0.55, 115, 2.00, 0.30, "Peak"),
    (26, 0.93, 58, 0.58, 115, 1.95, 0.30, "Peak"),
    (27, 0.94, 59, 0.60, 116, 1.90, 0.30, "Peak"),
    (28, 0.94, 59, 0.62, 116, 1.88, 0.30, "Peak"),
    (29, 0.94, 60, 0.64, 117, 1.85, 0.30, "Peak"),
    (30, 0.95, 60, 0.65, 117, 1.83, 0.30, "Peak"),
    (31, 0.95, 60, 0.66, 117, 1.82, 0.30, "Peak"),
    (32, 0.95, 61, 0.67, 117, 1.82, 0.30, "Peak"),
    (33, 0.94, 61, 0.68, 118, 1.83, 0.30, "Peak"),
    (34, 0.94, 61, 0.69, 118, 1.84, 0.30, "Peak"),
    (35, 0.94, 62, 0.70, 118, 1.80, 0.30, "Peak"),
    (36, 0.93, 62, 0.70, 118, 1.82, 0.30, "Peak"),
    (37, 0.93, 62, 0.70, 118, 1.83, 0.30, "Peak"),
    (38, 0.92, 62, 0.69, 118, 1.85, 0.30, "Peak"),
    (39, 0.92, 62, 0.69, 118, 1.86, 0.30, "Peak"),
    (40, 0.91, 63, 0.68, 118, 1.88, 0.30, "Peak"),
    (41, 0.91, 63, 0.68, 118, 1.90, 0.30, "Peak"),
    (42, 0.90, 63, 0.67, 118, 1.92, 0.30, "Peak"),
    (43, 0.90, 63, 0.67, 118, 1.93, 0.30, "Peak"),
    (44, 0.89, 63, 0.66, 118, 1.95, 0.30, "Peak"),
    (45, 0.90, 63, 0.68, 118, 1.90, 0.30, "Peak"),
    (46, 0.89, 63, 0.67, 118, 1.95, 0.35, "Post-peak"),
    (47, 0.88, 63, 0.66, 118, 1.98, 0.35, "Post-peak"),
    (48, 0.87, 64, 0.66, 118, 2.00, 0.35, "Post-peak"),
    (49, 0.87, 64, 0.65, 118, 2.02, 0.35, "Post-peak"),
    (50, 0.86, 64, 0.65, 119, 2.05, 0.35, "Post-peak"),
    (51, 0.85, 64, 0.64, 119, 2.08, 0.35, "Post-peak"),
    (52, 0.85, 64, 0.64, 119, 2.10, 0.35, "Post-peak"),
    (53, 0.84, 64, 0.63, 119, 2.12, 0.40, "Post-peak"),
    (54, 0.84, 64, 0.63, 119, 2.15, 0.40, "Post-peak"),
    (55, 0.85, 64, 0.65, 118, 2.20, 0.40, "Post-peak"),
    (56, 0.83, 64, 0.62, 119, 2.18, 0.40, "Post-peak"),
    (57, 0.82, 64, 0.62, 119, 2.20, 0.40, "Post-peak"),
    (58, 0.81, 64, 0.61, 120, 2.22, 0.40, "Post-peak"),
    (59, 0.80, 64, 0.60, 120, 2.25, 0.40, "Post-peak"),
    (60, 0.79, 64, 0.60, 120, 2.28, 0.40, "Post-peak"),
    (61, 0.78, 64, 0.60, 120, 2.50, 0.50, "Late-lay"),
    (62, 0.77, 64, 0.59, 120, 2.52, 0.50, "Late-lay"),
    (63, 0.76, 64, 0.58, 121, 2.55, 0.50, "Late-lay"),
    (64, 0.75, 64, 0.57, 121, 2.58, 0.50, "Late-lay"),
    (65, 0.78, 64, 0.60, 120, 2.50, 0.50, "Late-lay"),
    (66, 0.73, 63, 0.56, 121, 2.62, 0.55, "Late-lay"),
    (67, 0.72, 63, 0.55, 121, 2.65, 0.55, "Late-lay"),
    (68, 0.71, 63, 0.54, 122, 2.68, 0.55, "Late-lay"),
    (69, 0.70, 63, 0.55, 122, 2.70, 0.55, "Late-lay"),
    (70, 0.70, 63, 0.55, 122, 2.80, 0.60, "Late-lay"),
    (75, 0.65, 63, 0.50, 123, 2.90, 0.65, "Late-lay"),
    (80, 0.60, 62, 0.45, 124, 3.10, 0.70, "Late-lay"),
]

# Prices (GHS)
PRICES = {
    "egg_crate_wholesale": 42.00,
    "egg_crate_retail": 48.00,
    "egg_crate_walkin": 50.00,
    "egg_crate_default": 45.00,
    "manure_bag": 8.00,
    "manure_bulk_ton": 200.00,
    "culled_bird": 55.00,
    "ING001": 5.50,   # Maize per kg
    "ING002": 12.00,  # Soya per kg
    "ING003": 3.80,   # Wheat bran per kg
    "ING004": 2.00,   # Limestone fine per kg
    "ING005": 2.20,   # Limestone coarse per kg
    "ING006": 15.00,  # Vegetable oil per liter
    "ING007": 18.00,  # Layer premix per kg
    "ING008": 1.50,   # Salt per kg
    "ING009": 45.00,  # DL-Methionine per kg
    "ING010": 35.00,  # Mycotoxin binder per kg
    "FUL001": 15.50,  # Diesel per liter
    "FUL002": 16.00,  # Petrol per liter
}

# Alert thresholds (from architecture doc section 17)
THRESHOLDS = {
    "mortality_daily_yellow": 4,
    "mortality_daily_red": 6,
    "lay_rate_yellow": 0.85,
    "lay_rate_red": 0.82,
    "cash_mismatch_yellow_crates": 1,
    "cash_mismatch_red_crates": 2,
    "feed_stock_yellow_days": 7,
    "feed_stock_red_days": 3,
    "egg_stock_yellow_crates": 500,
    "egg_stock_red_crates": 1000,
    "equipment_uptime_yellow": 0.90,
    "equipment_uptime_red": 0.80,
    "cracked_pct_yellow": 0.06,
    "large_pct_yellow_peak": 0.60,
    "large_pct_yellow_postpeak": 0.55,
    "fcr_peak_yellow": 2.2,
    "fcr_peak_red": 2.5,
    "feed_variance_pct_yellow": 0.10,
    "procurement_approval_threshold": 500,
    "price_deviation_tolerance_pct": 0.15,
    "cash_deposit_window_hours": 24,
    "disease_mortality_immediate_red": 5,
    "heat_stress_hours_red": 2,
    "heat_stress_hours_day_yellow": 6,
    "water_drop_pct_yellow": 0.20,
    "water_spike_pct_yellow": 0.30,
    "biosecurity_score_yellow": 0.80,
    "attendance_min_crew": 6,
    "feed_discrepancy_bags_yellow": 0.5,
    "ingredient_shrinkage_yellow": 0.02,
    "ingredient_shrinkage_red": 0.05,
}

# Equipment list
EQUIPMENT_LIST = [
    "Generator",
    "Borehole pump",
    "Water pump",
    "Cooling pads",
    "Fans H1-2",
    "Fans H3-4",
    "Fans H5",
    "Egg belt H1",
    "Egg belt H2",
    "Egg belt H3-4",
    "Manure belt H1-2",
    "Manure belt H3-4",
    "Cross-conveyor",
    "Feed hoppers",
    "Solar panels",
    "Lighting system",
]

# Chart of Accounts
CHART_OF_ACCOUNTS = [
    ("4000", "Egg Sales", "Revenue", "Revenue — Primary"),
    ("4010", "Manure Sales", "Revenue", "Revenue — Secondary"),
    ("4020", "Culled Bird Sales", "Revenue", "Revenue — Secondary"),
    ("5000", "Feed Ingredients", "COGS", "COGS — Feed"),
    ("5010", "Premix & Supplements", "COGS", "COGS — Feed"),
    ("5020", "Medications", "COGS", "COGS — Health"),
    ("5030", "Feed Mixing Costs", "COGS", "COGS — Feed"),
    ("6000", "Labor — Wages", "Opex", "Opex — Labor"),
    ("6010", "Staff Food", "Opex", "Opex — Labor"),
    ("6020", "Fuel", "Opex", "Opex — Energy"),
    ("6030", "Repairs & Maintenance", "Opex", "Opex — Maintenance"),
    ("6040", "Utilities (Water, Electric)", "Opex", "Opex — Utilities"),
    ("6050", "Supplies (PPE, Chemicals, Packaging)", "Opex", "Opex — Supplies"),
    ("6060", "Transport", "Opex", "Opex — Logistics"),
    ("6070", "Other Operating Expenses", "Opex", "Opex — Other"),
    ("7000", "Equipment Purchases", "Capex", "Capex"),
    ("7010", "Infrastructure", "Capex", "Capex"),
]

# Dropdown value lists (in-cell lists for simple dropdowns)
DROPDOWN_LISTS = {
    "sides": ["A", "B"],
    "mortality_cause": ["Culled", "Disease", "Injury", "Unknown"],
    "disease_subcategory": ["Respiratory", "Digestive", "Neurological", "Other"],
    "cluster_pattern": ["Yes", "No"],
    "mortality_action": ["Disposed", "Lab sample", "Vet called", "Isolated section"],
    "yes_no": ["Yes", "No"],
    "feeding_round": ["AM-40%", "PM-30%", "EVE-30%"],
    "payment_method": ["MoMo", "Bank Transfer", "Cash"],
    "payment_status": ["Paid", "Part-Paid", "Unpaid"],
    "delivery_status": ["Dispatched", "Pending", "Pickup-Confirmed"],
    "egg_product": ["Egg crates", "Egg singles", "Manure (bags)", "Manure (bulk)", "Culled birds"],
    "egg_grade_sold": ["Mixed", "Large", "Medium", "Small"],
    "procurement_category": ["Feed ingredient", "Medication", "Fuel", "Repairs", "Labor", "Capex", "Utilities", "Supplies", "Other"],
    "approval_status": ["Not Required", "Pending", "Approved", "Rejected"],
    "equipment_status": ["Green", "Yellow", "Red"],
    "equipment_list": EQUIPMENT_LIST,
    "cooling_status": ["Operational", "Degraded", "Failed"],
    "water_status": ["Normal", "Low pressure", "Failure"],
    "feed_refusal": ["Full", "Normal", "Low"],
    "cause_category_equip": ["Mechanical failure", "Electrical failure", "Power outage", "Wear-and-tear", "Operator error", "External"],
    "impact_on_birds": ["None", "Minor", "Major", "Critical"],
    "resolution_status": ["Resolved", "In-progress", "Awaiting parts", "Awaiting specialist"],
    "technician": ["Maintenance Tech", "Specialist (on-call)", "Other"],
    "visitor_reason": ["Delivery", "Pickup", "Maintenance", "Inspection", "Vet visit", "Other"],
    "zone_access": ["Zone 1 (birds)", "Zone 2 (feed/stores)", "Zone 3 (perimeter only)"],
    "symptom_category": ["Respiratory", "Digestive", "Neurological", "Reproductive", "Skin-feather", "Behavioral", "Unknown"],
    "severity_1_5": ["1", "2", "3", "4", "5"],
    "health_action": ["Monitoring", "Isolated", "Treated", "Vet called", "Sample sent"],
    "health_resolution": ["Open", "Resolved", "Escalated"],
    "med_type": ["Vaccination", "Medication", "Supplement"],
    "med_route": ["Drinking water", "Spray", "Injection", "Feed additive"],
    "shift": ["Day", "Night", "Off-rotation"],
    "team": ["Team A", "Team B", "Non-rotating", "Management"],
    "violations": ["None", "Late arrival", "Task incomplete", "SOP breach", "Insubordination", "Other"],
    "strike_level": ["None", "Yellow", "Orange"],
    "count_trigger": ["Scheduled", "Random", "Handover", "Investigation", "Monthly-independent"],
    "movement_type": ["Received", "Issued-to-Mixing", "Adjusted", "Wasted", "Transfer"],
    "ingredient_category": ["Ingredient", "Finished Feed", "Medication", "Supplement", "Fuel", "Tools", "Spare parts", "Packaging", "PPE", "Chemicals", "Other"],
    "risk_class": ["A", "B", "C"],
    "override_scope": ["Cohort-specific", "Breed-wide"],
    "override_metric": ["Lay %", "Egg weight", "Feed intake", "FCR", "Mortality", "Large %"],
    "fraud_status": ["Open", "Investigating", "Resolved", "False-positive"],
    "fraud_severity": ["Red", "Yellow"],
    "power_source": ["Grid", "Solar", "Generator", "None"],
    "customer_type": ["Wholesale", "Regular Retail", "Walk-in", "Agent"],
    "staff_role": ["Site Coordinator", "Team Leader", "Production Hand", "Feed Handler",
                    "Cleaner", "Night Watch", "Maintenance Tech", "Cook", "Manure Crew", "BDS"],
    "staff_status": ["Active", "Suspended", "Terminated"],
    "flock_status": ["Active", "Closed"],
    "qc_visual": ["Pass", "Fail"],
    "vendor_category": ["Feed ingredients", "Medications", "Fuel", "Repairs", "Equipment", "Utilities", "Other"],
}
