"""
tab6_master_data.py -- Build Tab 6: Master Data (single sheet, 11 tables).
Replaces the old layer2_master_data.py which spread 11 tables across 11 tabs.
All master data tables now live on one "Master Data" sheet, stacked vertically
with section headers and 2-row gaps between them.

Returns master_rows dict mapping table_name -> (header_row, end_row) so that
downstream code can point dropdown validations at the correct ranges.
"""

from datetime import date
from openpyxl.utils import get_column_letter

from . import config as C
from .config import T
from .helpers import (
    create_excel_table, add_inline_dropdown, protect_sheet,
    freeze_panes, set_col_widths, write_section_header, create_named_range,
    add_text_status_cf,
)


# ---------------------------------------------------------------------------
# Individual table builders -- each returns (end_row, max_cols) and updates
# master_rows in place.  current_row is the first available row on the sheet.
# ---------------------------------------------------------------------------

def _build_config_table(ws, current_row, master_rows):
    """Table 1: tblConfig -- system parameters and alert thresholds."""
    table_name = T("config")
    headers = ["Section", "Parameter", "Value", "Unit", "Description"]
    max_cols = len(headers)

    rows = [
        # Global parameters
        ("Global", "Farm Name", C.FARM_NAME, "", ""),
        ("Global", "Farm Location", C.FARM_LOCATION, "", ""),
        ("Global", "Currency", C.CURRENCY, "", ""),
        ("Global", "Crate Size", C.CRATE_SIZE, "eggs", "Standard crate = 30 eggs"),
        ("Global", "Feeding Schedule AM", 0.40, "%", "40% morning feed"),
        ("Global", "Feeding Schedule PM", 0.30, "%", "30% afternoon feed"),
        ("Global", "Feeding Schedule EVE", 0.30, "%", "30% evening feed"),
        ("Global", "Target Intake Per Bird", 118, "g/day", "Default 118g (range 115-120g)"),
        ("Global", "Working Days Per Week", 6, "days", ""),
        # Alert thresholds
        ("Alert", "Mortality Daily Yellow", C.THRESHOLDS["mortality_daily_yellow"], "birds", "4-6 birds"),
        ("Alert", "Mortality Daily Red", C.THRESHOLDS["mortality_daily_red"], "birds", "> 6 birds OR cluster"),
        ("Alert", "Lay Rate Yellow", C.THRESHOLDS["lay_rate_yellow"], "decimal", "< 88% for 2 days"),
        ("Alert", "Lay Rate Red", C.THRESHOLDS["lay_rate_red"], "decimal", "< 85% for 2 consecutive days"),
        ("Alert", "Cash Mismatch Yellow (crates)", C.THRESHOLDS["cash_mismatch_yellow_crates"], "crates", "1-2 crates discrepancy"),
        ("Alert", "Cash Mismatch Red (crates)", C.THRESHOLDS["cash_mismatch_red_crates"], "crates", "> 2 crates OR any cash gap"),
        ("Alert", "Feed Stock Yellow (days)", C.THRESHOLDS["feed_stock_yellow_days"], "days", "3-7 days on hand"),
        ("Alert", "Feed Stock Red (days)", C.THRESHOLDS["feed_stock_red_days"], "days", "< 3 days on hand"),
        ("Alert", "Egg Stock Yellow (crates)", C.THRESHOLDS["egg_stock_yellow_crates"], "crates", "500-1000 crates"),
        ("Alert", "Egg Stock Red (crates)", C.THRESHOLDS["egg_stock_red_crates"], "crates", "> 1000 crates"),
        ("Alert", "Equipment Uptime Yellow", C.THRESHOLDS["equipment_uptime_yellow"], "decimal", "80-90%"),
        ("Alert", "Equipment Uptime Red", C.THRESHOLDS["equipment_uptime_red"], "decimal", "< 80%"),
        ("Alert", "Cracked % Yellow", C.THRESHOLDS["cracked_pct_yellow"], "decimal", "> 6% for 2 consecutive days"),
        ("Alert", "Large % Yellow (Peak)", C.THRESHOLDS["large_pct_yellow_peak"], "decimal", "< 60% for 3 days"),
        ("Alert", "Large % Yellow (Post-peak)", C.THRESHOLDS["large_pct_yellow_postpeak"], "decimal", "< 55% for 3 days"),
        ("Alert", "FCR Yellow (Peak)", C.THRESHOLDS["fcr_peak_yellow"], "kg/dz", "> 2.2 for 3 days"),
        ("Alert", "FCR Red (Peak)", C.THRESHOLDS["fcr_peak_red"], "kg/dz", "> 2.5 for 3 days"),
        ("Alert", "Disease Mortality Immediate Red", C.THRESHOLDS["disease_mortality_immediate_red"], "birds/day", "Absolute trigger"),
        ("Alert", "Heat Stress Hours Red", C.THRESHOLDS["heat_stress_hours_red"], "hours", "> 2 hrs in any zone"),
        ("Alert", "Heat Stress Day Yellow", C.THRESHOLDS["heat_stress_hours_day_yellow"], "hours/day", "> 6 cumulative hours"),
        ("Alert", "Water Drop % Yellow", C.THRESHOLDS["water_drop_pct_yellow"], "decimal", "> 20% drop vs 7-day avg"),
        ("Alert", "Water Spike % Yellow", C.THRESHOLDS["water_spike_pct_yellow"], "decimal", "> 30% spike vs 7-day avg"),
        ("Alert", "Biosecurity Score Yellow", C.THRESHOLDS["biosecurity_score_yellow"], "decimal", "< 80% compliance"),
        ("Alert", "Min Crew (Attendance Red)", C.THRESHOLDS["attendance_min_crew"], "staff", "< 6 present"),
        # Fraud detection
        ("Fraud", "Price Deviation Tolerance", C.THRESHOLDS["price_deviation_tolerance_pct"], "decimal", "15% tolerance"),
        ("Fraud", "Procurement Approval Threshold", C.THRESHOLDS["procurement_approval_threshold"], "GHS", "Above this = requires approval"),
        ("Fraud", "Cash Deposit Window", C.THRESHOLDS["cash_deposit_window_hours"], "hours", "Cash must be deposited within this"),
        ("Fraud", "Feed Discrepancy Yellow (bags)", C.THRESHOLDS["feed_discrepancy_bags_yellow"], "bags", "> 0.5 bags/day"),
        ("Fraud", "Ingredient Shrinkage Yellow", C.THRESHOLDS["ingredient_shrinkage_yellow"], "decimal", "> 2% over 7 days"),
        ("Fraud", "Ingredient Shrinkage Red", C.THRESHOLDS["ingredient_shrinkage_red"], "decimal", "> 5%"),
        # Cycle count
        ("CycleCount", "Class A Frequency (days)", 7, "days", "High-value items"),
        ("CycleCount", "Class B Frequency (days)", 14, "days", "Moderate-value items"),
        ("CycleCount", "Class C Frequency (days)", 30, "days", "Low-value items"),
        ("CycleCount", "Variance Threshold %", 0.05, "decimal", "Flag if > 5%"),
        # Cull logic
        ("Cull", "Late-lay Start Week", 61, "weeks", ""),
        ("Cull", "Consecutive Below-target Days", 14, "days", "Trigger for prepare status"),
        ("Cull", "Margin Floor Weeks", 4, "weeks", "Below floor for N weeks"),
        ("Cull", "Mortality Slope Threshold", 0.002, "rate", "Sustained slope trigger"),
    ]

    # Section header
    write_section_header(ws, current_row, 1, "CONFIGURATION & THRESHOLDS", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[12, 35, 12, 10, 45],
    )
    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_staff_table(ws, current_row, master_rows):
    """Table 2: tblStaff -- staff registry."""
    table_name = T("staff")
    headers = ["Staff ID", "Name", "Role", "Team", "Hire Date", "Phone", "Status"]
    max_cols = len(headers)

    rows = [[s[0], s[1], s[2], s[3], s[4], s[5], s[6]] for s in C.STAFF_DATA]

    write_section_header(ws, current_row, 1, "STAFF REGISTRY", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[10, 20, 18, 14, 12, 16, 10],
    )
    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_housing_table(ws, current_row, master_rows):
    """Table 3: tblHousing -- housing / cage map."""
    table_name = T("housing")
    headers = [
        "Cage ID", "House", "Zone", "Row", "Side", "Bird Count",
        "Cohort ID", "Active", "Total Birds", "Equipment Status",
    ]
    max_cols = len(headers)

    rows = []
    for h in C.HOUSING_DATA:
        rows.append([h[0], h[1], h[2], h[3], h[4], h[5], h[6], h[7], None, "Green"])

    # Total Birds = SUMIFS of Bird Count per House where Active="Yes"
    calculated = {8: 'SUMIFS([Bird Count],[House],[@House],[Active],"Yes")'}

    write_section_header(ws, current_row, 1, "HOUSING MAP", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[10, 12, 10, 6, 6, 12, 12, 8, 12, 14],
        calculated_columns=calculated,
    )

    # Equipment Status column is column J (index 10 = start_col + 9)
    equip_col_letter = get_column_letter(1 + 9)  # J
    add_text_status_cf(ws, f"{equip_col_letter}{current_row + 1}:{equip_col_letter}{end_row}")

    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_flock_table(ws, current_row, master_rows):
    """Table 4: tblFlock -- flock / cohort registry with calculated columns."""
    table_name = T("flock")
    headers = [
        "Cohort ID", "Breed", "Arrival Date", "Initial Count", "Status",
        "Current Bird Count", "Current Age (weeks)", "Production Phase",
        "Projected Cull Start", "Projected Cull End", "Cull Status", "Cull Triggers",
    ]
    max_cols = len(headers)

    rows = []
    for f in C.FLOCK_DATA:
        rows.append([
            f[0], f[1], f[2], f[3], f[4],
            None, None, None,   # calculated columns 5-7
            None, None, "Monitor", "",
        ])

    # Current Bird Count: initial minus deaths for cages belonging to this cohort
    dcl = T("daily_cage_log")  # "tblDailyCageLog"
    housing = T("housing")     # "tblHousing"
    calculated = {
        5: (
            f'[@[Initial Count]]-SUMPRODUCT(({dcl}[Deaths])'
            f'*(IFERROR(INDEX({housing}[Cohort ID],MATCH({dcl}[Cage ID],{housing}[Cage ID],0)),"")'
            f'=[@[Cohort ID]]))'
        ),
        6: 'INT((TODAY()-[@[Arrival Date]])/7)',
        7: (
            'IF([@[Current Age (weeks)]]<25,"Ramp-up",'
            'IF([@[Current Age (weeks)]]<=45,"Peak",'
            'IF([@[Current Age (weeks)]]<=60,"Post-peak","Late-lay")))'
        ),
    }

    write_section_header(ws, current_row, 1, "FLOCK REGISTRY", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[12, 22, 12, 12, 10, 16, 16, 14, 14, 14, 12, 20],
        calculated_columns=calculated,
    )
    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_breeder_curves_table(ws, current_row, master_rows):
    """Table 5: tblBreederCurves -- Lohmann Brown Classic performance by age."""
    table_name = T("breeder_curves")
    headers = [
        "Week", "Expected Lay %", "Expected Egg Wt (g)", "Expected Large %",
        "Expected Feed Intake (g/bird/day)", "Expected FCR (kg/dozen)",
        "Mortality Band (% monthly)", "Phase",
    ]
    max_cols = len(headers)

    rows = [list(entry) for entry in C.BREEDER_CURVE]

    write_section_header(ws, current_row, 1, "BREEDER CURVES -- LOHMANN BROWN CLASSIC", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[8, 14, 18, 14, 28, 22, 22, 12],
        number_formats={
            1: C.FMT_PERCENT,
            3: C.FMT_PERCENT,
            5: C.FMT_DECIMAL_2,
            6: C.FMT_PERCENT,
        },
    )
    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_owner_overrides_table(ws, current_row, master_rows):
    """Table 6: tblOwnerOverrides -- owner target overrides."""
    table_name = T("owner_overrides")
    headers = [
        "Override ID", "Scope", "Cohort ID", "Breed", "Metric",
        "Start Week", "End Week", "Override Target", "Min Band", "Max Band",
        "Reason", "Set By", "Set Date",
    ]
    max_cols = len(headers)

    rows = [
        [1, "Cohort-specific", "FL2024A", "Lohmann Brown Classic", "Lay %",
         45, 60, 0.88, 0.85, 0.92,
         "Adjusted for Ghana heat conditions", "Owner", "2025-12-01"],
        [2, "Breed-wide", "", "Lohmann Brown Classic", "Feed intake",
         25, 80, 120, 115, 125,
         "Tropical climate adjustment", "Owner", "2025-12-01"],
    ]

    write_section_header(ws, current_row, 1, "OWNER TARGET OVERRIDES", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[10, 14, 12, 22, 12, 10, 10, 14, 10, 10, 30, 10, 12],
    )
    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_customer_master_table(ws, current_row, master_rows):
    """Table 7: tblCustomerMaster -- merged CRM + Profile."""
    table_name = T("customer_master")
    headers = [
        "Customer ID", "Name", "Type", "Contact Person", "Phone",
        "Location", "Credit Allowed", "Credit Limit (GHS)",
        "Typical Order (crates/wk)", "Reliability (1-5)",
        "Price Sensitivity (1-5)", "Preferred Payment",
        "Churn Risk", "Customer Since",
        "Last Order Date", "Lifetime Crates", "Notes",
    ]
    max_cols = len(headers)

    rows = []
    for c in C.CUSTOMER_DATA:
        rows.append([
            c[0], c[1], c[2], c[3], c[4], c[5],
            c[6], c[7], c[8], c[9], c[10], c[11],
            None,            # col 12 -- Churn Risk (formula)
            "2024-01-15",    # col 13 -- Customer Since
            None,            # col 14 -- Last Order Date (to be filled)
            None,            # col 15 -- Lifetime Crates (to be filled)
            "",              # col 16 -- Notes
        ])

    # Churn Risk: Red if > 14 days since last order, Yellow if > 7, else Green
    calculated = {
        12: (
            'IF([@[Last Order Date]]="","N/A",'
            'IF(TODAY()-[@[Last Order Date]]>14,"Red",'
            'IF(TODAY()-[@[Last Order Date]]>7,"Yellow","Green")))'
        ),
    }

    write_section_header(ws, current_row, 1, "CUSTOMER MASTER", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[10, 22, 14, 16, 16, 18, 12, 14, 18, 12, 14, 16, 10, 14, 14, 14, 20],
        calculated_columns=calculated,
        number_formats={7: C.FMT_CURRENCY},
    )

    # Churn Risk conditional formatting (column M relative to table start)
    churn_col_letter = get_column_letter(1 + 12)  # M
    add_text_status_cf(ws, f"{churn_col_letter}{current_row + 1}:{churn_col_letter}{end_row}")

    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_vendor_table(ws, current_row, master_rows):
    """Table 8: tblVendor -- vendor / supplier master."""
    table_name = T("vendor")
    headers = [
        "Vendor ID", "Name", "Category", "Contact", "Phone",
        "Location", "Payment Terms", "Reliability (1-5)",
        "Price Score (1-5)", "Approved", "Notes",
    ]
    max_cols = len(headers)

    rows = []
    for v in C.VENDOR_DATA:
        rows.append([v[0], v[1], v[2], v[3], v[4], v[5], v[6], v[7], v[8], v[9], ""])

    write_section_header(ws, current_row, 1, "VENDOR / SUPPLIER MASTER", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[10, 22, 16, 14, 16, 14, 12, 14, 14, 10, 20],
    )
    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_item_master_table(ws, current_row, master_rows):
    """Table 9: tblItemMaster -- inventory item master."""
    table_name = T("item_master")
    headers = [
        "Item ID", "Item Name", "Category", "Unit", "Bag Weight (kg)",
        "Reorder Threshold (units)", "Reorder Threshold (days)",
        "Risk Class", "Preferred Supplier", "Storage Location", "Notes",
    ]
    max_cols = len(headers)

    rows = []
    for item in C.ITEM_DATA:
        rows.append([
            item[0], item[1], item[2], item[3], item[4],
            item[5], item[6], item[7], item[8] or "",
            "Main Store", "",
        ])

    write_section_header(ws, current_row, 1, "INVENTORY ITEM MASTER", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[10, 25, 14, 8, 14, 20, 20, 10, 16, 14, 20],
    )
    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_feed_formulation_table(ws, current_row, master_rows):
    """Table 10: tblFeedFormulation -- formula summary."""
    table_name = T("feed_formulation")
    headers = [
        "Formula ID", "Name", "Age Range", "Target ME (kcal/kg)",
        "Target CP (%)", "Target Calcium (%)", "Target Avail P (%)",
    ]
    max_cols = len(headers)

    rows = []
    for fid, fdata in C.FEED_FORMULAS.items():
        rows.append([
            fid, fdata["name"], fdata["age_range"],
            fdata["target_ME_kcal"], fdata["target_CP_pct"],
            fdata["target_calcium_pct"], fdata["target_phos_pct"],
        ])

    write_section_header(ws, current_row, 1, "FEED FORMULATION -- SUMMARY", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[12, 28, 14, 18, 12, 16, 16],
    )
    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


def _build_formula_ingredients_table(ws, current_row, master_rows):
    """Table 11: tblFormulaIngredients -- ingredient proportions per formula."""
    table_name = T("formula_ingredients")
    headers = [
        "Formula ID", "Ingredient Code", "Ingredient Name",
        "Target %", "Min %", "Max %",
    ]
    max_cols = len(headers)

    rows = []
    for fid, fdata in C.FEED_FORMULAS.items():
        for ing in fdata["ingredients"]:
            rows.append([fid, ing[0], ing[1], ing[2], ing[3], ing[4]])

    write_section_header(ws, current_row, 1, "FEED FORMULATION -- INGREDIENT PROPORTIONS", merge_end_col=max_cols)
    current_row += 1

    tab, end_row = create_excel_table(
        ws, table_name, headers, rows, start_row=current_row,
        col_widths=[12, 14, 22, 10, 8, 8],
        number_formats={
            3: C.FMT_DECIMAL_1,
            4: C.FMT_DECIMAL_1,
            5: C.FMT_DECIMAL_1,
        },
    )
    master_rows[table_name] = (current_row, end_row)
    return end_row, max_cols


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def build_tab6_master_data(wb):
    """Build Tab 6: Master Data -- 11 tables on one sheet. Returns master_rows dict."""
    ws = wb.create_sheet(title=C.TAB_NAMES[6])
    ws.sheet_properties.tabColor = C.TAB_COLORS[6]

    master_rows = {}
    current_row = 1

    # --- Table 1: Config & Thresholds ---
    end_row, _ = _build_config_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 2: Staff Registry ---
    end_row, _ = _build_staff_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 3: Housing Map ---
    end_row, _ = _build_housing_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 4: Flock Registry ---
    end_row, _ = _build_flock_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 5: Breeder Curves ---
    end_row, _ = _build_breeder_curves_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 6: Owner Overrides ---
    end_row, _ = _build_owner_overrides_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 7: Customer Master ---
    end_row, _ = _build_customer_master_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 8: Vendor Master ---
    end_row, _ = _build_vendor_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 9: Item Master ---
    end_row, _ = _build_item_master_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 10: Feed Formulation Summary ---
    end_row, _ = _build_feed_formulation_table(ws, current_row, master_rows)
    current_row = end_row + 3

    # --- Table 11: Formula Ingredients ---
    end_row, _ = _build_formula_ingredients_table(ws, current_row, master_rows)

    # ------------------------------------------------------------------
    # Named ranges -- use actual row positions from master_rows
    # All ranges reference the single "Master Data" sheet.
    # ------------------------------------------------------------------
    sheet = C.MASTER_SHEET

    # tblStaff: Staff IDs (col A) and Names (col B)
    staff_hdr, staff_end = master_rows[T("staff")]
    create_named_range(wb, "StaffIDs", sheet, f"$A${staff_hdr + 1}:$A${staff_end}")
    create_named_range(wb, "StaffNames", sheet, f"$B${staff_hdr + 1}:$B${staff_end}")

    # tblHousing: Cage IDs (col A) and House Names (col B)
    housing_hdr, housing_end = master_rows[T("housing")]
    create_named_range(wb, "CageIDs", sheet, f"$A${housing_hdr + 1}:$A${housing_end}")
    create_named_range(wb, "HouseNames", sheet, f"$B${housing_hdr + 1}:$B${housing_end}")

    # tblFlock: Flock / Cohort IDs (col A)
    flock_hdr, flock_end = master_rows[T("flock")]
    create_named_range(wb, "FlockIDs", sheet, f"$A${flock_hdr + 1}:$A${flock_end}")

    # tblCustomerMaster: Customer IDs (col A) and Names (col B)
    cust_hdr, cust_end = master_rows[T("customer_master")]
    create_named_range(wb, "CustomerIDs", sheet, f"$A${cust_hdr + 1}:$A${cust_end}")
    create_named_range(wb, "CustomerNames", sheet, f"$B${cust_hdr + 1}:$B${cust_end}")

    # tblVendor: Vendor IDs (col A) and Names (col B)
    vendor_hdr, vendor_end = master_rows[T("vendor")]
    create_named_range(wb, "VendorIDs", sheet, f"$A${vendor_hdr + 1}:$A${vendor_end}")
    create_named_range(wb, "VendorNames", sheet, f"$B${vendor_hdr + 1}:$B${vendor_end}")

    # tblItemMaster: Item IDs (col A) and Names (col B)
    item_hdr, item_end = master_rows[T("item_master")]
    create_named_range(wb, "ItemIDs", sheet, f"$A${item_hdr + 1}:$A${item_end}")
    create_named_range(wb, "ItemNames", sheet, f"$B${item_hdr + 1}:$B${item_end}")

    # tblBreederCurves: Weeks (col A)
    bc_hdr, bc_end = master_rows[T("breeder_curves")]
    create_named_range(wb, "BreederWeeks", sheet, f"$A${bc_hdr + 1}:$A${bc_end}")

    # tblFeedFormulation: Formula IDs (col A)
    ff_hdr, ff_end = master_rows[T("feed_formulation")]
    create_named_range(wb, "FormulaIDs", sheet, f"$A${ff_hdr + 1}:$A${ff_end}")

    # ------------------------------------------------------------------
    # Final sheet formatting
    # ------------------------------------------------------------------
    freeze_panes(ws, "A2")
    protect_sheet(ws)

    print("  Tab 6: 11 master data tables + 13 named ranges created on single sheet")
    return master_rows
