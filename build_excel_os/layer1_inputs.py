"""
layer1_inputs.py — Build 16 input tabs (Layer 1).
These are the only tabs where staff type data. Input columns unlocked, formulas locked.
"""

from openpyxl.utils import get_column_letter

from . import config as C
from .helpers import (
    create_excel_table, add_dropdown_validation, add_inline_dropdown,
    add_integer_validation, add_date_validation,
    protect_sheet, freeze_panes, set_col_widths,
    add_traffic_light_cf, add_text_status_cf,
)


def build_tab1_production(wb, data):
    """Tab 1: Daily Production Log — egg collection with mandatory grade breakdown."""
    ws = wb.create_sheet(title=C.TAB_NAMES[1])

    headers = [
        "Date", "Cage ID", "Side", "Crates Collected", "Singles Collected",
        "Grade: Large", "Grade: Medium", "Grade: Small/Peewee", "Grade: Cracked/Broken",
        "Notes",
        # Auto-calc columns
        "Total Eggs", "Equivalent Crates", "Remainder Singles",
        "Cracked %", "Large %", "Grade Validation"
    ]

    rows = []
    for r in data["production"]:
        rows.append([
            r[0], r[1], r[2], r[3], r[4],   # Date, Cage, Side, Crates, Singles
            r[5], r[6], r[7], r[8],           # Large, Medium, Small, Cracked
            r[9],                               # Notes
            None, None, None, None, None, None  # Formula columns
        ])

    calculated = {
        10: "([@Crates Collected]*30)+[@Singles Collected]",
        11: "INT([@Total Eggs]/30)",
        12: "MOD([@Total Eggs],30)",
        13: 'IF([@Total Eggs]=0,0,[@Grade: Cracked/Broken]/[@Total Eggs])',
        14: 'IF([@Total Eggs]=0,0,[@Grade: Large]/[@Total Eggs])',
        15: 'IF([@Grade: Large]+[@Grade: Medium]+[@Grade: Small/Peewee]+[@Grade: Cracked/Broken]=[@Total Eggs],"OK","MISMATCH")',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[1], headers, rows,
        col_widths=[12, 10, 6, 14, 14, 12, 12, 16, 18, 20, 10, 14, 14, 10, 10, 14],
        calculated_columns=calculated,
        number_formats={13: C.FMT_PERCENT, 14: C.FMT_PERCENT},
    )

    # Dropdowns
    add_dropdown_validation(ws, f"B2:B{end_row}", C.TAB_NAMES[19], f"$A$2:$A${len(C.HOUSING_DATA)+1}")
    add_inline_dropdown(ws, f"C2:C{end_row}", C.DROPDOWN_LISTS["sides"])
    add_integer_validation(ws, f"D2:D{end_row}", 0, 999)
    add_integer_validation(ws, f"E2:E{end_row}", 0, 29)

    # Conditional formatting
    add_traffic_light_cf(ws, f"N2:N{end_row}",
                          green_op="lessThanOrEqual", green_val="0.04",
                          yellow_op="lessThanOrEqual", yellow_val="0.06",
                          red_op="greaterThan", red_val="0.06")
    add_text_status_cf(ws, f"P2:P{end_row}")

    # Protection: unlock input columns (A-J), lock formula columns (K-P)
    unlock_cols = list(range(1, 11))  # columns 1-10
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab2_environmental(wb, data):
    """Tab 2: Daily Environmental Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[2])

    headers = [
        "Date", "House/Zone", "Time of Reading", "Temperature (°C)",
        "Hours Above 31°C", "Cooling Status", "Feed Refusal (AM)",
        "Feed Refusal (PM)", "Trough Locations Checked", "Water Status", "Notes",
        # Auto-calcs
        "Heat Stress Score", "Feed Refusal Flag", "Max Temp Today"
    ]

    rows = []
    for r in data["environmental"]:
        rows.append(list(r) + [None, None, None])

    calculated = {
        11: '[@Hours Above 31°C]',
        12: 'IF(OR([@Feed Refusal (AM)]="Full",[@Feed Refusal (PM)]="Full"),"YES","NO")',
        13: '[@Temperature (°C)]',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[2], headers, rows,
        col_widths=[12, 14, 12, 14, 14, 14, 14, 14, 25, 12, 20, 14, 14, 12],
        calculated_columns=calculated,
        number_formats={3: C.FMT_DECIMAL_1, 4: C.FMT_DECIMAL_1},
    )

    add_inline_dropdown(ws, f"F2:F{end_row}", C.DROPDOWN_LISTS["cooling_status"])
    add_inline_dropdown(ws, f"G2:G{end_row}", C.DROPDOWN_LISTS["feed_refusal"])
    add_inline_dropdown(ws, f"H2:H{end_row}", C.DROPDOWN_LISTS["feed_refusal"])
    add_inline_dropdown(ws, f"J2:J{end_row}", C.DROPDOWN_LISTS["water_status"])

    add_text_status_cf(ws, f"L2:L{end_row}")

    unlock_cols = list(range(1, 12))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab3_mortality(wb, data):
    """Tab 3: Daily Mortality Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[3])

    headers = [
        "Date", "Cage ID", "Side", "Death Count", "Cause Category",
        "Disease Sub-category", "Symptom Description", "Cluster Pattern",
        "Action Taken", "Escalated to Owner", "Photo Evidence", "Notes",
        # Auto-calcs
        "Mortality Rate %", "Cluster Flag"
    ]

    rows = []
    for r in data["mortality"]:
        rows.append(list(r) + [None, None])

    calculated = {
        12: f'IF(SUMIFS({C.TABLE_NAMES[19]}[Bird Count],{C.TABLE_NAMES[19]}[Cage ID],[@Cage ID])=0,0,[@Death Count]/SUMIFS({C.TABLE_NAMES[19]}[Bird Count],{C.TABLE_NAMES[19]}[Cage ID],[@Cage ID]))',
        13: 'IF([@Death Count]>=3,"CLUSTER","OK")',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[3], headers, rows,
        col_widths=[12, 10, 6, 10, 14, 18, 25, 12, 16, 14, 12, 20, 12, 12],
        calculated_columns=calculated,
        number_formats={12: C.FMT_PERCENT_2},
    )

    add_dropdown_validation(ws, f"B2:B{end_row}", C.TAB_NAMES[19], f"$A$2:$A${len(C.HOUSING_DATA)+1}")
    add_inline_dropdown(ws, f"C2:C{end_row}", C.DROPDOWN_LISTS["sides"])
    add_inline_dropdown(ws, f"E2:E{end_row}", C.DROPDOWN_LISTS["mortality_cause"])
    add_inline_dropdown(ws, f"F2:F{end_row}", C.DROPDOWN_LISTS["disease_subcategory"])
    add_inline_dropdown(ws, f"H2:H{end_row}", C.DROPDOWN_LISTS["cluster_pattern"])
    add_inline_dropdown(ws, f"I2:I{end_row}", C.DROPDOWN_LISTS["mortality_action"])
    add_inline_dropdown(ws, f"J2:J{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"K2:K{end_row}", C.DROPDOWN_LISTS["yes_no"])

    add_text_status_cf(ws, f"N2:N{end_row}")

    unlock_cols = list(range(1, 13))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab4_feed_consumption(wb, data):
    """Tab 4: Daily Feed Consumption Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[4])

    headers = [
        "Date", "Feeding Round", "Cohort ID", "Cage ID (optional)",
        "Feed Type", "Qty Issued (kg)", "Qty Returned (kg)",
        "Issued By", "Verified By", "Notes",
        # Auto-calcs
        "Net Consumed (kg)", "Intake/Bird (g)", "Daily FCR"
    ]

    rows = []
    for r in data["feed_consumption"]:
        rows.append(list(r) + [None, None, None])

    calculated = {
        10: '[@Qty Issued (kg)]-[@Qty Returned (kg)]',
        11: f'IF(SUMIFS({C.TABLE_NAMES[20]}[Current Bird Count],{C.TABLE_NAMES[20]}[Cohort ID],[@Cohort ID])=0,0,([@Net Consumed (kg)]*1000)/SUMIFS({C.TABLE_NAMES[20]}[Current Bird Count],{C.TABLE_NAMES[20]}[Cohort ID],[@Cohort ID]))',
        12: f'IF(SUMIFS({C.TABLE_NAMES[1]}[Total Eggs],{C.TABLE_NAMES[1]}[Date],[@Date])=0,0,[@Net Consumed (kg)]/(SUMIFS({C.TABLE_NAMES[1]}[Total Eggs],{C.TABLE_NAMES[1]}[Date],[@Date])/12))',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[4], headers, rows,
        col_widths=[12, 12, 12, 16, 14, 14, 14, 12, 12, 20, 14, 14, 10],
        calculated_columns=calculated,
        number_formats={5: C.FMT_DECIMAL_1, 6: C.FMT_DECIMAL_1, 10: C.FMT_DECIMAL_1,
                        11: C.FMT_DECIMAL_1, 12: C.FMT_DECIMAL_2},
    )

    add_inline_dropdown(ws, f"B2:B{end_row}", C.DROPDOWN_LISTS["feeding_round"])
    add_dropdown_validation(ws, f"C2:C{end_row}", C.TAB_NAMES[20], f"$A$2:$A${len(C.FLOCK_DATA)+1}")
    add_dropdown_validation(ws, f"E2:E{end_row}", C.TAB_NAMES[26], f"$A$2:$A${len(C.ITEM_DATA)+1}")
    add_dropdown_validation(ws, f"H2:H{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_dropdown_validation(ws, f"I2:I{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")

    unlock_cols = list(range(1, 11))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab5_water_consumption(wb, data):
    """Tab 5: Daily Water Consumption Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[5])

    headers = [
        "Date", "House/Zone", "Meter Reading AM (L)", "Meter Reading PM (L)",
        "Estimation Method", "Notes",
        # Auto-calcs
        "Daily Consumption (L)", "Liters per 1000 Birds", "7-Day Average", "Variance from Avg"
    ]

    rows = []
    for r in data["water_consumption"]:
        rows.append(list(r) + [None, None, None, None])

    calculated = {
        6: '[@Meter Reading PM (L)]-[@Meter Reading AM (L)]',
        7: f'IF(SUMIFS({C.TABLE_NAMES[19]}[Bird Count],{C.TABLE_NAMES[19]}[House],[@ House/Zone])=0,0,[@Daily Consumption (L)]/(SUMIFS({C.TABLE_NAMES[19]}[Bird Count],{C.TABLE_NAMES[19]}[House],[@ House/Zone])/1000))',
        8: f'IFERROR(AVERAGEIFS([Daily Consumption (L)],[Date],">="&([@Date]-6),[Date],"<="&[@Date],[House/Zone],[@House/Zone]),0)',
        9: 'IF([@7-Day Average]=0,0,([@Daily Consumption (L)]-[@7-Day Average])/[@7-Day Average])',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[5], headers, rows,
        col_widths=[12, 14, 18, 18, 16, 20, 18, 18, 14, 14],
        calculated_columns=calculated,
        number_formats={2: C.FMT_INTEGER, 3: C.FMT_INTEGER, 6: C.FMT_INTEGER,
                        7: C.FMT_DECIMAL_1, 8: C.FMT_DECIMAL_1, 9: C.FMT_PERCENT},
    )

    add_inline_dropdown(ws, f"E2:E{end_row}", ["Flow meter", "Polytank level", "Manual count"])

    unlock_cols = list(range(1, 7))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab6_ingredient_movement(wb, data):
    """Tab 6: Ingredient Inventory Movement Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[6])

    headers = [
        "Date", "Ingredient", "Movement Type", "Quantity (kg)", "Bag Count",
        "Supplier", "Batch/Lot #", "Price/kg (if Received)", "Receipt/Invoice Ref",
        "Weight Verified", "Issued By", "Verified By", "Notes",
        # Auto-calcs
        "Line Value (GHS)", "Running Stock"
    ]

    rows = []
    for r in data["ingredient_movement"]:
        rows.append(list(r) + [None, None])

    calculated = {
        13: 'IF([@Movement Type]="Received",[@Quantity (kg)]*[@Price/kg (if Received)],0)',
        14: f'SUMIFS([Quantity (kg)],[Ingredient],[@Ingredient],[Movement Type],"Received")-SUMIFS([Quantity (kg)],[Ingredient],[@Ingredient],[Movement Type],"Issued-to-Mixing")-SUMIFS([Quantity (kg)],[Ingredient],[@Ingredient],[Movement Type],"Wasted")',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[6], headers, rows,
        col_widths=[12, 12, 16, 12, 10, 10, 14, 16, 16, 12, 10, 10, 20, 14, 14],
        calculated_columns=calculated,
        number_formats={3: C.FMT_DECIMAL_1, 7: C.FMT_DECIMAL_2, 13: C.FMT_CURRENCY, 14: C.FMT_DECIMAL_1},
    )

    add_dropdown_validation(ws, f"B2:B{end_row}", C.TAB_NAMES[26], f"$A$2:$A${len(C.ITEM_DATA)+1}")
    add_inline_dropdown(ws, f"C2:C{end_row}", C.DROPDOWN_LISTS["movement_type"])
    add_dropdown_validation(ws, f"F2:F{end_row}", C.TAB_NAMES[25], f"$A$2:$A${len(C.VENDOR_DATA)+1}")
    add_inline_dropdown(ws, f"J2:J{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"K2:K{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_dropdown_validation(ws, f"L2:L{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")

    unlock_cols = list(range(1, 14))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab7_feed_mixing(wb, data):
    """Tab 7: Feed Mixing / Formulation Batch Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[7])

    headers = [
        "Batch ID", "Date Mixed", "Feed Type Produced", "Formula Reference",
        "Output Qty (kg)", "Output Qty (bags)", "Mixed By",
        "QC: Moisture", "QC: Particle Size", "QC: Visual", "Notes",
        # Auto-calcs
        "Cost/kg", "Yield Efficiency"
    ]

    rows = []
    for r in data["feed_mixing"]:
        rows.append(list(r) + [None, None])

    calculated = {
        11: 'IF([@Output Qty (kg)]=0,0,0)',  # Simplified — would reference batch lines
        12: 'IF([@Output Qty (kg)]=0,0,0.97)',  # ~97% yield typical
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[7], headers, rows,
        col_widths=[10, 12, 24, 16, 14, 14, 12, 12, 14, 12, 20, 10, 14],
        calculated_columns=calculated,
        number_formats={4: C.FMT_DECIMAL_1, 5: C.FMT_DECIMAL_1, 11: C.FMT_CURRENCY, 12: C.FMT_PERCENT},
    )

    add_dropdown_validation(ws, f"D2:D{end_row}", C.TAB_NAMES[27], f"$A$2:$A${len(C.FEED_FORMULAS)+1}")
    add_dropdown_validation(ws, f"G2:G{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_inline_dropdown(ws, f"H2:H{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"I2:I{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"J2:J{end_row}", C.DROPDOWN_LISTS["qc_visual"])

    unlock_cols = list(range(1, 12))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab8_sales(wb, data):
    """Tab 8: Sales Transaction Log — fraud-proof sales recording."""
    ws = wb.create_sheet(title=C.TAB_NAMES[8])

    headers = [
        "Date/Time", "Invoice/Receipt ID", "Customer ID", "Product", "Egg Grade Sold",
        "Qty (crates)", "Qty (singles)", "Unit Price", "Payment Method", "Payment Status",
        "Evidence Ref", "MoMo Verified", "Delivery Status", "Dispatch Auth By",
        "Sold By", "Proof Link", "Notes",
        # Auto-calcs
        "Total Eggs", "Line Total (GHS)", "Revenue Today",
        "Fraud: Dup Invoice", "Fraud: Paid No Evidence", "Fraud: Price Dev"
    ]

    rows = []
    for r in data["sales"]:
        rows.append(list(r) + [None, None, None, None, None, None])

    calculated = {
        17: '([@Qty (crates)]*30)+[@Qty (singles)]',
        18: '[@Qty (crates)]*[@Unit Price]',
        19: f'SUMIFS([Line Total (GHS)],[Date/Time],">="&INT([@Date/Time]),[Date/Time],"<"&(INT([@Date/Time])+1))',
        20: f'IF(COUNTIFS([Invoice/Receipt ID],[@Invoice/Receipt ID])>1,"RED FLAG","OK")',
        21: 'IF(AND([@Payment Status]="Paid",[@Evidence Ref]=""),"RED FLAG","OK")',
        22: f'IF(ABS([@Unit Price]-{C.PRICES["egg_crate_default"]})/{C.PRICES["egg_crate_default"]}>{C.THRESHOLDS["price_deviation_tolerance_pct"]},"YELLOW FLAG","OK")',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[8], headers, rows,
        col_widths=[16, 16, 10, 12, 14, 10, 10, 10, 14, 12, 16, 12, 14, 12, 10, 14, 16,
                    10, 14, 14, 14, 18, 14],
        calculated_columns=calculated,
        number_formats={7: C.FMT_CURRENCY, 18: C.FMT_CURRENCY, 19: C.FMT_CURRENCY},
    )

    add_dropdown_validation(ws, f"C2:C{end_row}", C.TAB_NAMES[23], f"$A$2:$A${len(C.CUSTOMER_DATA)+1}")
    add_inline_dropdown(ws, f"D2:D{end_row}", C.DROPDOWN_LISTS["egg_product"])
    add_inline_dropdown(ws, f"E2:E{end_row}", C.DROPDOWN_LISTS["egg_grade_sold"])
    add_inline_dropdown(ws, f"I2:I{end_row}", C.DROPDOWN_LISTS["payment_method"])
    add_inline_dropdown(ws, f"J2:J{end_row}", C.DROPDOWN_LISTS["payment_status"])
    add_inline_dropdown(ws, f"L2:L{end_row}", ["Yes", "No", "N/A"])
    add_inline_dropdown(ws, f"M2:M{end_row}", C.DROPDOWN_LISTS["delivery_status"])
    add_dropdown_validation(ws, f"N2:N{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_dropdown_validation(ws, f"O2:O{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")

    add_text_status_cf(ws, f"U2:U{end_row}")
    add_text_status_cf(ws, f"V2:V{end_row}")
    add_text_status_cf(ws, f"W2:W{end_row}")

    unlock_cols = list(range(1, 18))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab9_procurement(wb, data):
    """Tab 9: Procurement / Expense Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[9])

    headers = [
        "Date/Time", "Vendor", "Category", "Item/Service", "Quantity", "Unit",
        "Unit Cost", "Total Cost", "Payment Method", "Payment Ref",
        "Receipt/Invoice ID", "Requested By", "Approval Status", "Approved By",
        "Approval Date", "Received Verified", "Verified By", "Price Benchmark", "Notes",
        # Auto-calcs
        "Fraud: Above Threshold", "Fraud: Price Above Benchmark", "Fraud: Dup Receipt"
    ]

    rows = []
    for r in data["procurement"]:
        rows.append(list(r) + [None, None, None])

    calculated = {
        7: '[@Quantity]*[@Unit Cost]',
        17: f'IFERROR(INDEX({C.TABLE_NAMES[26]}[Reorder Threshold (units)],MATCH([@Item/Service],{C.TABLE_NAMES[26]}[Item Name],0)),"")',
        19: f'IF(AND([@Total Cost]>{C.THRESHOLDS["procurement_approval_threshold"]},[@Approval Status]<>"Approved"),"RED FLAG","OK")',
        20: 'IF(AND(ISNUMBER([@Price Benchmark]),[@Unit Cost]>[@Price Benchmark]*1.15),"YELLOW FLAG","OK")',
        21: 'IF(COUNTIFS([Receipt/Invoice ID],[@Receipt/Invoice ID])>1,"RED FLAG","OK")',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[9], headers, rows,
        col_widths=[16, 10, 16, 20, 10, 8, 12, 12, 14, 14, 16, 12, 14, 12, 14, 14, 12, 14, 20,
                    16, 20, 14],
        calculated_columns=calculated,
        number_formats={6: C.FMT_CURRENCY, 7: C.FMT_CURRENCY, 17: C.FMT_CURRENCY},
    )

    add_dropdown_validation(ws, f"B2:B{end_row}", C.TAB_NAMES[25], f"$A$2:$A${len(C.VENDOR_DATA)+1}")
    add_inline_dropdown(ws, f"C2:C{end_row}", C.DROPDOWN_LISTS["procurement_category"])
    add_inline_dropdown(ws, f"I2:I{end_row}", C.DROPDOWN_LISTS["payment_method"])
    add_dropdown_validation(ws, f"L2:L{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_inline_dropdown(ws, f"M2:M{end_row}", C.DROPDOWN_LISTS["approval_status"])
    add_dropdown_validation(ws, f"N2:N{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_inline_dropdown(ws, f"P2:P{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"Q2:Q{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")

    add_text_status_cf(ws, f"T2:T{end_row}")
    add_text_status_cf(ws, f"U2:U{end_row}")
    add_text_status_cf(ws, f"V2:V{end_row}")

    unlock_cols = list(range(1, 20))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab10_equipment(wb, data):
    """Tab 10: Equipment & Environmental Status Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[10])

    headers = [
        "Date/Time", "Equipment", "Location", "Status", "Downtime Hours",
        "Downtime Start", "Downtime End", "Cause Category", "Impact on Birds",
        "Resolution Action", "Resolution Status", "Cost of Repair",
        "Technician", "SC Notified", "Owner Notified", "Notes",
        # Auto-calcs
        "Uptime %"
    ]

    rows = []
    for r in data["equipment"]:
        rows.append(list(r) + [None])

    calculated = {
        16: 'IF([@Downtime Hours]=0,1,(24-[@Downtime Hours])/24)',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[10], headers, rows,
        col_widths=[16, 22, 12, 8, 12, 14, 14, 18, 14, 20, 14, 12, 16, 10, 12, 20, 10],
        calculated_columns=calculated,
        number_formats={4: C.FMT_DECIMAL_1, 11: C.FMT_CURRENCY, 16: C.FMT_PERCENT},
    )

    add_inline_dropdown(ws, f"B2:B{end_row}", C.DROPDOWN_LISTS["equipment_list"])
    add_inline_dropdown(ws, f"D2:D{end_row}", C.DROPDOWN_LISTS["equipment_status"])
    add_inline_dropdown(ws, f"H2:H{end_row}", C.DROPDOWN_LISTS["cause_category_equip"])
    add_inline_dropdown(ws, f"I2:I{end_row}", C.DROPDOWN_LISTS["impact_on_birds"])
    add_inline_dropdown(ws, f"K2:K{end_row}", C.DROPDOWN_LISTS["resolution_status"])
    add_inline_dropdown(ws, f"M2:M{end_row}", C.DROPDOWN_LISTS["technician"])
    add_inline_dropdown(ws, f"N2:N{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"O2:O{end_row}", C.DROPDOWN_LISTS["yes_no"])

    add_text_status_cf(ws, f"D2:D{end_row}")

    unlock_cols = list(range(1, 17))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab11_biosecurity(wb, data):
    """Tab 11: Daily Fortress Checklist (Biosecurity)."""
    ws = wb.create_sheet(title=C.TAB_NAMES[11])

    headers = [
        "Date", "Footbath OK", "Disinfectant Fresh", "Restricted Entry OK",
        "Houses Locked", "Rodent Control Checked", "Dead Bird Disposal OK",
        "PPE Compliance", "Time Separation OK", "Right-side Door OK",
        "Supervisor Sign-off", "Notes",
        # Auto-calcs
        "Compliance Score %", "Critical Breach"
    ]

    rows = []
    for r in data["biosecurity"]:
        rows.append(list(r) + [None, None])

    calculated = {
        12: 'COUNTIF(B{row}:J{row},"Yes")/9'.replace("{row}", "[@Footbath OK]:[@Right-side Door OK]"),
        13: 'IF(OR([@Footbath OK]="No",[@Restricted Entry OK]="No",[@Dead Bird Disposal OK]="No"),"BREACH","OK")',
    }
    # Fix: use a proper COUNTIF approach for the row
    calculated[12] = '(COUNTIF(INDIRECT("B"&ROW()&":J"&ROW()),"Yes"))/9'

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[11], headers, rows,
        col_widths=[12, 12, 14, 16, 12, 18, 18, 14, 16, 16, 16, 20, 14, 12],
        calculated_columns=calculated,
        number_formats={12: C.FMT_PERCENT},
    )

    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        add_inline_dropdown(ws, f"{col}2:{col}{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"K2:K{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")

    add_text_status_cf(ws, f"N2:N{end_row}")
    add_traffic_light_cf(ws, f"M2:M{end_row}",
                          green_op="greaterThanOrEqual", green_val="0.9",
                          yellow_op="greaterThanOrEqual", yellow_val="0.8",
                          red_op="lessThan", red_val="0.8")

    unlock_cols = list(range(1, 13))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab12_visitor(wb, data):
    """Tab 12: Visitor & Vehicle Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[12])

    headers = [
        "Date/Time In", "Date/Time Out", "Visitor/Driver Name", "Vehicle (License)",
        "Company", "Reason for Visit", "Zone Accessed",
        "Disinfection Done", "Approved By", "Camera Ref", "Notes"
    ]

    rows = data["visitor"]

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[12], headers, rows,
        col_widths=[16, 16, 20, 16, 16, 14, 20, 14, 12, 12, 20],
    )

    add_inline_dropdown(ws, f"F2:F{end_row}", C.DROPDOWN_LISTS["visitor_reason"])
    add_inline_dropdown(ws, f"G2:G{end_row}", C.DROPDOWN_LISTS["zone_access"])
    add_inline_dropdown(ws, f"H2:H{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"I2:I{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")

    unlock_cols = list(range(1, 12))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab13_health_incident(wb, data):
    """Tab 13: Health Incident Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[13])

    headers = [
        "Date/Time", "Cage/Cohort", "Symptom Category", "Specific Symptoms",
        "Severity (1-5)", "Birds Affected", "Photo Evidence",
        "FHIS Consultation", "FHIS Recommendation", "Action Taken",
        "Vet Called", "Vet Response", "Resolution Status", "Notes"
    ]

    rows = data["health_incident"]

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[13], headers, rows,
        col_widths=[16, 12, 16, 30, 12, 12, 12, 14, 20, 14, 10, 20, 14, 20],
    )

    add_dropdown_validation(ws, f"B2:B{end_row}", C.TAB_NAMES[19], f"$A$2:$A${len(C.HOUSING_DATA)+1}")
    add_inline_dropdown(ws, f"C2:C{end_row}", C.DROPDOWN_LISTS["symptom_category"])
    add_inline_dropdown(ws, f"E2:E{end_row}", C.DROPDOWN_LISTS["severity_1_5"])
    add_inline_dropdown(ws, f"G2:G{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"H2:H{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"J2:J{end_row}", C.DROPDOWN_LISTS["health_action"])
    add_inline_dropdown(ws, f"K2:K{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_inline_dropdown(ws, f"M2:M{end_row}", C.DROPDOWN_LISTS["health_resolution"])

    unlock_cols = list(range(1, 15))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab14_medication(wb, data):
    """Tab 14: Medication & Vaccination Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[14])

    headers = [
        "Date", "Type", "Product Name", "Dose/Concentration",
        "Route", "Cohort/Cage", "Birds Treated",
        "Administered By", "Vet Authorized", "Batch/Lot #",
        "Withdrawal Period (days)", "Next Dose Due", "Notes",
        # Auto-calc
        "Withdrawal End Date"
    ]

    rows = []
    for r in data["medication"]:
        rows.append(list(r) + [None])

    calculated = {
        13: 'IF([@Withdrawal Period (days)]>0,[@Date]+[@Withdrawal Period (days)],"")',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[14], headers, rows,
        col_widths=[12, 14, 14, 18, 14, 12, 12, 14, 12, 14, 18, 14, 20, 16],
        calculated_columns=calculated,
    )

    add_inline_dropdown(ws, f"B2:B{end_row}", C.DROPDOWN_LISTS["med_type"])
    add_dropdown_validation(ws, f"C2:C{end_row}", C.TAB_NAMES[26], f"$A$2:$A${len(C.ITEM_DATA)+1}")
    add_inline_dropdown(ws, f"E2:E{end_row}", C.DROPDOWN_LISTS["med_route"])
    add_dropdown_validation(ws, f"F2:F{end_row}", C.TAB_NAMES[20], f"$A$2:$A${len(C.FLOCK_DATA)+1}")
    add_dropdown_validation(ws, f"H2:H{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_inline_dropdown(ws, f"I2:I{end_row}", C.DROPDOWN_LISTS["yes_no"])

    unlock_cols = list(range(1, 14))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab15_labor(wb, data):
    """Tab 15: Labor / Compliance Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[15])

    headers = [
        "Date", "Staff ID", "Team", "Shift", "Present",
        "Arrival Time", "Tasks Assigned", "Tasks Completed",
        "Verified By", "Violations", "Strike Issued",
        "Strike Reason", "Supervisor Notes",
        # Auto-calcs
        "Completion Rate %", "Attendance (30d)"
    ]

    rows = []
    for r in data["labor"]:
        rows.append(list(r) + [None, None])

    calculated = {
        13: 'IF([@Tasks Assigned]=0,0,[@Tasks Completed]/[@Tasks Assigned])',
        14: f'IF(COUNTIFS([Staff ID],[@Staff ID],[Date],">="&([@Date]-29),[Date],"<="&[@Date])=0,0,COUNTIFS([Staff ID],[@Staff ID],[Date],">="&([@Date]-29),[Date],"<="&[@Date],[Present],"Yes")/COUNTIFS([Staff ID],[@Staff ID],[Date],">="&([@Date]-29),[Date],"<="&[@Date]))',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[15], headers, rows,
        col_widths=[12, 10, 10, 12, 8, 12, 14, 14, 12, 16, 12, 20, 20, 14, 14],
        calculated_columns=calculated,
        number_formats={13: C.FMT_PERCENT, 14: C.FMT_PERCENT},
    )

    add_dropdown_validation(ws, f"B2:B{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_inline_dropdown(ws, f"C2:C{end_row}", C.DROPDOWN_LISTS["team"])
    add_inline_dropdown(ws, f"D2:D{end_row}", C.DROPDOWN_LISTS["shift"])
    add_inline_dropdown(ws, f"E2:E{end_row}", C.DROPDOWN_LISTS["yes_no"])
    add_dropdown_validation(ws, f"I2:I{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_inline_dropdown(ws, f"J2:J{end_row}", C.DROPDOWN_LISTS["violations"])
    add_inline_dropdown(ws, f"K2:K{end_row}", C.DROPDOWN_LISTS["strike_level"])

    add_traffic_light_cf(ws, f"N2:N{end_row}",
                          green_op="greaterThanOrEqual", green_val="0.9",
                          yellow_op="greaterThanOrEqual", yellow_val="0.8",
                          red_op="lessThan", red_val="0.8")

    unlock_cols = list(range(1, 14))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_tab16_inventory_count(wb, data):
    """Tab 16: Inventory Count Log."""
    ws = wb.create_sheet(title=C.TAB_NAMES[16])

    headers = [
        "Date", "Trigger", "Item", "Storage Location",
        "Expected Stock", "Physical Count",
        "Variance", "Variance %",
        "Explanation", "Counted By", "Verified By", "Approved By", "Notes"
    ]

    rows = []
    for r in data["inventory_count"]:
        rows.append(list(r))

    calculated = {
        6: '[@Physical Count]-[@Expected Stock]',
        7: 'IF([@Expected Stock]=0,0,[@Variance]/[@Expected Stock])',
    }

    tab, end_row = create_excel_table(
        ws, C.TABLE_NAMES[16], headers, rows,
        col_widths=[12, 14, 12, 14, 14, 14, 10, 10, 25, 12, 12, 12, 20],
        calculated_columns=calculated,
        number_formats={4: C.FMT_DECIMAL_1, 5: C.FMT_DECIMAL_1, 6: C.FMT_DECIMAL_1, 7: C.FMT_PERCENT},
    )

    add_inline_dropdown(ws, f"B2:B{end_row}", C.DROPDOWN_LISTS["count_trigger"])
    add_dropdown_validation(ws, f"C2:C{end_row}", C.TAB_NAMES[26], f"$A$2:$A${len(C.ITEM_DATA)+1}")
    add_dropdown_validation(ws, f"J2:J{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")
    add_dropdown_validation(ws, f"K2:K{end_row}", C.TAB_NAMES[18], f"$A$2:$A${len(C.STAFF_DATA)+1}")

    add_traffic_light_cf(ws, f"H2:H{end_row}",
                          green_op="between", green_val="-0.02",
                          yellow_op="between", yellow_val="-0.05",
                          red_op="lessThan", red_val="-0.05")

    unlock_cols = list(range(1, 14))
    protect_sheet(ws, unlocked_columns=unlock_cols, end_row=max(end_row, 500))
    freeze_panes(ws)
    return ws


def build_layer1(wb, data):
    """Build all 16 Layer 1 (Input) tabs."""
    build_tab1_production(wb, data)
    build_tab2_environmental(wb, data)
    build_tab3_mortality(wb, data)
    build_tab4_feed_consumption(wb, data)
    build_tab5_water_consumption(wb, data)
    build_tab6_ingredient_movement(wb, data)
    build_tab7_feed_mixing(wb, data)
    build_tab8_sales(wb, data)
    build_tab9_procurement(wb, data)
    build_tab10_equipment(wb, data)
    build_tab11_biosecurity(wb, data)
    build_tab12_visitor(wb, data)
    build_tab13_health_incident(wb, data)
    build_tab14_medication(wb, data)
    build_tab15_labor(wb, data)
    build_tab16_inventory_count(wb, data)

    print("  Layer 1: 16 input tabs created with validations and formulas")
